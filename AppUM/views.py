from django.shortcuts import render, get_object_or_404
from .models import *
from datetime import datetime
from django.shortcuts import render
from django.shortcuts import redirect
from django.contrib import messages
from django.contrib.auth.hashers import make_password
from num2words import num2words
from datetime import date
from django.urls import reverse
from django.db.models import Count
from dateutil.relativedelta import relativedelta
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from .decorators import admin_required, mecanico_jefe_required, mecanico_empleado_required
from django.core.mail import send_mail
from .inserts import *
from .functions import *
from django.http import JsonResponse
from django.db.models import Max
from docx.shared import Pt
from docx.oxml.ns import qn
from docx.oxml import OxmlElement
import io
from django.http import HttpResponse
from docx import Document
from django.db.models import Sum
from django.utils.timezone import now
import openpyxl
from openpyxl.styles import Font
from django.http import HttpResponse
from io import BytesIO
from openpyxl.styles import PatternFill, Font, Alignment
# import json
# from docx import Document

# Create your views here.

# for usuario in Personal.objects.all():
#     usuario.contrasena = make_password(usuario.contrasena)
#     usuario.save()

##VISTA DEL LOGIN AL ENTRAR AL SITIO##S
def editar_usuario(req):
    try:
            usuario = req.user
            usuario_actual = Personal.objects.filter(usuario=usuario.username).first()
            if req.method == "POST":
                passw = req.POST.get('password_actual')
                exito = authenticate(username=usuario.username,password=passw)
                if exito:
                    documento = req.POST['tipo_documento'] + str(req.POST['documento'])
                    telefono = str(req.POST['telefono'])
                    correo = req.POST['correo'] + req.POST['dominio_correo']
                    correo = correo.lower()
                    existe_doc = Personal.objects.filter(documento=documento).first()
                    existe_tel = Personal.objects.filter(telefono=telefono).first()
                    existe_cor = Personal.objects.filter(correo=correo).first()

                    if (existe_doc) and (existe_doc.usuario != usuario.username):
                        contexto = contexto_editar_usuario(req,"Ya existe el documento")
                        return render(req,"perfil_administrativo/usuario/editar_usuario.html",contexto)
                    elif (existe_cor) and (existe_cor.usuario != usuario.username):
                        contexto = contexto_editar_usuario(req,"Ya existe el correo")
                        return render(req,"perfil_administrativo/usuario/editar_usuario.html",{contexto})
                    elif (existe_tel) and (existe_tel.usuario != usuario.username):
                        contexto = contexto_editar_usuario(req,"Ya existe el telefono")
                        return render(req,"perfil_administrativo/usuario/editar_usuario.html",contexto)
                    else:
                        f_nac_str = req.POST.get('fecha_nacimiento')  # Cambiado a paréntesis
                        f_nac = datetime.strptime(f_nac_str, '%Y-%m-%d').date() if f_nac_str else None
                        usuario_actual.documento = documento
                        usuario_actual.nombre = req.POST['nombre']
                        usuario_actual.apellido = req.POST['apellido']
                        usuario_actual.fecha_nacimiento = f_nac
                        usuario_actual.correo = correo
                        usuario_actual.telefono = telefono
                        usuario_actual.primera_sesion = 0
                        usuario_actual.save()
                        messages.success(req, "Datos modificados con éxito.")
                        return redirect('EditarUsuario')

                else:
                    contexto = contexto_editar_usuario(req,"La contraseña es incorrecta")
                    return render(req,"perfil_administrativo/usuario/editar_usuario.html",contexto)
            else:
                contexto = contexto_editar_usuario(req,None)
                return render(req,"perfil_administrativo/usuario/editar_usuario.html",contexto)
    except Exception as e:
        contexto = contexto_editar_usuario(req,"Algo salió mal" + str(e))
        return render(req,"perfil_administrativo/usuario/editar_usuario.html",contexto)


def editar_password(req):
    # try:
        usuario = req.user
        passw = req.POST.get('password_actual')
        exito = authenticate(username=usuario.username,password=passw)
        if exito:
            usuario = req.user
            usuario_actual = Personal.objects.filter(usuario=usuario.username).first()
            pass_nueva = req.POST.get('confirm_password')
            usuario_actual.contrasena = make_password(pass_nueva)
            usuario.set_password(pass_nueva)  # Actualiza la contraseña y la encripta automáticamente
            usuario.save()
            usuario_actual.save()
            
            messages.success(req, "Contraseña modificada con éxito, debe iniciar sesión nuevamente.")
            return redirect('Login')
        else:
            contexto = contexto_editar_usuario(req,"La contraseña es incorrecta")
            return render(req,"perfil_administrativo/usuario/editar_usuario.html",contexto)
    # except Exception as e:
    #     pass

def resetear_usuario(req,id_u):
    # try:
        if req.method == "POST":
            usuario = Personal.objects.get(id=id_u)
            usuario.contrasena = make_password("Inicio1234")
            usuario.primera_sesion = 1
            usuario.save()
            user = User.objects.filter(username=usuario.usuario).first()
            user.password = make_password("Inicio1234")
            user.save()
            return render(req,"perfil_administrativo/personal/resetear_usuario.html",{"message":"Usuario reseteado con éxito"})
        else:
            return render(req,"perfil_administrativo/personal/resetear_usuario.html",{})

    # except Exception as e:
    #     pass


def acceso_login(req):
    try:
        if req.method == "POST":
            usuario = req.POST.get('usuario_login')
            passw = req.POST.get('pass_login')

            exito = authenticate(username=usuario,password=passw)
            if exito:
                login(req,exito)
                usuario_consulta = Personal.objects.filter(usuario=usuario).first()
                

                if usuario_consulta:
                    obtener_administrativo = Administrativo.objects.filter(personal_ptr_id=usuario_consulta.id).first()
                    obtener_mecanico = Mecanico.objects.filter(personal_ptr_id=usuario_consulta.id).first()

                    administrativo = obtener_administrativo.activo
                    mecanico_jefe = obtener_mecanico.activo and obtener_mecanico.jefe
                    mecanico_empleado = obtener_mecanico.activo and not obtener_mecanico.jefe

                    if usuario_consulta.primera_sesion:
                        return render(req,"login/cambio_pass.html",{"usuario":usuario})
                    else:
                    
                        if administrativo and mecanico_jefe:
                            renderizar_en = "login/rol_usuario.html"
                            roles = ['Administrativo','Mecanico jefe']
                            contexto = {"roles":roles}
                            # ejecutar_cada_5_min()
                        elif administrativo and mecanico_empleado:
                            renderizar_en = "login/rol_usuario.html"
                            roles = ['Administrativo','Mecanico empleado']
                            contexto = {"roles":roles}
                            # ejecutar_cada_5_min()
                        elif administrativo:
                            renderizar_en = "perfil_administrativo/bienvenida.html"
                            contexto = {}
                            # ejecutar_cada_5_min()
                        elif mecanico_jefe:
                            renderizar_en = "perfil_taller/padre_perfil_taller.html"
                            contexto = {"resultado":"MEC JEFE"}
                            # ejecutar_cada_5_min()
                        elif mecanico_empleado:
                            renderizar_en = "perfil_taller/padre_perfil_taller.html"
                            roles = ['Mecanico empleado']
                            contexto = {"roles":roles}
                            # ejecutar_cada_5_min()
                        else:
                            renderizar_en = "login/login.html"
                            contexto = {"resultado":"Usuario dado de baja, consulte con el administrador del sistema"}
                            # ejecutar_cada_5_min()
                else:
                    #ACCEDER CON SUPER USUARIO
                    renderizar_en = "login/rol_usuario.html"
                    roles = ['Administrativo','Mecanico jefe','Mecanico empleado']
                    contexto = {"roles":roles}

                return render(req,renderizar_en,contexto)
            else:
                return render(req,"login/login.html",{"resultado":"Error usuario y/o pass"})
        else:
            # ejecutar_cada_5_min()
            # ejecutar_hilo()
            return render(req,"login/login.html",{})
    except Exception as e:
        return render(req,"login/login.html",{"resultado":e})

@login_required()
def cerrar_sesion(req):
    try:
        logout(req)
        messages.success(req, "Has cerrado sesión correctamente.")  # Agrega el mensaje
        return redirect('Login')
    except:
        pass

@login_required()
def cambio_pass(req):
    try:
        pass1 = req.POST.get('nueva_contrasena')
        pass2 = req.POST.get('nueva_contrasena_2')
       
        if (len(pass1) < 8):
            return render(req,"login/cambio_pass.html",{"resultado":"La contraseña debe tener un mínimo de 8 caracteres"})
        elif pass1 != pass2:
            return render(req,"login/cambio_pass.html",{"resultado":"Las contraseñas deben coincidir"})
        else:
            #usuario.contrasena = make_password(usuario.contrasena)
            usuario = req.user
            personal = Personal.objects.filter(usuario=usuario.username).first()
            personal.primera_sesion = 0
            personal.contrasena = make_password(pass1)
            personal.save()

            usuario.password = make_password(pass1)
            usuario.save()
            messages.success(req, "Contraseña cambiada correctamente.")  # Agrega el mensaje
            return redirect('Login')
            
    except Exception as e:
        pass

@login_required()
def seleccion_rol(req):
    # try:
        rol = req.POST['rol_usuario']
        if rol == "Administrativo":
            renderizar_en = "perfil_administrativo/bienvenida.html"
            contexto={}
        elif rol == "Mecanico jefe":
            renderizar_en = "perfil_taller/padre_perfil_taller.html"
            contexto={}
        else:
            renderizar_en = "perfil_taller/padre_perfil_taller.html"
            contexto={}

        return render(req,renderizar_en,contexto)
    # except Exception as e:
    #     return render(req,"login/login.html",{"resultado":e})


@admin_required
def vista_inventario_motos(req):
    motos = Moto.objects.filter(pertenece_tienda=1).order_by('-fecha_ingreso')
    data_motos = []
    for moto in motos:
        reservada = ComprasVentas.objects.filter(tipo="R",moto_id=moto.id).first()
        data_motos.append({
            "moto":moto,
            "estado":"Reservada" if reservada else "En stock"
        })
    logo_um = Logos.objects.get(id=1)
    page_obj = funcion_paginas_varias(req,data_motos)
    return render(req,"perfil_administrativo/motos/motos.html",{'page_obj': page_obj,"motos":motos,"logo_um":logo_um.logo_UM.url if logo_um.logo_UM else None,"active_page": 'Motos'})

@admin_required
def cliente_moto(req):
    # try:
        num_motor = req.POST['num_motor_moto'].upper()
        documento = req.POST['tipo_documento'] + str(req.POST['documento'])

        moto = Moto.objects.filter(num_motor=num_motor).first()
        cliente = Cliente.objects.filter(documento=documento).first()
        if cliente:
            telefono = ClienteTelefono.objects.filter(cliente_id=cliente.id,principal=1).first()
            correo = ClienteCorreo.objects.filter(cliente_id=cliente.id,principal=1).first()

        if not cliente:
            return render(req,"perfil_administrativo/motos/alta_moto.html",{"error_message_cliente":"El cliente no existe, para ingresarlo haga clic ",
                                                                            "active_page": 'Motos',
                                                                            "form_moto_usada":False,
                                                                            "form_moto_ingresada":False})
        elif not moto:
            #SI LA MOTO NO EXISTE, SE INGRESA LA MOTO DESDE 0
            return render(req,"perfil_administrativo/motos/alta_moto.html",{"cliente":cliente,
                                                                            "telefono":telefono.telefono,
                                                                            "correo":correo.correo if correo else "El cliente no cuenta con correo",
                                                                            "active_page": 'Motos',
                                                                            "form_moto_usada":True,
                                                                            "form_moto_ingresada":False,
                                                                            "consultar_moto_cliente":False,
                                                                            "error_message":"La moto no se encuentra registrada en el sistema, debe ingresar la misma completando todos los datos requeridos"})
        else:
            #SI LA MOTO Y EL CLIENTE EXISTEN, ENTONCES SE MODIFICAN CAMPOS EN MOT (PERT_TIENDA = 1)
            if moto.pertenece_tienda:
                return render(req,"perfil_administrativo/motos/alta_moto.html",{"error_message":"La moto que usted desea ingresar ya existe en el inventario de la tienda",
                                                                                "active_page": 'Motos',
                                                                                "telefono":telefono.telefono,
                                                                                 "correo":correo.correo if correo else "El cliente no cuenta con correo",
                                                                                "form_moto_usada":False,
                                                                                "form_moto_ingresada":False})
            else:
                existe_matricula = Matriculas.objects.filter(moto_id=moto.id).first()
                if existe_matricula:
                    matricula = existe_matricula.matricula
                else:
                    matricula = "Sin matrícula"
                return render(req,"perfil_administrativo/motos/alta_moto.html",{"cliente":cliente,
                                                                                "moto":moto,
                                                                                "matricula":matricula,
                                                                                "telefono":telefono.telefono,
                                                                                "correo":correo.correo if correo else "El cliente no cuenta con correo",
                                                                                "active_page": 'Motos',
                                                                                "form_moto_usada":False,
                                                                                "form_moto_ingresada":True,
                                                                                "consultar_moto_cliente":False})
    # except Exception as e:
    #     return render(req,"perfil_administrativo/motos/alta_moto.html",{"error_message":e})

@admin_required
def alta_moto_usada(req,id_cliente):
    # try:
    #INSERT EN MOTO
        cliente = Cliente.objects.get(id=id_cliente)
        num_chasis = req.POST['num_chasis_moto'].upper()
        num_motor = req.POST['num_motor_moto'].upper()
        existe_num_motor = Moto.objects.filter(num_motor=num_motor).first()
        existe_num_chasis = Moto.objects.filter(num_chasis=num_chasis).first()
        matricula = req.POST['matricula_letras'].upper() + str(req.POST['matricula_numeros'].upper())
        padron = req.POST['num_padron']
        existe_matricula = Matriculas.objects.filter(matricula=matricula).first()
        existe_padron = Matriculas.objects.filter(padron=padron).first()
        departamento = departamento_matricula(matricula)
        if existe_num_motor:
            return render(req,"perfil_administrativo/motos/alta_moto.html",{"cliente":cliente,
                                                                            "active_page": 'Motos',
                                                                            "form_moto_usada":True,
                                                                            "form_moto_ingresada":False,
                                                                            "consultar_moto_cliente":False,
                                                                            "error_message":"Ya existe el número de motor ingresado"})
        elif existe_num_chasis:
            return render(req,"perfil_administrativo/motos/alta_moto.html",{"cliente":cliente,
                                                                            "active_page": 'Motos',
                                                                            "form_moto_usada":True,
                                                                            "form_moto_ingresada":False,
                                                                            "consultar_moto_cliente":False,
                                                                            "error_message":"Ya existe el número de chasis ingresado"})
        elif existe_matricula:
            return render(req,"perfil_administrativo/motos/alta_moto.html",{"cliente":cliente,
                                                                            "active_page": 'Motos',
                                                                            "form_moto_usada":True,
                                                                            "form_moto_ingresada":False,
                                                                            "consultar_moto_cliente":False,
                                                                            "error_message":"Ya existe la matricula ingresada"})
        elif existe_padron:
            return render(req,"perfil_administrativo/motos/alta_moto.html",{"cliente":cliente,
                                                                            "active_page": 'Motos',
                                                                            "form_moto_usada":True,
                                                                            "form_moto_ingresada":False,
                                                                            "consultar_moto_cliente":False,
                                                                            "error_message":"Ya existe el número de padrón ingresado"})
        elif not departamento:
            return render(req,"perfil_administrativo/motos/alta_moto.html",{"cliente":cliente,
                                                                            "active_page": 'Motos',
                                                                            "form_moto_usada":True,
                                                                            "form_moto_ingresada":False,
                                                                            "consultar_moto_cliente":False,
                                                                            "error_message":"La matrícula es incorrecta"})
        else:
            marca = req.POST['marca_moto'].upper()
            modelo = req.POST['modelo_moto'].upper()
            color = req.POST['color_moto'].upper()

            foto = req.FILES.get('foto_moto')

            checkbox_num_motor = 'sin_num_motor' in req.POST
            if checkbox_num_motor:
                numero_de_motor = crear_num_motor()
                tiene_numero_motor = 0
            else:
                numero_de_motor = num_motor
                tiene_numero_motor = 1


            checkbox_num_chasis = 'sin_num_chasis' in req.POST
            if checkbox_num_chasis:
                numero_de_chasis = crear_num_chasis()
                tiene_numero_chasis = 0
            else:
                numero_de_chasis = num_chasis
                tiene_numero_chasis = 1
            
            nueva_moto = insert_moto(marca,modelo,req.POST['anio_moto'],"Usada",req.POST['motor_moto'],req.POST['km_moto'],req.POST['moneda_precio'],req.POST['precio_moto'],color,numero_de_motor,numero_de_chasis,req.POST['num_cilindros'],req.POST['num_pasajeros'],1,0,req.POST['descripcion_moto'],foto,req.POST['tipo_moto'],tiene_numero_motor,tiene_numero_chasis)
            libreta_propiedad = req.FILES.get('libreta_propiedad_moto')
           
            insert_compras_ventas("CV",libreta_propiedad,id_cliente,nueva_moto.id,None,None,None)

            # generar_compromiso_compra_venta_moto_ingreso(req,nueva_moto.id,id_cliente)


            insert_matriculas(matricula,padron,nueva_moto.id)
            checkbox = 'crear_pdf' in req.POST
            if checkbox:
                
                pdf = checkbox_pdf(req,nueva_moto)
                return pdf
            else:
                # messages.success(req, "Moto ingresada con éxito.")
                # return redirect('Motos')
                messages.success(req, "Moto ingresada con éxito, debe registrar el compromiso compraventa para que la misma se ingrese en el inventario.")
                return redirect(reverse('ClienteFicha', kwargs={'id_cliente': id_cliente}))

            
    # except Exception as e:
    #          return render(req,"perfil_administrativo/motos/alta_moto.html",{"cliente":cliente,
    #                                                                          "active_page": 'Motos',
    #                                                                          "form_moto_usada":True,
    #                                                                          "form_moto_ingresada":False,
    #                                                                          "consultar_moto_cliente":False,
    #                                                                          "error_message":e})

@admin_required
def reingresar_moto_usada(req,id_moto,id_cliente):
    try:
    #UPDATE MOTO PERT_TIENDA = 1
        moto = Moto.objects.get(id=id_moto)
        
        moto.kilometros = req.POST['km_moto']
        moto.moneda_precio = req.POST['moneda_precio']
        moto.precio = req.POST['precio_moto']
        moto.observaciones = req.POST['descripcion_moto']
        moto.foto = req.FILES.get('foto_moto')
        moto.pertenece_tienda = 1
        moto.fecha_ingreso = datetime.now()
        moto.save()

        libreta_propiedad = req.FILES.get('libreta_propiedad_moto')
        insert_compras_ventas("CV",libreta_propiedad,id_cliente,moto.id,None,None,None)
        checkbox = 'crear_pdf' in req.POST
        if checkbox:
            pdf = checkbox_pdf(req,moto)
            return pdf
        else:
            messages.success(req, "Moto ingresada con éxito.")
            return redirect('Motos')

    except Exception as e:
        return render(req,"perfil_administrativo/motos/alta_moto.html",{"active_page": 'Motos',
                                                                            "form_moto_usada":True,
                                                                            "form_moto_ingresada":False,
                                                                            "consultar_moto_cliente":False,
                                                                            "error_message":e})

@admin_required
def alta_moto_nueva(req):
    try:
    #INSERT EN MOTO
        checkbox_num_motor = 'sin_num_motor' in req.POST
        if checkbox_num_motor:
            num_motor = crear_num_motor()
            contiene_num_motor = 0
        else:
            num_motor = req.POST['num_motor_moto'].upper()
            contiene_num_motor = 1

        checkbox_num_chasis = 'sin_num_chasis' in req.POST
        if checkbox_num_chasis:
            num_chasis = crear_num_chasis()
            contiene_num_chasis = 0
        else:
            num_chasis = req.POST['num_chasis_moto'].upper()
            contiene_num_chasis = 1
            

        existe_num_motor = Moto.objects.filter(num_motor=num_motor).first()
        existe_num_chasis = Moto.objects.filter(num_chasis=num_chasis).first()
        if existe_num_motor:
            return render(req,"perfil_administrativo/motos/alta_moto.html",{
                                                                            "active_page": 'Motos',
                                                                            "form_moto_usada":False,
                                                                            "form_moto_ingresada":False,
                                                                            "consultar_moto_cliente":False,
                                                                            "error_message":"Ya existe el número de motor ingresado",
                                                                            "es_nueva":True})
        elif existe_num_chasis:
            return render(req,"perfil_administrativo/motos/alta_moto.html",{
                                                                            "active_page": 'Motos',
                                                                            "form_moto_usada":False,
                                                                            "form_moto_ingresada":False,
                                                                            "consultar_moto_cliente":False,
                                                                            "error_message":"Ya existe el número de chasis ingresado",
                                                                            "es_nueva":True})
        else:
            marca = req.POST['marca_moto'].upper()
            modelo = req.POST['modelo_moto'].upper()
            color = req.POST['color_moto'].upper()

            cbox_precio = 'cbox_precio_moto' in req.POST
            if cbox_precio:
                precio = 0
                moneda = "Sin precio"
            else:
                precio = req.POST['precio_moto']
                moneda = req.POST['moneda_precio']

            foto = req.FILES.get('foto_moto')
            nueva_moto = insert_moto(marca,modelo,req.POST['anio_moto'],"Nueva",req.POST['motor_moto'],0,moneda,precio,color,num_motor,num_chasis,req.POST['num_cilindros'],req.POST['num_pasajeros'],1,0,req.POST['descripcion_moto'],foto,req.POST['tipo_moto'],contiene_num_motor,contiene_num_chasis)

            checkbox = 'crear_pdf' in req.POST
            if checkbox:
                pdf = checkbox_pdf(req,nueva_moto)
                return pdf
            else:
                messages.success(req, "Moto ingresada con éxito.")
                return redirect('Motos')

            
    except Exception as e:
            return render(req,"perfil_administrativo/motos/alta_moto.html",{"active_page": 'Motos',
                                                                            "form_moto_usada":True,
                                                                            "form_moto_ingresada":False,
                                                                            "consultar_moto_cliente":False,
                                                                            "error_message":e,
                                                                            "es_nueva":True})    

def alta_moto_usada_sin_dueno(req):                
    # try:
    #INSERT EN MOTO
        
        num_chasis = req.POST['num_chasis_moto'].upper()
        num_motor = req.POST['num_motor_moto'].upper()
        existe_num_motor = Moto.objects.filter(num_motor=num_motor).first()
        existe_num_chasis = Moto.objects.filter(num_chasis=num_chasis).first()
        matricula = req.POST['matricula_letras'].upper() + str(req.POST['matricula_numeros'].upper())
        checkbox_padron = 'cb_sin_padron_sn' in req.POST
        if checkbox_padron:
            padron = None
            existe_padron = None
        else:
            padron = req.POST['num_padron']
            existe_padron = Matriculas.objects.filter(padron=padron).first()
        existe_matricula = Matriculas.objects.filter(matricula=matricula).first()
        departamento = departamento_matricula(matricula)
        if existe_num_motor:
            return render(req,"perfil_administrativo/motos/alta_moto.html",{
                                                                            "active_page": 'Motos',
                                    
                                                                            "error_message":"Ya existe el número de motor ingresado"})
        elif existe_num_chasis:
            return render(req,"perfil_administrativo/motos/alta_moto.html",{
                                                                            "active_page": 'Motos',
                                                                            
                                                                            "error_message":"Ya existe el número de chasis ingresado"})
        elif existe_matricula:
            return render(req,"perfil_administrativo/motos/alta_moto.html",{
                                                                            "active_page": 'Motos',
                                                                            
                                                                            "error_message":"Ya existe la matricula ingresada"})
        elif existe_padron:
            return render(req,"perfil_administrativo/motos/alta_moto.html",{
                                                                            "active_page": 'Motos',
                                                                            
                                                                            "error_message":"Ya existe el número de padrón ingresado"})
        elif not departamento:
            return render(req,"perfil_administrativo/motos/alta_moto.html",{
                                                                            "active_page": 'Motos',
                                                                            
                                                                            "error_message":"La matrícula es incorrecta"})
        else:
            marca = req.POST['marca_moto'].upper()
            modelo = req.POST['modelo_moto'].upper()
            color = req.POST['color_moto'].upper()

            foto = req.FILES.get('foto_moto')

            checkbox_num_motor = 'sin_num_motor' in req.POST
            if checkbox_num_motor:
                numero_de_motor = crear_num_motor()
                tiene_numero_motor = 0
            else:
                numero_de_motor = num_motor
                tiene_numero_motor = 1


            checkbox_num_chasis = 'sin_num_chasis' in req.POST
            if checkbox_num_chasis:
                numero_de_chasis = crear_num_chasis()
                tiene_numero_chasis = 0
            else:
                numero_de_chasis = num_chasis
                tiene_numero_chasis = 1
            
            nueva_moto = insert_moto(marca,modelo,req.POST['anio_moto'],"Usada",req.POST['motor_moto'],req.POST['km_moto'],req.POST['moneda_precio'],req.POST['precio_moto'],color,numero_de_motor,numero_de_chasis,req.POST['num_cilindros'],req.POST['num_pasajeros'],1,0,req.POST['descripcion_moto'],foto,req.POST['tipo_moto'],tiene_numero_motor,tiene_numero_chasis)
            #libreta_propiedad = req.FILES.get('libreta_propiedad_moto')
           
            # insert_compras_ventas("CV",libreta_propiedad,id_cliente,nueva_moto.id,None,None,None)


            insert_matriculas(matricula,padron,nueva_moto.id)
            checkbox = 'crear_pdf' in req.POST
            if checkbox:
                
                pdf = checkbox_pdf(req,nueva_moto)
                return pdf
            else:
                messages.success(req, "Moto ingresada con éxito.")
                return redirect('Motos')

            
    # except Exception as e:
    #          return render(req,"perfil_administrativo/motos/alta_moto.html",{"cliente":cliente,
    #                                                                          "active_page": 'Motos',
    #                                                                          "form_moto_usada":True,
    #                                                                          "form_moto_ingresada":False,
    #                                                                          "consultar_moto_cliente":False,
    #                                                                          "error_message":e})



@admin_required
def form_alta_moto(req):
    return render(req,"perfil_administrativo/motos/alta_moto.html",{"active_page": 'Motos',
                                                                            "consultar_moto_cliente":True})


@admin_required
def baja_moto(req,id_moto):
    if req.method == 'POST':

      motoDel = Moto.objects.get(id=id_moto)
      motoDel.pertenece_tienda = 0
      motoDel.save()
      return render(req, "perfil_administrativo/motos/baja_moto.html", {"message":"Moto borrada con éxito","id_moto":0,"active_page": 'Motos'})
    else:
       return render(req, "perfil_administrativo/motos/baja_moto.html", {"id_moto":id_moto,"active_page": 'Motos'})

@admin_required
def busqueda_codigo(req):
    codigo = req.GET.get('codigo')
    # motos = Moto.objects.filter(id=codigo,pertenece_tienda=1)
    # page_obj = funcion_paginas_varias(req,motos)
    # return render(req,"perfil_administrativo/motos/motos.html",{'page_obj': page_obj,"motos":motos,"active_page": 'Motos'})
    motos = Moto.objects.filter(id=codigo,pertenece_tienda=1).order_by('-fecha_ingreso')
    data_motos = []
    for moto in motos:
        reservada = ComprasVentas.objects.filter(tipo="R",moto_id=moto.id).first()
        data_motos.append({
            "moto":moto,
            "estado":"Reservada" if reservada else "En stock"
        })
    logo_um = Logos.objects.get(id=1)
    page_obj = funcion_paginas_varias(req,data_motos)
    return render(req,"perfil_administrativo/motos/motos.html",{'page_obj': page_obj,"motos":motos,"logo_um":logo_um.logo_UM.url if logo_um.logo_UM else None,"active_page": 'Motos'})

@admin_required
def busqueda_marca(req):
    marca = req.GET.get('marca')
    # motos = Moto.objects.filter(marca__icontains=marca,pertenece_tienda=1).order_by('-fecha_ingreso')
    # page_obj = funcion_paginas_varias(req,motos)
    # return render(req,"perfil_administrativo/motos/motos.html",{'page_obj': page_obj,"motos":motos,"active_page": 'Motos'})
    motos = Moto.objects.filter(marca__icontains=marca,pertenece_tienda=1).order_by('-fecha_ingreso')
    data_motos = []
    for moto in motos:
        reservada = ComprasVentas.objects.filter(tipo="R",moto_id=moto.id).first()
        data_motos.append({
            "moto":moto,
            "estado":"Reservada" if reservada else "En stock"
        })
    logo_um = Logos.objects.get(id=1)
    page_obj = funcion_paginas_varias(req,data_motos)
    return render(req,"perfil_administrativo/motos/motos.html",{'page_obj': page_obj,"motos":motos,"logo_um":logo_um.logo_UM.url if logo_um.logo_UM else None,"active_page": 'Motos'})

@admin_required
def busqueda_modelo(req):
    modelo = req.GET.get('modelo')
    motos = Moto.objects.filter(modelo__icontains=modelo,pertenece_tienda=1).order_by('-fecha_ingreso')
    data_motos = []
    for moto in motos:
        reservada = ComprasVentas.objects.filter(tipo="R",moto_id=moto.id).first()
        data_motos.append({
            "moto":moto,
            "estado":"Reservada" if reservada else "En stock"
        })
    logo_um = Logos.objects.get(id=1)
    page_obj = funcion_paginas_varias(req,data_motos)
    return render(req,"perfil_administrativo/motos/motos.html",{'page_obj': page_obj,"motos":motos,"logo_um":logo_um.logo_UM.url if logo_um.logo_UM else None,"active_page": 'Motos'})


@admin_required
def busqueda_marca_modelo(req):
    marca = req.GET.get('marca_modelo')
    modelo = req.GET.get('modelo_marca')
    motos = Moto.objects.filter(marca__icontains=marca,modelo__icontains=modelo,pertenece_tienda=1).order_by('-fecha_ingreso')
    data_motos = []
    for moto in motos:
        reservada = ComprasVentas.objects.filter(tipo="R",moto_id=moto.id).first()
        data_motos.append({
            "moto":moto,
            "estado":"Reservada" if reservada else "En stock"
        })
    logo_um = Logos.objects.get(id=1)
    page_obj = funcion_paginas_varias(req,data_motos)
    return render(req,"perfil_administrativo/motos/motos.html",{'page_obj': page_obj,"motos":motos,"logo_um":logo_um.logo_UM.url if logo_um.logo_UM else None,"active_page": 'Motos'})

@admin_required
def busqueda_anio(req):
    anio = req.GET.get('anio')
    motos = Moto.objects.filter(anio=anio,pertenece_tienda=1).order_by('-fecha_ingreso')
    data_motos = []
    for moto in motos:
        reservada = ComprasVentas.objects.filter(tipo="R",moto_id=moto.id).first()
        data_motos.append({
            "moto":moto,
            "estado":"Reservada" if reservada else "En stock"
        })
    logo_um = Logos.objects.get(id=1)
    page_obj = funcion_paginas_varias(req,data_motos)
    return render(req,"perfil_administrativo/motos/motos.html",{'page_obj': page_obj,"motos":motos,"logo_um":logo_um.logo_UM.url if logo_um.logo_UM else None,"active_page": 'Motos'})

@admin_required
def busqueda_kms(req):
    km_minimo = req.GET.get('km_minimo')
    km_maximo = req.GET.get('km_maximo')
    motos = Moto.objects.filter(kilometros__range=(km_minimo, km_maximo),pertenece_tienda=1).order_by('-fecha_ingreso')
    data_motos = []
    for moto in motos:
        reservada = ComprasVentas.objects.filter(tipo="R",moto_id=moto.id).first()
        data_motos.append({
            "moto":moto,
            "estado":"Reservada" if reservada else "En stock"
        })
    logo_um = Logos.objects.get(id=1)
    page_obj = funcion_paginas_varias(req,data_motos)
    return render(req,"perfil_administrativo/motos/motos.html",{'page_obj': page_obj,"motos":motos,"logo_um":logo_um.logo_UM.url if logo_um.logo_UM else None,"active_page": 'Motos'})


@admin_required
def busqueda_precio(req):
    precio_minimo = req.GET.get('precio_minimo')
    precio_maximo = req.GET.get('precio_maximo')
    motos = Moto.objects.filter(precio__range=(precio_minimo, precio_maximo),pertenece_tienda=1).order_by('-fecha_ingreso')
    data_motos = []
    for moto in motos:
        reservada = ComprasVentas.objects.filter(tipo="R",moto_id=moto.id).first()
        data_motos.append({
            "moto":moto,
            "estado":"Reservada" if reservada else "En stock"
        })
    logo_um = Logos.objects.get(id=1)
    page_obj = funcion_paginas_varias(req,data_motos)
    return render(req,"perfil_administrativo/motos/motos.html",{'page_obj': page_obj,"motos":motos,"logo_um":logo_um.logo_UM.url if logo_um.logo_UM else None,"active_page": 'Motos'})


@admin_required
def busqueda_matricula(req):
    matricula_letras = req.GET.get('letras_matricula').upper()
    matricula_numeros = str(req.GET.get('numeros_matricula'))
    matricula = matricula_letras + matricula_numeros

    busq_matricula = Matriculas.objects.filter(matricula = matricula).first()
    if busq_matricula:
        motos = Moto.objects.filter(id=busq_matricula.moto_id,pertenece_tienda=1)
        # motos = Moto.objects.filter(precio__range=(precio_minimo, precio_maximo),pertenece_tienda=1).order_by('-fecha_ingreso')
        data_motos = []
        for moto in motos:
            reservada = ComprasVentas.objects.filter(tipo="R",moto_id=moto.id).first()
            data_motos.append({
                "moto":moto,
                "estado":"Reservada" if reservada else "En stock"
            })
        logo_um = Logos.objects.get(id=1)
        page_obj = funcion_paginas_varias(req,data_motos)
        # return render(req,"perfil_administrativo/motos/motos.html",{'page_obj': page_obj,"motos":motos,"logo_um":logo_um.logo_UM.url if logo_um.logo_UM else None,"active_page": 'Motos'})

        # page_obj = funcion_paginas_varias(req,moto)
        contexto = {'page_obj': page_obj,"motos":moto,"active_page": 'Motos'}
    else:
        contexto = {'page_obj': None,"motos":None,"active_page": 'Motos'}

    return render(req,"perfil_administrativo/motos/motos.html",contexto)

@admin_required
def busqueda_tipo_moto(req):
    tipo = req.GET.get('tipo_moto')
    motos = Moto.objects.filter(tipo=tipo,pertenece_tienda=1).order_by('-fecha_ingreso')
    data_motos = []
    for moto in motos:
        reservada = ComprasVentas.objects.filter(tipo="R",moto_id=moto.id).first()
        data_motos.append({
            "moto":moto,
            "estado":"Reservada" if reservada else "En stock"
        })
    logo_um = Logos.objects.get(id=1)
    page_obj = funcion_paginas_varias(req,data_motos)
    return render(req,"perfil_administrativo/motos/motos.html",{'page_obj': page_obj,"motos":motos,"logo_um":logo_um.logo_UM.url if logo_um.logo_UM else None,"active_page": 'Motos'})


@admin_required
def busqueda_num_motor(req):
    num_motor = req.GET.get('num_motor').upper()
    motos = Moto.objects.filter(num_motor=num_motor)
    data_motos = []
    for moto in motos:
        reservada = ComprasVentas.objects.filter(tipo="R",moto_id=moto.id).first()
        data_motos.append({
            "moto":moto,
            "estado":"Reservada" if reservada else "En stock"
        })
    logo_um = Logos.objects.get(id=1)
    page_obj = funcion_paginas_varias(req,data_motos)
    return render(req,"perfil_administrativo/motos/motos.html",{'page_obj': page_obj,"motos":motos,"logo_um":logo_um.logo_UM.url if logo_um.logo_UM else None,"active_page": 'Motos'})


@admin_required
def busqueda_num_chasis(req):
    num_chasis = req.GET.get('num_chasis').upper()
    motos = Moto.objects.filter(num_chasis=num_chasis)
    data_motos = []
    for moto in motos:
        reservada = ComprasVentas.objects.filter(tipo="R",moto_id=moto.id).first()
        data_motos.append({
            "moto":moto,
            "estado":"Reservada" if reservada else "En stock"
        })
    logo_um = Logos.objects.get(id=1)
    page_obj = funcion_paginas_varias(req,data_motos)
    return render(req,"perfil_administrativo/motos/motos.html",{'page_obj': page_obj,"motos":motos,"logo_um":logo_um.logo_UM.url if logo_um.logo_UM else None,"active_page": 'Motos'})




    
@admin_required
def modificacion_moto(req,id_moto):
    try:
        if req.method == "POST":
                
                moto_upd = Moto.objects.get(id=id_moto)
                checkbox_num_motor = 'con_num_motor' in req.POST
                if moto_upd.contiene_num_motor:
                    num_motor = req.POST['num_motor_moto'].upper()
                else:
                    if checkbox_num_motor:
                        num_motor = req.POST['num_motor_moto_agregado'].upper()
                    else:
                        num_motor = None
                
                checkbox_num_chasis = 'con_num_chasis' in req.POST
                if moto_upd.contiene_num_chasis:
                    num_chasis = req.POST['num_chasis_moto'].upper()
                else:
                    if checkbox_num_chasis:
                        num_chasis = req.POST['num_chasis_moto_agregado'].upper()
                    else:
                        num_chasis = None
                # num_chasis = req.POST['num_chasis_moto'].upper()
                matricula = req.POST['matricula_letras'].upper() + str(req.POST['matricula_numeros'])
                if req.POST['matricula_letras'] and req.POST['matricula_numeros']:
                    letras_matricula = req.POST['matricula_letras']
                    num_matricula = req.POST['matricula_numeros']
                    validar_letra_matricula = departamento_matricula(matricula)
                else:
                    letras_matricula = None
                    num_matricula = None
                    validar_letra_matricula = "No Ingresada"
                padron = req.POST['num_padron']
                if not padron:
                    padron = None
                validar_matricula = matricula_valid(matricula,padron,id_moto)
                validacion_datos_moto = validacion_moto(id_moto,num_motor,num_chasis)
                precio = int(moto_upd.precio) if moto_upd.precio else 0

                if validacion_datos_moto == "existe_num_motor":
                    return render(req,"perfil_administrativo/motos/modificacion_moto.html",{"error_message":"El número de motor ya se encuentra asignado a otra moto",
                                                                                            'datos_moto': moto_upd,
                                                                                            "letras_matricula":letras_matricula,
                                                                                            "num_matricula":num_matricula,
                                                                                            "precio":precio,
                                                                                            "active_page": 'Motos'}) 
                elif validacion_datos_moto == "existe_num_chasis":
                    return render(req,"perfil_administrativo/motos/modificacion_moto.html",{"error_message":"El número de chasis ya se encuentra asignado a otra moto",
                                                                                            'datos_moto': moto_upd,
                                                                                            "letras_matricula":letras_matricula,
                                                                                            "num_matricula":num_matricula,
                                                                                            "precio":precio,
                                                                                            "active_page": 'Motos'}) 
                elif (req.POST['matricula_letras'] and req.POST['matricula_numeros']) and (validar_matricula == "matricula_existe"): #EXISTE UNA MATRICULA REGISTRADA Y ADEMAS ES DIFERENTE
                    return render(req,"perfil_administrativo/motos/modificacion_moto.html",{"error_message":"La matricula ingresada ya existe",
                                                                                            'datos_moto': moto_upd,
                                                                                            "letras_matricula":letras_matricula,
                                                                                            "num_matricula":num_matricula,
                                                                                            "precio":precio,
                                                                                            "active_page": 'Motos'}) 
                elif (req.POST['num_padron']) and (validar_matricula == "padron_existe"):
                    return render(req,"perfil_administrativo/motos/modificacion_moto.html",{"error_message":"El número de padrón ya existe",
                                                                                            'datos_moto': moto_upd,
                                                                                            "letras_matricula":letras_matricula,
                                                                                            "num_matricula":num_matricula,
                                                                                            "precio":precio,
                                                                                            "active_page": 'Motos'})
                elif not validar_letra_matricula:
                    return render(req,"perfil_administrativo/motos/modificacion_moto.html",{"error_message":"La matrícula es incorrecta",
                                                                                            'datos_moto': moto_upd,
                                                                                            "letras_matricula":letras_matricula,
                                                                                            "num_matricula":num_matricula,
                                                                                            "precio":precio,
                                                                                            "active_page": 'Motos'})
                else:
                            
                    marca = req.POST['marca_moto'].upper()
                    modelo = req.POST['modelo_moto'].upper()
                    color = req.POST['color_moto'].upper()

                    foto = req.FILES.get('foto_moto')
                    
                    moto_upd = Moto.objects.get(id=id_moto)
                    moto_upd.marca = marca
                    moto_upd.modelo = modelo
                    moto_upd.motor = req.POST['motor_moto']
                    
                    
                    if moto_upd.contiene_num_motor or checkbox_num_motor:
                        moto_upd.num_motor = num_motor
                        moto_upd.contiene_num_motor = 1
                    
                    if moto_upd.contiene_num_chasis or checkbox_num_chasis:
                        moto_upd.num_chasis = num_chasis
                        moto_upd.contiene_num_chasis = 1
                
                    moto_upd.anio = req.POST['anio_moto']
                    moto_upd.tipo = req.POST['tipo_moto']

                    if req.POST['estado_moto'] == "Nueva":
                        estado = "Nueva"
                        moto_upd.kilometros = 0
                    else:
                        estado = "Usada"
                        moto_upd.kilometros = req.POST['km_moto']
                        
                    moto_upd.estado = estado
                    
                    
                    # moto_upd.num_chasis = num_chasis
                    moto_upd.num_cilindros = req.POST['num_cilindros']
                    moto_upd.cantidad_pasajeros = req.POST['num_pasajeros']
                    moto_upd.color = color
                    moto_upd.moneda_precio = req.POST['moneda_precio']
                    moto_upd.precio = req.POST['precio_moto']
                    moto_upd.observaciones = req.POST['descripcion_moto']
                    moto_upd.foto = foto
                    moto_upd.save()

                    matricula_actual = Matriculas.objects.filter(moto_id=id_moto).first()
                    if req.POST['matricula_letras'] and req.POST['matricula_numeros'] and req.POST['num_padron']:
                        if matricula_actual:
                            if matricula_actual.matricula != matricula: #EN CASO DE QUE LA MATRICULA INGRESADA EN EL TEMPLATE SEA DIFERENTE A LA EXISTENTE EN LA BASE DE DATOS, ENTONCES ACCEDE
                                matricula_actual.delete()
                                insert_matriculas(matricula,padron,moto_upd.id)
                        else:
                                insert_matriculas(matricula,padron,moto_upd.id)


                    checkbox = 'crear_pdf' in req.POST
                    if checkbox:
                            pdf = checkbox_pdf(req,moto_upd)
                            return pdf
                    else:
                        messages.success(req, "La moto ha sido modificada con éxito.")
                        return redirect('Motos')
        else:
            moto_upd = Moto.objects.get(id=id_moto)
            matricula_upd = Matriculas.objects.filter(moto_id=moto_upd.id).first()

            if matricula_upd:
                letras_matricula = matricula_upd.matricula[0:3:1]
                num_matricula = matricula_upd.matricula[3:7:1]
                padron = matricula_upd.padron
            else:
                letras_matricula = None
                num_matricula = None
                padron = None

            if moto_upd.observaciones == None:
                moto_upd.observaciones = "Sin descripción"
            precio = int(moto_upd.precio) if moto_upd.precio else 0
            return render(req,"perfil_administrativo/motos/modificacion_moto.html",{'datos_moto': moto_upd,
            "letras_matricula":letras_matricula,
            "num_matricula":num_matricula,
            "padron":padron,
            "precio":precio,
            "active_page": 'Motos'}) 
    except Exception as e:
        return render(req,"perfil_administrativo/motos/modificacion_moto.html",{'datos_moto': moto_upd,
            "letras_matricula":letras_matricula,
            "num_matricula":num_matricula,"active_page": 'Motos',"error_message":str(e)}) 

def form_cambio_propietario_moto(req,id_moto):
    try:
        if req.method == "POST":
            moto = Moto.objects.get(id=id_moto)
            documento = req.POST['tipo_documento'] + str(req.POST['documento'])
            print("DOCUMENTO ES: " + documento)
            cliente = Cliente.objects.filter(documento=documento).first()
            if cliente:
                return render(req,"perfil_administrativo/motos/cambio_duenio.html",{"id_moto":id_moto,
                                                                                    "cliente":cliente,
                                                                                    "moto":moto})
            else:
                return render(req,"perfil_administrativo/motos/cambio_duenio.html",{"id_moto":id_moto,
                                                                                    "cliente":None,
                                                                                    "moto":None,
                                                                                    "error_message_cliente":"El cliente no existe, para ingresarlo haga clic "})
        else:
            return render(req,"perfil_administrativo/motos/cambio_duenio.html",{"id_moto":id_moto})
    except Exception as e:
        pass

def cambio_duenio_moto(req,id_moto,id_cliente):
    try:
        nuevo_duenio = ComprasVentas(
            fecha_compra = datetime.now(),
            tipo = "CV",
            cliente_id = id_cliente,
            moto_id = id_moto
        )
        nuevo_duenio.save()
        messages.success(req, "Propietario actual modificado con éxito.")
        return redirect('Motos')
    except Exception as e:
        pass
    

@admin_required
def detalles_moto(req,id_moto):
    moto = Moto.objects.get(id=id_moto)
    
    matriculas_moto = Matriculas.objects.filter(moto_id=id_moto).first()

    if not moto.observaciones:
        descripcion = "Sin descripción"
    else:
        descripcion = moto.observaciones
    
    if moto.identificacion_pdf:
        pdf = moto.identificacion_pdf.url
    else:
        pdf = None
    
    compra_venta = ComprasVentas.objects.filter(moto_id=id_moto,tipo="CV").first()

    if compra_venta:
        compra_venta = ComprasVentas.objects.filter(moto_id=id_moto,tipo="CV").latest('id')
        cliente_id = compra_venta.cliente_id
        telefono_principal_cliente = ClienteTelefono.objects.filter(cliente_id=cliente_id,principal=1).first()
        correo_cliente = ClienteCorreo.objects.filter(cliente_id=cliente_id,principal=1).first()

        if correo_cliente:
            correo = correo_cliente.correo
        else:
            correo = None

        telefono_principal = telefono_principal_cliente.telefono
        libreta = req.build_absolute_uri(compra_venta.fotocopia_libreta.url) if compra_venta.fotocopia_libreta else None
        # print("ID CLIENTE --->>>" + str(cliente_id))
        cliente = Cliente.objects.get(id=cliente_id)
    else:
        cliente = None
        libreta = None
        telefono_principal = None
        correo = None

    if matriculas_moto:
    # Obtiene la matrícula actual, o None si no existe
        matricula = Matriculas.objects.filter(moto_id=id_moto).first()
    
        if not matricula: 
            matr_act = None
        else:
            matr_act = matricula.matricula
        
        contexto = {"moto":moto,
                    "descripcion":descripcion,
                    "matr_actual":matr_act,
                    "foto_moto":moto.foto.url if moto.foto else None,
                    "pdf":pdf,
                    "cliente":cliente,
                    "libreta":libreta,
                    "telefono_principal":telefono_principal,
                    "correo":correo,
                    "active_page": 'Motos'
                    }
    else:
        contexto = {"moto":moto,"descripcion":descripcion,"pdf":pdf,"cliente":cliente,"libreta":libreta,"telefono_principal":telefono_principal,"correo":correo,"active_page": 'Motos'}

    return render(req,"perfil_administrativo/motos/detalles_moto.html",contexto)

@admin_required
def vista_inventario_accesorios(req):
    accesorios = Accesorio.objects.filter(activo=1).order_by('-fecha_ingreso')
    logo_um = Logos.objects.get(id=1)
    page_obj = funcion_paginas_varias(req,accesorios)  # Obtiene la página solicitada
    seleccionados = Accesorio.objects.filter(activo=1,seleccionado=1)

    return render(req,"perfil_administrativo/accesorios/accesorios.html",{"seleccionados":seleccionados,'page_obj': page_obj,"accesorios":accesorios,"logo_um":logo_um.logo_UM.url if logo_um.logo_UM else None})

@admin_required
def seleccionar_accesorio(req,id_accesorio):
    accesorios = Accesorio.objects.filter(activo=1).order_by('-fecha_ingreso')
    logo_um = Logos.objects.get(id=1)
    page_obj = funcion_paginas_varias(req,accesorios)  # Obtiene la página solicitada

    accesorio = Accesorio.objects.get(id=id_accesorio)
    accesorio.seleccionado = 1
    accesorio.save()

    seleccionados = Accesorio.objects.filter(activo=1,seleccionado=1)

    return render(req,"perfil_administrativo/accesorios/accesorios.html",{"seleccionados":seleccionados,'page_obj': page_obj,"accesorios":accesorios,"logo_um":logo_um.logo_UM.url if logo_um.logo_UM else None})


@admin_required
def borrar_seleccion_accesorio(req,id_accesorio):
    accesorios = Accesorio.objects.filter(activo=1).order_by('-fecha_ingreso')
    logo_um = Logos.objects.get(id=1)
    page_obj = funcion_paginas_varias(req,accesorios)  # Obtiene la página solicitada

    accesorio = Accesorio.objects.get(id=id_accesorio)
    accesorio.seleccionado = 0
    accesorio.save()

    seleccionados = Accesorio.objects.filter(activo=1,seleccionado=1)

    return render(req,"perfil_administrativo/accesorios/accesorios.html",{"seleccionados":seleccionados,'page_obj': page_obj,"accesorios":accesorios,"logo_um":logo_um.logo_UM.url if logo_um.logo_UM else None})



@admin_required
def alta_accesorio(req):
    # try:
        if req.method == "POST":
            cantidad = int(req.POST['cantidad_accesorios'])
            if cantidad <= 0:
                return render(req,"perfil_administrativo/accesorios/alta_accesorio.html",{"error_message":"La cantidad ingresada es incorrecta"})
            else:

                tipo = req.POST['tipo_accesorio'].upper()     
                marca = req.POST['marca_accesorio'].upper()
                modelo = req.POST['modelo_accesorio'].upper()
                precio = req.POST['precio_accesorio']
                moneda = req.POST['moneda_precio']

                foto = req.FILES.get('foto_accesorio')
                if req.POST['tipo_accesorio'] == "Otro":
                    tipo = "Accesorio"
                else:
                    tipo = req.POST['tipo_accesorio']

                for i in range(1,cantidad + 1):
                    nuevo_accesorio = Accesorio(
                        tipo = tipo,
                        marca = marca,
                        modelo = modelo,
                        activo = 1,
                        foto = foto,
                        precio = precio,
                        fecha_ingreso = datetime.now(),
                        moneda_precio = moneda,
                        talle = req.POST['talle_accesorio']
                    )

                    nuevo_accesorio.save()
                    # print("ACCESORIO: " + str(i))

                
                # return render(req,"perfil_administrativo/accesorios/accesorios.html",{"message":"Accesorio ingresado con éxito"})
                messages.success(req, f"{tipo} ingresado con éxito.")
                return redirect('Accesorios')
        else:
            return render(req,"perfil_administrativo/accesorios/alta_accesorio.html",{})
    # except Exception as e:
    #     pass

@admin_required
def modificacion_accesorio(req,id_accesorio):
    try:
        
        if req.method == "POST": 
            accesorio_upd = Accesorio.objects.get(id=id_accesorio)
            accesorio_upd.tipo = req.POST['tipo_accesorio']
            accesorio_upd.marca = req.POST['marca_accesorio'].upper()
            accesorio_upd.modelo = req.POST['modelo_accesorio'].upper()
            accesorio_upd.precio = req.POST['precio_accesorio']
            accesorio_upd.foto = req.FILES.get('foto_accesorio')
            accesorio_upd.moneda_precio = req.POST['moneda_accesorio']
            accesorio_upd.talle = req.POST['talle_accesorio']
            accesorio_upd.save()
            messages.success(req, "El accesorio ha sido modificado con éxito.")
            return redirect('Accesorios')
        else:
            accesorio = Accesorio.objects.get(id=id_accesorio)
            precio = int(accesorio.precio) if accesorio.precio else 0
            return render(req,"perfil_administrativo/accesorios/modificacion_accesorio.html",{"datos_accesorio":accesorio,"precio_accesorio":precio})
    except Exception as e:
        pass

@admin_required
def baja_accesorio(req,id_accesorio):
    try:
        if req.method == 'POST':
            accesorio_baja = Accesorio.objects.get(id=id_accesorio)
            accesorio_baja.activo = 0
            accesorio_baja.save()
            return render(req, "perfil_administrativo/accesorios/baja_accesorio.html", {"message":"Accesorio borrado con éxito","id_accesorio":0})
        else:
            return render(req, "perfil_administrativo/accesorios/baja_accesorio.html", {"id_accesorio":id_accesorio})
    except Exception as e:
        pass

@admin_required
def detalles_accesorio(req,id_accesorio):
    try:
        accesorio_detalle = Accesorio.objects.get(id=id_accesorio)
        return render(req, "perfil_administrativo/accesorios/detalles_accesorio.html", {"accesorio":accesorio_detalle})
    except Exception as e:
        pass


@admin_required
def venta_accesorio(req,id_cliente):
    # try:
        validar_caja = validar_caja_abierta()
        if not validar_caja:
            cliente = Cliente.objects.get(id=id_cliente)
            documento = cliente.documento
            contexto = contexto_cliente_accesorio(req,"No existe caja del día abierta",documento)
            return render(req,"perfil_administrativo/accesorios/venta_accesorio.html",contexto)
        else:
            fondos = ClienteFondos.objects.filter(cliente_id=id_cliente).first()
            if fondos:
                ult_fondo = ClienteFondos.objects.filter(cliente_id=id_cliente).latest('id')
                ult_fondo_pesos = float(ult_fondo.total_pesos)
                ult_fondo_dolares = float(ult_fondo.total_dolares)
            else: 
                ult_fondo = None
            cbox_fondos = 'incluir_fondos' in req.POST
            if cbox_fondos and ult_fondo and ((req.POST['moneda_venta'] == "Pesos" and float(req.POST['cantidad_destinada_pesos']) > ult_fondo_pesos) or (req.POST['moneda_venta'] == "Dolares" and float(req.POST['cantidad_destinada_dolares']) > ult_fondo_dolares)):
                cliente = Cliente.objects.get(id=id_cliente)
                contexto = contexto_cliente_accesorio(req,"La cantidad destinada no puede exceder el fondo total del cliente",cliente.documento)
                return render(req,"perfil_administrativo/accesorios/venta_accesorio.html",contexto)
            
            elif cbox_fondos and ult_fondo and ((req.POST['moneda_venta'] == "Pesos" and float(req.POST['cantidad_destinada_pesos']) > float(req.POST['precio_total_pesos'])) or (req.POST['moneda_venta'] == "Dolares" and float(req.POST['cantidad_destinada_dolares']) > float(req.POST['precio_total_dolares']))):
                cliente = Cliente.objects.get(id=id_cliente)
                contexto = contexto_cliente_accesorio(req,"La cantidad destinada no puede exceder el precio total del o los accesorios",cliente.documento)
                return render(req,"perfil_administrativo/accesorios/venta_accesorio.html",contexto)
            
            else:
                doc_factura = req.FILES.get('factura_accesorio')
                ult_compra_accesorio = ClienteAccesorio.objects.exists()
                if ult_compra_accesorio:
                    ult_compra_accesorio = ClienteAccesorio.objects.latest('codigo_compra')
                    codigo_compra = int(ult_compra_accesorio.codigo_compra) + 1
                    print("CODIGO: " + str(codigo_compra))
                    print("ID ULT COMPRA: " + str(ult_compra_accesorio.id))
                else:
                    codigo_compra = 1
                accesorios = req.session["accesorios_json"]
                total_pesos = 0
                total_dolares = 0
                dolar = PrecioDolar.objects.get(id=1)
                precio_dolar = float(dolar.precio_dolar_tienda)
                for accesorio in accesorios:
                    # print("INSERT ACCESORIO ID: " + str(accesorio))
                    # print("ID CLIENTE: " + str(id_cliente))
                    # print(codigo_venta)
                    if not doc_factura:
                        doc_factura = None
                    nueva_venta_accesorio = ClienteAccesorio(
                        fecha_compra = datetime.now(),
                        accesorio_id = accesorio,
                        cliente_id = id_cliente,
                        factura_documento = doc_factura,
                        codigo_compra = codigo_compra,
                        forma_de_pago = req.POST['forma_pago']
                    )
                    nueva_venta_accesorio.save()
                    a = Accesorio.objects.get(id=accesorio)
                    if a.moneda_precio == "Pesos":
                        total_pesos = total_pesos + float(a.precio)
                        total_dolares = total_pesos / precio_dolar
                    else:
                        total_dolares = total_dolares + float(a.precio)
                        total_pesos = total_dolares * precio_dolar

                    
                    a = Accesorio.objects.get(id=accesorio)
                    a.activo = 0
                    a.seleccionado = 0
                    a.save()
                    if cbox_fondos:
                        if fondos:
                            ult_fondo = ClienteFondos.objects.filter(cliente_id=id_cliente).latest('id')
                            if float(ult_fondo.total_pesos) > 0 or float(ult_fondo.total_dolares) > 0:
                            
                                if req.POST['moneda_venta'] == "Pesos": 
                                    resta_fondo_pesos = float(ult_fondo.total_pesos) - float(req.POST['cantidad_destinada_pesos'])
                                    resta_fondo_dolares = resta_fondo_pesos / precio_dolar
                                    ingreso_pesos = 0 - float(req.POST['cantidad_destinada_pesos'])
                                    ingreso_dolares = ingreso_pesos / precio_dolar
                                    entrega_pesos = abs(ingreso_pesos)
                                    entrega_dolares = abs(ingreso_dolares)

                                else:
                                    resta_fondo_dolares = float(ult_fondo.total_dolares) - float(req.POST['cantidad_destinada_dolares'])
                                    resta_fondo_pesos = resta_fondo_dolares * precio_dolar
                                    ingreso_dolares = 0 - float(req.POST['cantidad_destinada_dolares'])
                                    ingreso_pesos = ingreso_dolares * precio_dolar
                                
                                resto_pesos = total_pesos - entrega_pesos
                                resto_dolares = total_dolares - entrega_dolares
                                if resto_pesos <= 0 or resto_dolares <= 0:
                                    resto_dolares = 0
                                    resto_pesos = 0
                                alta_cuota_accesorio(req,nueva_venta_accesorio.id,resto_dolares,resto_pesos,req.POST['moneda_venta'],"Pago realizado con los fondos del cliente",precio_dolar,entrega_dolares,entrega_pesos,None,"Efectivo",0,1,None)
                                nuevo_fondo = ClienteFondos(
                                        cliente_id = id_cliente,
                                        moneda = req.POST['moneda_venta'],
                                        ingreso_pesos = ingreso_pesos,
                                        ingreso_dolares = ingreso_dolares,
                                        fecha = datetime.now(),
                                        total_pesos = resta_fondo_pesos,
                                        total_dolares = resta_fondo_dolares,
                                        tipo = "Retiro",
                                        metodo = "----------",
                                        comprobante = None 
                                    )
                                nuevo_fondo.save()

                messages.success(req, "Venta generada con éxito")
                return redirect(f"{reverse('ClienteFicha',kwargs={'id_cliente':id_cliente})}")
    # except Exception as e:
    #     return render(req, "perfil_administrativo/accesorios/venta_accesorio.html", {"error_message":e})
    
def prueba_varios_accesorios(req):
    # accesorios_json = req.POST.get("accesorios", "[]")  # Capturamos el JSON enviado
    # accesorios_seleccionados = json.loads(accesorios_json)  # Convertimos a lista Python
    # accesorios_json = req.POST.get("accesorios", req.session.get("accesorios_json", "[]"))
    # accesorios_seleccionados = json.loads(accesorios_json)

    # # Guardar en sesión para que no se pierdan
    # req.session["accesorios_json"] = accesorios_seleccionados

    #     # Aquí puedes procesar la venta de los 
    # for accesorio in req.session["accesorios_json"]:
    #     #TIPO str
    #     print("Accesorios seleccionados:", accesorio)
    if req.method == "POST":
        # Capturamos la lista de accesorios enviados desde el formulario
        accesorios_seleccionados = req.POST.getlist("accesorios[]")

        # Guardamos en sesión para futuras referencias
        req.session["accesorios_json"] = accesorios_seleccionados

        # Depuración: imprimir los accesorios seleccionados
        print("Accesorios seleccionados:", accesorios_seleccionados)
    
        return render(req, "perfil_administrativo/accesorios/venta_accesorio.html", {})

@admin_required
def cliente_venta_accesorio(req,mostrar,vender):
    # try:
    #     accesorios_json = req.POST.get("accesorios", req.session.get("accesorios_json", "[]"))
    #     accesorios_seleccionados = json.loads(accesorios_json)

    # # Guardar en sesión para que no se pierdan
    #     req.session["accesorios_json"] = accesorios_json
    documento = req.POST['tipo_documento'] + str(req.POST['documento'])
    contexto = contexto_cliente_accesorio(req,None,documento)
    return render(req,"perfil_administrativo/accesorios/venta_accesorio.html",contexto)
                                                                            

    # except Exception as e:
    #     pass

def pagos_accesorio(req,codigo_compra):
    # try: 
        print("CODIGO COMPRA: " + str(codigo_compra))
        data = obtener_compras_accesorios(req,codigo_compra)
        # c_v = ClienteAccesorio.objects.get(id=id_venta)
        return render(req,"perfil_administrativo/accesorios/pagos_accesorios.html",{
                                                                                    "page_obj":data[4],
                                                                                    "id_venta":data[3],
                                                                                    # "id_venta":id_venta,
                                                                                    "id_cliente":data[9],
                                                                                    "cliente_json":data[5],
                                                                                    "accesorio_json":data[6],
                                                                                    "pagos_json":data[7],
                                                                                    # "accesorio":data[4]}
                                                                                    "accesorio":data[0],
                                                                                    "total_pesos":data[1],
                                                                                    "total_dolares":data[2],
                                                                                    "total_precios_json":data[8],
                                                                                    "codigo_compra":codigo_compra,
                                                                                    "forma_pago":data[10]
                                                                                    }
                                                                                    )
    # except Exception as e:
    #     return render(req,"perfil_administrativo/accesorios/pagos_accesorios.html",{"error_message":e})

def alta_paga_accesorio(req,id_venta):
    
        cv = ClienteAccesorio.objects.get(id=id_venta)
        page_obj = obtener_compras_accesorios(req,cv.codigo_compra)
    # try:
        accesorios = ClienteAccesorio.objects.filter(codigo_compra=cv.codigo_compra)
        existe_cuota = CuotasAccesorios.objects.filter(venta_id=id_venta).exists()
        dolar = PrecioDolar.objects.get(id=1)
        precio_dolar = dolar.precio_dolar_tienda
        precio_total_pesos = 0
        precio_total_dolares = 0
        total = req.POST['total_luego_recargo']
        fecha_proximo_pago = req.POST['f_prox_pago'] if req.POST['f_prox_pago'] else None 
        

        if not existe_cuota:
            for suma in accesorios:
                accesorio = Accesorio.objects.get(id=suma.accesorio_id)
                if accesorio.moneda_precio == "Pesos":
                    precio_total_pesos = precio_total_pesos + int(accesorio.precio)
                    precio_total_dolares = precio_total_dolares + (precio_total_pesos / float(precio_dolar))
                else:
                    precio_total_dolares = precio_total_dolares + int(accesorio.precio)
                    precio_total_pesos = precio_total_pesos + (precio_total_dolares * float(precio_dolar))
        else:
            ult_cuota = CuotasAccesorios.objects.filter(venta_id=id_venta).latest('id')
            precio_total_pesos = float(ult_cuota.cant_restante_pesos)
            precio_total_dolares = float(ult_cuota.cant_restante_dolares)

        
        moneda = req.POST['moneda_entrega']
        if moneda == "Pesos":
            resto_pesos = precio_total_pesos - float(total)
            resto_dolares = resto_pesos / float(precio_dolar)
            entrega_pesos = float(total)
            entrega_dolares = entrega_pesos / float(precio_dolar)
        else:
            resto_dolares = precio_total_dolares - float(total)
            resto_pesos = resto_dolares * float(precio_dolar)
            entrega_dolares = float(total)
            entrega_pesos = entrega_dolares * float(precio_dolar)
        
        


        comprobante = req.FILES.get('comprobante_pago')
        # caja = Caja.objects.filter(estado="Abierto").first()
        forma_pago = req.POST['forma_pago']
        # fecha_proximo_pago = req.POST['f_prox_pago']
        observaciones_pago = req.POST['observaciones_pago']
        
        recargo = req.POST['recargo']
        accesorio = Accesorio.objects.get(id=cv.accesorio_id)
        # validar_precio = validar_entrega_menor_precio(moneda,req.POST['valor_a_pagar'],accesorio,precio_dolar,"Accesorio",existe_cuota,id_venta)
        # validar_fecha_proximo_pago = datetime.strptime(fecha_proximo_pago, '%Y-%m-%d')
        fecha_actual = datetime.now().date()
        validar_caja = validar_caja_abierta()
        if forma_pago == "Fondos":
            fondos_cliente = ClienteFondos.objects.filter(cliente=cv.cliente).first()
            if fondos_cliente:
                ult_fondo = ClienteFondos.objects.filter(cliente=cv.cliente).latest('id')
                if moneda == "Pesos":
                    sin_fondos = float(total) > float(ult_fondo.total_pesos)
                else:
                    sin_fondos = float(total) > float(ult_fondo.total_dolares)
            else:
                sin_fondos = True
        else:
            sin_fondos = False
        # valores = valores_compras(existe_cuota,moneda,req.POST['valor_a_pagar'],id_venta,accesorio,"Accesorio",precio_dolar)
        if (entrega_pesos > precio_total_pesos and moneda == "Pesos") or (entrega_dolares > precio_total_dolares and moneda == "Dolares"):
            # return render(req,"perfil_administrativo/accesorios/pagos_accesorios.html",{"error_message":"La entrega no puede ser superior al precio Total del o los accesorios","id_venta":id_venta,"page_obj": page_obj,"id_cliente":cv.cliente_id})
            messages.error(req,"La entrega no puede ser superior al precio Total del o los accesorios.")
            return redirect(reverse('DetallesCompraAccesorio', kwargs={'codigo_compra': cv.codigo_compra}))
        # elif validar_fecha_proximo_pago.date() < fecha_actual:
        #     return render(req,"perfil_administrativo/accesorios/pagos_accesorios.html",{"error_message":"La fecha del próximo pago no debe ser anterior a la actual","id_venta":id_venta,"page_obj": page_obj,"id_cliente":cv.cliente_id})
        elif not validar_caja:
            # return render(req,"perfil_administrativo/accesorios/pagos_accesorios.html",{"error_message":"No existe caja del día abierta.","id_venta":id_venta,"page_obj": page_obj,"id_cliente":cv.cliente_id})
            messages.error(req,"No existe caja del día abierta.")
            return redirect(reverse('DetallesCompraAccesorio', kwargs={'codigo_compra': cv.codigo_compra}))
        elif sin_fondos:
            messages.error(req, "El cliente no dispone de suficientes fondos para realizar el pago correspondiente.")
            return redirect(reverse('DetallesCompraAccesorio', kwargs={'codigo_compra': cv.codigo_compra}))
        else:
            
            # alta = alta_cuota_accesorio(req,id_venta,valores[0],valores[1],moneda,observaciones_pago,precio_dolar,valores[3],valores[2],comprobante,forma_pago,recargo)
            #alta_cuota_accesorio(req,id_cv,resto_dolares,resto_pesos,moneda,observaciones_pago,precio_dolar,entrega_dolares,entrega_pesos,comprobante,forma_pago,recargo)
            #lista = [resto_dolares,resto_pesos,entrega_pesos,entrega_dolares]
            alta = alta_cuota_accesorio(req,id_venta,resto_dolares,resto_pesos,moneda,observaciones_pago,precio_dolar,entrega_dolares,entrega_pesos,comprobante,forma_pago,recargo,1,fecha_proximo_pago)
            if alta.comprobante_pago:
                comprobante_url = alta
            else:
                comprobante_url = None
            
            if forma_pago == "Fondos":
                ult_fondo = ClienteFondos.objects.filter(cliente=cv.cliente).latest('id')
                total_pesos = float(ult_fondo.total_pesos) - float(entrega_pesos)
                total_dolares = float(ult_fondo.total_dolares) - float(entrega_dolares) 
                ingreso_pesos = 0 - float(entrega_pesos)
                ingreso_dolares = 0 - float(entrega_dolares) 
                retiro_fondos = ClienteFondos(
                    moneda = moneda,
                    fecha = fecha_actual,
                    cliente = cv.cliente,
                    ingreso_pesos = ingreso_pesos,
                    ingreso_dolares = ingreso_dolares,
                    total_pesos = total_pesos,
                    total_dolares = total_dolares,
                    comprobante = None,
                    metodo = "----------",
                    tipo = "Retiro"
                )
                retiro_fondos.save()

            if forma_pago != "Fondos":
                if cv.fecha_compra == fecha_actual:
                    tipo = "Ingreso"
                else:
                    tipo = "Ingreso extra"
                movimiento_caja_por_pago_accesorio(req,float(total),id_venta,moneda,forma_pago,tipo,alta,"accesorio")
                #,metodo,tipo,id_venta,producto
            messages.success(req, "Pago ingresado con éxito")
            return redirect(f"{reverse('DetallesCompraAccesorio',kwargs={'codigo_compra':cv.codigo_compra})}?comprobante_url={comprobante_url}")
    # except Exception as e:
    #     return render(req,"perfil_administrativo/accesorios/pagos_accesorios.html",{"error_message":e,"page_obj": page_obj})

def baja_paga_accesorio(req,id_ca):
    # try:
        cuota = CuotasAccesorios.objects.get(id=id_ca)
        id_cv = cuota.venta_id
        venta = ClienteAccesorio.objects.get(id=cuota.venta_id)
        codigo_compra = venta.codigo_compra
        if req.method == "POST":
            validar_caja = validar_caja_abierta()
            if not validar_caja:
                return render(req, "perfil_administrativo/accesorios/baja_pago_accesorio.html", {"error_message":"No existe caja del día abierta","id_cv":id_cv,"codigo_compra":codigo_compra})
            else:
            # if cuota.metodo_pago == "Efectivo":
                if cuota.moneda == "Pesos":
                    quitar_deposito = cuota.valor_pago_pesos
                else:
                    dolar = PrecioDolar.objects.get(id=1)
                    precio_dolar = dolar.precio_dolar_tienda
                    quitar_deposito = cuota.valor_pago_dolares * precio_dolar
                
                if cuota.metodo_pago == "Fondos":
                    dolar = PrecioDolar.objects.get(id=1)
                    precio_dolar = float(dolar.precio_dolar_tienda)
                    # venta = ClienteAccesorio.objects.get(id=id_cv)
                    fondos_cliente = ClienteFondos.objects.filter(cliente=venta.cliente).latest('id')
                    if cuota.moneda == "Pesos":
                        entrega_pesos = float(cuota.valor_pago_pesos)
                        entrega_dolares = entrega_pesos / precio_dolar
                    else:
                        entrega_dolares = float(cuota.valor_pago_dolares)
                        entrega_pesos = entrega_dolares * precio_dolar
                    
                    total_pesos = float(fondos_cliente.total_pesos) + entrega_pesos
                    total_dolares = float(fondos_cliente.total_dolares) + entrega_dolares
                    nuevo_fondo = ClienteFondos(
                        moneda = cuota.moneda,
                        fecha = datetime.now().date(),
                        cliente = fondos_cliente.cliente,
                        ingreso_pesos = entrega_pesos,
                        ingreso_dolares = entrega_dolares,
                        total_pesos = total_pesos,
                        total_dolares = total_dolares,
                        comprobante = None,
                        metodo = "----------",
                        tipo = "Ingreso"
                    )   
                    nuevo_fondo.save()
                
                
                    # insert_movimientos_caja("Se borra pago de accesorio ingresado por error","Egreso",quitar_deposito,caja.id,personal.id,0,0,None,None) 
                mov = MovimientoPagoAccesorio.objects.filter(pago=cuota).first()
                if mov:
                    usuario = req.user
                    caja = Caja.objects.latest('id')
                    
                    movimiento = Movimientos.objects.get(id=mov.movimiento_id)
                    if movimiento.metodo == "Efectivo":
                        caja.depositos = caja.depositos - quitar_deposito
                        caja.save()
                        # if movimiento.moneda == "Pesos":
                        #     monto_a_quitar = float(movimiento.monto)
                        # else:
                        #     dolar = PrecioDolar.objects.get(id=1)
                        #     precio_dolar = float(dolar.precio_dolar_tienda)
                        #     monto_a_quitar = float(mov.monto) * precio_dolar
                        
                        # if movimiento.tipo == "Ingreso" or movimiento.tipo == "Ingreso extra":
                        #     nueva_diferencia = float(caja.diferencia) - monto_a_quitar
                        # else:
                        #     nueva_diferencia = float(caja.diferencia) + monto_a_quitar
                        
                        # caja.diferencia = nueva_diferencia
                    mov.delete()
                    movimiento.delete()
                cuota.delete()    
                return render(req, "perfil_administrativo/accesorios/baja_pago_accesorio.html", {"message":"Pago borrado con éxito","id_cv":id_cv,"codigo_compra":codigo_compra})
        else:
            
            return render(req,"perfil_administrativo/accesorios/baja_pago_accesorio.html",{"id_cv":id_cv,"codigo_compra":codigo_compra})
    # except Exception as e:
    #     return render(req,"perfil_administrativo/accesorios/baja_pago_accesorio.html",{"error_message":e})


@admin_required
def busqueda_tipo_accesorio(req):
    try:
        tipo = req.GET.get('tipo_accesorio')
        accesorio = Accesorio.objects.filter(tipo__icontains=tipo,activo=1)
        page_obj = funcion_paginas_varias(req,accesorio)  # Obtiene la página solicitada
        return render(req,"perfil_administrativo/accesorios/accesorios.html",{'page_obj': page_obj,"accesorios":accesorio})
    except Exception as e:
        pass

@admin_required
def busqueda_codigo_accesorio(req):
    try:
        codigo = req.GET.get('codigo')
        accesorio = Accesorio.objects.filter(id=codigo)
        page_obj = funcion_paginas_varias(req,accesorio)
        return render(req,"perfil_administrativo/accesorios/accesorios.html",{'page_obj': page_obj,"accesorios":accesorio})
    except Exception as e:
        return render(req,"perfil_administrativo/accesorios/accesorios.html",{'message': e})

@admin_required
def busqueda_marca_modelo_accesorio(req):
    try:
        marca = req.GET.get('marca_modelo')
        modelo = req.GET.get('modelo_marca')
        accesorio = Accesorio.objects.filter(marca__icontains=marca,modelo__icontains=modelo,activo=1)
        page_obj = funcion_paginas_varias(req,accesorio)
        return render(req,"perfil_administrativo/accesorios/accesorios.html",{'page_obj': page_obj,"accesorios":accesorio})
    except Exception as e:
        pass

@admin_required
def vista_clientes(req):
    clientes = Cliente.objects.filter(
        cliente_telefono__principal=True
    ).values('id','nombre', 'apellido', 'cliente_telefono__telefono').order_by('nombre')
    page_obj = funcion_paginas_varias(req,clientes)
    return render(req,"perfil_administrativo/cliente/clientes.html",{'page_obj': page_obj,"clientes":clientes})


@admin_required
def alta_cliente(req):
    try:
        if req.method == "POST":
            tipo_doc = req.POST['tipo_doc']
            doc = req.POST['doc']
            nombre = req.POST['nombre'].title()
            apellido = req.POST['apellido'].title()
            f_nac_str = req.POST.get('f_nac')  # Cambiado a paréntesis
            f_nac = datetime.strptime(f_nac_str, '%Y-%m-%d').date() if f_nac_str else None
            telefono_principal = req.POST['telefono_principal']
            telefono_secundario = req.POST['telefono_secundario']
            correo = req.POST['correo_1']
            dominio_correo = req.POST['dominio_correo']
            correo_2 = req.POST['correo_2']
            dominio_correo_2 = req.POST['dominio_correo_2']
            # localidad = req.POST['localidad'].title()
            # calle = req.POST['calle'].title()
            # numero = req.POST['numero']
            # num_apto = req.POST['num_apto']
            
            if correo:
                correo_principal = correo + dominio_correo
                correo_principal = correo_principal.lower()
            else:
                correo_principal = None
            
            if correo_2:
                correo_secundario = correo_2 + dominio_correo_2
                correo_secundario = correo_secundario.lower()
            else:
                correo_secundario = None
            
            doc_compuesto = tipo_doc + str(doc)
            # existe_cliente = valid_cliente(doc_compuesto,telefono_principal,telefono_secundario,correo_principal,correo_secundario)
            existe_cliente = valid_cliente(doc_compuesto)
            if existe_cliente == "existe_cliente":
                return render(req,"perfil_administrativo/cliente/alta_cliente.html",{"error_message":"El cliente ya existe"})
            # elif existe_cliente == "existe_telefono_1":
            #     return render(req,"perfil_administrativo/cliente/alta_cliente.html",{"error_message":"El telefono 1 ya existe"})
            # elif existe_cliente == "existe_telefono_2":
            #     return render(req,"perfil_administrativo/cliente/alta_cliente.html",{"error_message":"El telefono 2 ya existe"})
            # elif existe_cliente == "existe_correo_1":
            #     return render(req,"perfil_administrativo/cliente/alta_cliente.html",{"error_message":"El correo 1 ya existe"})
            # elif existe_cliente == "existe_correo_2":
                # return render(req,"perfil_administrativo/cliente/alta_cliente.html",{"error_message":"El correo 2 ya existe"})
            elif telefono_principal == telefono_secundario:
                return render(req,"perfil_administrativo/cliente/alta_cliente.html",{"error_message":"Los numeros de teléfono no pueden ser iguales"})
            elif (correo_principal and correo_secundario) and (correo_principal == correo_secundario):
                return render(req,"perfil_administrativo/cliente/alta_cliente.html",{"error_message":"Los correos no pueden ser iguales"})
            else:
                # if localidad == "Otro":
                #     ciudad = req.POST['localidad_otro'].title()
                # else:
                #     ciudad = localidad

                # if num_apto:
                #     n_a = num_apto
                # else:
                #     n_a = 0

                nuevo_cliente = Cliente(
                    documento = doc_compuesto,
                    nombre = nombre,
                    apellido = apellido,
                    fecha_nacimiento = f_nac,
                    domicilio = req.POST['domicilio'].title()
                )

                nuevo_cliente.save()
                insert_cliente_telefono(telefono_principal,1,nuevo_cliente.id)

                if telefono_secundario:
                    insert_cliente_telefono(telefono_secundario,0,nuevo_cliente.id)

                if correo_principal:
                    correo_principal = correo_principal.lower()
                    insert_cliente_correo(correo_principal,1,nuevo_cliente.id)
                
                if correo_secundario:
                    correo_secundario = correo_secundario.lower()
                    insert_cliente_correo(correo_secundario,0,nuevo_cliente.id)
                messages.success(req, "El cliente fue ingresado con éxito.")
                return redirect('Clientes')
        else:
            return render(req,"perfil_administrativo/cliente/alta_cliente.html",{})
    except Exception as e:
        pass  

@admin_required
def modificacion_cliente(req,id_cliente):
    try:
        if req.method == "POST":
            tipo_doc = req.POST['tipo_doc']
            doc = req.POST['doc']
            documento = tipo_doc + str(doc)
            tel1 = req.POST['telefono_principal']
            tel2 = req.POST['telefono_secundario']
            correo1 = req.POST['correo_1'] + req.POST['dominio_correo']
            correo2 = req.POST['correo_2'] + req.POST['dominio_correo_2']

            correo1 = correo1.lower()
            correo2 = correo2.lower()
            # valid_cliente = valid_cliente_mod(id_cliente,documento,tel1,tel2,correo1,correo2)
            valid_cliente = valid_cliente_mod(id_cliente,documento)
            if valid_cliente == "existe_cliente":
                contexto = contexto_para_cliente(id_cliente,"El documento ingresado ya existe")
                return render(req,"perfil_administrativo/cliente/modificacion_cliente.html",contexto)
            
            # elif valid_cliente == "existe_tel_principal":
            #     contexto = contexto_para_cliente(id_cliente,"El telefono 1 ingresado ya existe")
            #     return render(req,"perfil_administrativo/cliente/modificacion_cliente.html",contexto)
            # elif valid_cliente == "existe_tel_secundario":
            #     contexto = contexto_para_cliente(id_cliente,"El telefono 2 ingresado ya existe")
            #     return render(req,"perfil_administrativo/cliente/modificacion_cliente.html",contexto)
            # elif valid_cliente == "existe_correo_principal":
            #     contexto = contexto_para_cliente(id_cliente,"El correo 1 ingresado ya existe")
            #     return render(req,"perfil_administrativo/cliente/modificacion_cliente.html",contexto)
            # elif valid_cliente == "existe_correo_secundario":
            #     contexto = contexto_para_cliente(id_cliente,"El correo 2 ingresado ya existe")
            #     return render(req,"perfil_administrativo/cliente/modificacion_cliente.html",contexto)
            elif tel1 == tel2:
                contexto = contexto_para_cliente(id_cliente,"Los telefonos no pueden ser iguales")
                return render(req,"perfil_administrativo/cliente/modificacion_cliente.html",contexto)
            elif (correo1 == correo2) and (req.POST['correo_1'] and req.POST['correo_2']):
                contexto = contexto_para_cliente(id_cliente,"Los correos no pueden ser iguales")
                return render(req,"perfil_administrativo/cliente/modificacion_cliente.html",contexto)
            else:
            
                tel1_actual = ClienteTelefono.objects.filter(cliente_id=id_cliente,principal=1).first()
                if tel1_actual.telefono != tel1:
                    #SI EL TEL1 INGRESADO ES DISTINTO DEL ACTUAL --->>> BORRAR ACTUAL E INGRESAR NUEVO TEL1
                    tel1_actual.delete()
                    insert_cliente_telefono(tel1,1,id_cliente)

                if tel2:
                    tel2_actual = ClienteTelefono.objects.filter(cliente_id=id_cliente,principal=0).first()
                    checkbox = 'convert_to_tel1' in req.POST    
                    if tel2_actual:
                        if tel2_actual.telefono != tel2:
                            #SI EL TEL2 INGRESADO ES DISTINTO DEL ACTUAL --->>> BORRAR ACTUAL E INGRESAR NUEVO TEL2
                            tel2_actual.delete()
                            insert_cliente_telefono(tel2,0,id_cliente)
                    else:
                        insert_cliente_telefono(tel2,0,id_cliente)
                    if checkbox:
                        tel2_actual = ClienteTelefono.objects.filter(cliente_id=id_cliente,principal=0).first()
                        tel2_actual.principal = 1
                        tel1_actual.principal = 0
                        tel2_actual.save()
                        tel1_actual.save()
                else:
                    tel2_actual = ClienteTelefono.objects.filter(cliente_id=id_cliente,principal=0).first()
                    if tel2_actual:
                        tel2_actual.delete()


                if req.POST['correo_1']:
                    correo1_actual = ClienteCorreo.objects.filter(cliente_id=id_cliente,principal=1).first()
                    if correo1_actual:
                        correo1 = correo1.lower()
                        if correo1_actual.correo != correo1:
                            #SI EL CORREO1 INGRESADO ES DISTINTO DEL ACTUAL --->>> BORRAR ACTUAL E INGRESAR NUEVO CORREO1
                            correo1_actual.delete()
                            insert_cliente_correo(correo1,1,id_cliente)
                        else:
                            insert_cliente_correo(correo1,1,id_cliente)
                    else:
                        insert_cliente_correo(correo1,1,id_cliente)
                else:
                    correo1_actual = ClienteCorreo.objects.filter(cliente_id=id_cliente,principal=1).first()
                    if correo1_actual:
                        correo1_actual.delete()
                
                if req.POST['correo_2']:
                    tiene_correo = ClienteCorreo.objects.filter(cliente_id=id_cliente).first()
                    if not tiene_correo:
                        insert_cliente_correo(correo2,1,id_cliente)
                    else:
                        correo2_actual = ClienteCorreo.objects.filter(cliente_id=id_cliente,principal=0).first()
                        if correo2_actual:
                            correo2 = correo2.lower()
                            if correo2_actual.correo != correo2:
                                #SI EL CORREO2 INGRESADO ES DISTINTO DEL ACTUAL --->>> BORRAR ACTUAL E INGRESAR NUEVO CORREO2
                                correo2_actual.delete()
                                insert_cliente_correo(correo2,0,id_cliente)
                        else:
                            insert_cliente_correo(correo2,0,id_cliente)
                    
                    
                    
                    checkbox_correo = 'convert_to_correo1' in req.POST
                    if checkbox_correo:
                        correo2_actual = ClienteCorreo.objects.filter(cliente_id=id_cliente,principal=0).first()
                        correo2_actual.principal = 1
                        correo1_actual.principal = 0 
                        correo1_actual.save()
                        correo2_actual.save()
                else:
                    correo2_actual = ClienteCorreo.objects.filter(cliente_id=id_cliente,principal=0).first()
                    if correo2_actual:
                        correo2_actual.delete()
                
                f_nac_str = req.POST.get('f_nac')  # Cambiado a paréntesis
                f_nac = datetime.strptime(f_nac_str, '%Y-%m-%d').date() if f_nac_str else None

                mod_cliente = Cliente.objects.get(id=id_cliente)
                mod_cliente.documento = documento
                mod_cliente.nombre = req.POST['nombre'].title()
                mod_cliente.apellido = req.POST['apellido'].title()
                mod_cliente.fecha_nacimiento = f_nac
                # mod_cliente.ciudad = req.POST['localidad'].title()
                # mod_cliente.calle = req.POST['calle'].title()
                # mod_cliente.numero = req.POST['numero']
                # mod_cliente.num_apartamento = req.POST['num_apto']
                mod_cliente.domicilio = req.POST['domicilio'].title()
                
                mod_cliente.save()
                messages.success(req, "Cliente modificado con éxito")
                return redirect('Clientes')
        else:
            contexto = contexto_para_cliente(id_cliente,None)
            return render(req,"perfil_administrativo/cliente/modificacion_cliente.html",contexto)
    except Exception as e:
        return render(req,"perfil_administrativo/cliente/modificacion_cliente.html",{"error_message":e})

@admin_required
def buscar_por_doc(req):
    tipo_doc = req.GET.get('tipo_doc_busq')
    doc_num = req.GET.get('documento')
    documento = tipo_doc + str(doc_num)

    cliente = Cliente.objects.filter(
         documento = documento,
         cliente_telefono__principal=True
     ).values('id','nombre', 'apellido', 'cliente_telefono__telefono')
    page_obj = funcion_paginas_varias(req,cliente)  # Obtiene la página solicitada
    return render(req,"perfil_administrativo/cliente/clientes.html",{'page_obj': page_obj,"clientes":cliente})

@admin_required
def buscar_nom_ape(req):
    nombre = req.GET.get('nombre')
    apellido = req.GET.get('apellido')
    if nombre and apellido:
        cliente = Cliente.objects.filter(
            nombre__icontains = nombre,
            apellido__icontains = apellido,
            cliente_telefono__principal=True
        ).values('id','nombre', 'apellido', 'cliente_telefono__telefono')
    elif nombre:
        cliente = Cliente.objects.filter(
            nombre__icontains = nombre,
            cliente_telefono__principal=True
        ).values('id','nombre', 'apellido', 'cliente_telefono__telefono')
    else:
        cliente = Cliente.objects.filter(
            apellido__icontains = apellido,
            cliente_telefono__principal=True
        ).values('id','nombre', 'apellido', 'cliente_telefono__telefono')

    page_obj = funcion_paginas_varias(req,cliente)  # Obtiene la página solicitada
    return render(req,"perfil_administrativo/cliente/clientes.html",{'page_obj': page_obj,"clientes":cliente})

@admin_required
def ficha_cliente(req,id_cliente):
    cliente = Cliente.objects.get(id=id_cliente)
    tel1 = ClienteTelefono.objects.filter(principal=1,cliente_id=cliente.id).first()
    tel_1 = tel1.telefono
    tel2 = ClienteTelefono.objects.filter(principal=0,cliente_id=cliente.id).first()

    correo1 = ClienteCorreo.objects.filter(principal=1,cliente_id=cliente.id).first()
    correo2 = ClienteCorreo.objects.filter(principal=0,cliente_id=cliente.id).first()
    if tel2:
        tel_2 = tel2.telefono
    else:
        tel_2 = None

    if correo1:
        c_1 = correo1.correo
    else:
        c_1 = None
    
    if correo2:
        c_2 = correo2.correo
    else:
        c_2 = None


    resultados_motos_anteriores = (
        ComprasVentas.objects
        .filter(cliente__id=id_cliente, tipo="CV")
        .select_related('moto', 'cliente')
        .values(
            'id',
            'moto__marca', 
            'moto__modelo', 
            'moto__estado',
            'moto__id',
            'cliente__id',
            'fecha_compra', 
            # 'cantidad_cuotas', 
            # 'cuotas_pagas', 
            'moto__precio', 
            'fotocopia_libreta', 
            'compra_venta', 
            'certificado_venta',
            'facturas'
            # 'valor_cuota'
        ).order_by('-id')
    )

    res_documentacion_ma = []
    for resultado in resultados_motos_anteriores:
            cv = ComprasVentas.objects.get(id=resultado['id'])
            c_actual = ComprasVentas.objects.filter(moto_id=resultado['moto__id'],tipo="V").first()
            if c_actual:
                c_actual = ComprasVentas.objects.filter(moto_id=resultado['moto__id'],tipo="V").latest('id')
                id_cliente_actual = c_actual.cliente_id
            else:
                id_cliente_actual = None
            res_documentacion_ma.append({
            'moto': resultado,
            'libreta': cv.fotocopia_libreta.url if cv.fotocopia_libreta else None,
            'compra_venta': cv.compra_venta.url if cv.compra_venta else None,
            "id_cliente_actual":id_cliente_actual
        })


    page_obj_ma = funcion_paginas_varias(req,res_documentacion_ma)
    
    resultados_motos = (
        ComprasVentas.objects
        .filter(cliente__id=id_cliente, tipo__in=['V','R'])
        .select_related('moto', 'cliente')
        .values(
            'id',
            'moto__marca', 
            'moto__modelo', 
            'moto__estado',
            'moto__id',
            'cliente__id',
            'fecha_compra', 
            # 'cantidad_cuotas', 
            # 'cuotas_pagas', 
            'moto__precio', 
            'fotocopia_libreta', 
            'compra_venta', 
            'certificado_venta',
            'facturas'
            # 'valor_cuota'
        ).order_by('-id')
    )

    res_documentacion = []
    for resultado in resultados_motos:
            cv = ComprasVentas.objects.get(id=resultado['id'])
            c_anterior = ComprasVentas.objects.filter(moto_id=resultado['moto__id'],tipo="CV").first()
            if c_anterior:
                c_anterior = ComprasVentas.objects.filter(moto_id=resultado['moto__id'],tipo="CV").latest('id')
                id_cliente_anterior = c_anterior.cliente_id
            else:
                id_cliente_anterior = None
            pagos = CuotasMoto.objects.filter(venta_id=resultado['id']).first()
            if pagos:
                mostrar_boton = False
            else:
                mostrar_boton = True
            res_documentacion.append({
            'moto': resultado,
            'libreta': cv.fotocopia_libreta.url if cv.fotocopia_libreta else None,
            'compra_venta': cv.compra_venta.url if cv.compra_venta else None,
            'certificado_venta': cv.certificado_venta.url if cv.certificado_venta else None,
            'facturas': cv.facturas.url if cv.facturas else None,
            'tipo':cv.tipo,
            "id_cliente_anterior":id_cliente_anterior,
            "boton_borrar":mostrar_boton
            # 'cantidad_cuotas':cv.cantidad_cuotas
        })


    page_obj = funcion_paginas_varias(req,res_documentacion)

    resultados_accesorios = (
        ClienteAccesorio.objects
        .filter(cliente__id=id_cliente,codigo_compra__gte=1)
        .select_related('accesorio', 'cliente')
        .values(
            'id',
            'accesorio__id',
            'accesorio__tipo',
            'accesorio__marca',
            'accesorio__modelo',
            'fecha_compra',
            'factura_documento',
            'codigo_compra'
        ).order_by('-id')
    )
    res_facturas = []
    i = 0
    j = 0
    for resultado_accesorio in resultados_accesorios:
        ca = ClienteAccesorio.objects.filter(codigo_compra=resultado_accesorio['codigo_compra']).first()
        cc = ClienteAccesorio.objects.filter(codigo_compra=resultado_accesorio['codigo_compra']).count()
        # prueba = int(resultado_accesorio['codigo_compra'])
        if cc > 1: 
            if i == 0 or j != int(ca.codigo_compra):
                i = 1
                j = int(ca.codigo_compra)
                # cc = ClienteAccesorio.objects.filter(codigo_compra=resultado_accesorio['codigo_compra'])
                detalles = "Varios"
                res_facturas.append({
                    "detalles":detalles,
                    "fecha":ca.fecha_compra,
                    "codigo_venta":ca.codigo_compra
                })      
        else:
            if resultado_accesorio['codigo_compra'] and int(resultado_accesorio['codigo_compra']) > 0:
                i = 0
                detalles = resultado_accesorio['accesorio__tipo'] + " " + resultado_accesorio['accesorio__marca'] + " " + resultado_accesorio['accesorio__modelo']
                res_facturas.append({
                        "detalles":detalles,
                        "fecha":ca.fecha_compra,
                        "codigo_venta":ca.codigo_compra
                    }) 
        
        
    page_obj_accesorio = funcion_paginas_varias(req,res_facturas)
    fondos = ClienteFondos.objects.filter(cliente=cliente)
    existen_fondos = ClienteFondos.objects.filter(cliente=cliente).first()
    if existen_fondos:
        ultimo_fondo = ClienteFondos.objects.filter(cliente=cliente).latest('id')
        total_pesos = ultimo_fondo.total_pesos
        total_dolares = ultimo_fondo.total_dolares
    else:
        total_pesos = 0
        total_dolares = 0


    return render(req,"perfil_administrativo/cliente/detalles_cliente.html",{"cliente":cliente,
                                                                             "tel1":tel_1,
                                                                             "tel2":tel_2,
                                                                             "correo1":c_1,
                                                                             "correo2":c_2,
                                                                             "page_obj":page_obj,
                                                                             "page_obj_accesorio":page_obj_accesorio,
                                                                             "fondos":fondos,
                                                                             "total_pesos":float(total_pesos),
                                                                             "total_dolares":float(total_dolares),
                                                                             "page_obj_ma":page_obj_ma
                                                                             })

def modificar_moto_vendida(req,id_moto,id_cliente):
    try:
        if req.method == "POST":
                
                moto_upd = Moto.objects.get(id=id_moto)
                checkbox_num_motor = 'con_num_motor' in req.POST
                if moto_upd.contiene_num_motor:
                    num_motor = req.POST['num_motor_moto'].upper()
                else:
                    if checkbox_num_motor:
                        num_motor = req.POST['num_motor_moto_agregado'].upper()
                    else:
                        num_motor = None
                
                checkbox_num_chasis = 'con_num_chasis' in req.POST
                if moto_upd.contiene_num_chasis:
                    num_chasis = req.POST['num_chasis_moto'].upper()
                else:
                    if checkbox_num_chasis:
                        num_chasis = req.POST['num_chasis_moto_agregado'].upper()
                    else:
                        num_chasis = None
                # num_chasis = req.POST['num_chasis_moto'].upper()
                matricula = req.POST['matricula_letras'].upper() + str(req.POST['matricula_numeros'])
                if req.POST['matricula_letras'] and req.POST['matricula_numeros']:
                    letras_matricula = req.POST['matricula_letras']
                    num_matricula = req.POST['matricula_numeros']
                    validar_letra_matricula = departamento_matricula(matricula)
                else:
                    letras_matricula = None
                    num_matricula = None
                    validar_letra_matricula = "No Ingresada"
                padron = req.POST['num_padron']
                if not padron:
                    padron = None
                validar_matricula = matricula_valid(matricula,padron,id_moto)
                validacion_datos_moto = validacion_moto(id_moto,num_motor,num_chasis)
                precio = int(moto_upd.precio) if moto_upd.precio else 0

                if validacion_datos_moto == "existe_num_motor":
                    return render(req,"perfil_administrativo/cliente/modificacion_moto_vendida.html",{"error_message":"El número de motor ya se encuentra asignado a otra moto",
                                                                                            'datos_moto': moto_upd,
                                                                                            "letras_matricula":letras_matricula,
                                                                                            "num_matricula":num_matricula,
                                                                                            "precio":precio
                                                                                            ,"id_cliente":id_cliente}) 
                elif validacion_datos_moto == "existe_num_chasis":
                    return render(req,"perfil_administrativo/cliente/modificacion_moto_vendida.html",{"error_message":"El número de chasis ya se encuentra asignado a otra moto",
                                                                                            'datos_moto': moto_upd,
                                                                                            "letras_matricula":letras_matricula,
                                                                                            "num_matricula":num_matricula,
                                                                                            "precio":precio
                                                                                            ,"id_cliente":id_cliente}) 
                elif (req.POST['matricula_letras'] and req.POST['matricula_numeros']) and (validar_matricula == "matricula_existe"): #EXISTE UNA MATRICULA REGISTRADA Y ADEMAS ES DIFERENTE
                    return render(req,"perfil_administrativo/cliente/modificacion_moto_vendida.html",{"error_message":"La matricula ingresada ya existe",
                                                                                            'datos_moto': moto_upd,
                                                                                            "letras_matricula":letras_matricula,
                                                                                            "num_matricula":num_matricula,
                                                                                            "precio":precio
                                                                                            
                                                                                            ,"id_cliente":id_cliente}) 
                elif (req.POST['num_padron']) and (validar_matricula == "padron_existe"):
                    return render(req,"perfil_administrativo/cliente/modificacion_moto_vendida.html",{"error_message":"El número de padrón ya existe",
                                                                                            'datos_moto': moto_upd,
                                                                                            "letras_matricula":letras_matricula,
                                                                                            "num_matricula":num_matricula,
                                                                                            "precio":precio
                                                                                            
                                                                                            ,"id_cliente":id_cliente})
                elif not validar_letra_matricula:
                    return render(req,"perfil_administrativo/cliente/modificacion_moto_vendida.html",{"error_message":"La matrícula es incorrecta",
                                                                                            'datos_moto': moto_upd,
                                                                                            "letras_matricula":letras_matricula,
                                                                                            "num_matricula":num_matricula,
                                                                                            "precio":precio
                                                                                            
                                                                                            ,"id_cliente":id_cliente})
                else:
                            
                    marca = req.POST['marca_moto'].upper()
                    modelo = req.POST['modelo_moto'].upper()
                    color = req.POST['color_moto'].upper()

                    foto = req.FILES.get('foto_moto')
                    
                    moto_upd = Moto.objects.get(id=id_moto)
                    moto_upd.marca = marca
                    moto_upd.modelo = modelo
                    moto_upd.motor = req.POST['motor_moto']
                    
                    
                    if moto_upd.contiene_num_motor or checkbox_num_motor:
                        moto_upd.num_motor = num_motor
                        moto_upd.contiene_num_motor = 1
                    
                    if moto_upd.contiene_num_chasis or checkbox_num_chasis:
                        moto_upd.num_chasis = num_chasis
                        moto_upd.contiene_num_chasis = 1
                
                    moto_upd.anio = req.POST['anio_moto']
                    moto_upd.tipo = req.POST['tipo_moto']

                    if req.POST['estado_moto'] == "Nueva":
                        estado = "Nueva"
                        moto_upd.kilometros = 0
                    else:
                        estado = "Usada"
                        moto_upd.kilometros = req.POST['km_moto']
                        
                    moto_upd.estado = estado
                    
                    
                    # moto_upd.num_chasis = num_chasis
                    moto_upd.num_cilindros = req.POST['num_cilindros']
                    moto_upd.cantidad_pasajeros = req.POST['num_pasajeros']
                    moto_upd.color = color
                    moto_upd.moneda_precio = req.POST['moneda_precio']
                    moto_upd.precio = req.POST['precio_moto']
                    moto_upd.observaciones = req.POST['descripcion_moto']
                    moto_upd.foto = foto
                    moto_upd.save()

                    matricula_actual = Matriculas.objects.filter(moto_id=id_moto).first()
                    if req.POST['matricula_letras'] and req.POST['matricula_numeros'] and req.POST['num_padron']:
                        if matricula_actual:
                            if matricula_actual.matricula != matricula: #EN CASO DE QUE LA MATRICULA INGRESADA EN EL TEMPLATE SEA DIFERENTE A LA EXISTENTE EN LA BASE DE DATOS, ENTONCES ACCEDE
                                matricula_actual.delete()
                                insert_matriculas(matricula,padron,moto_upd.id)
                        else:
                                insert_matriculas(matricula,padron,moto_upd.id)

                    messages.success(req, "La moto ha sido modificada con éxito.")
                    return redirect(reverse('ClienteFicha', kwargs={'id_cliente': id_cliente}))
        else:
            moto_upd = Moto.objects.get(id=id_moto)
            matricula_upd = Matriculas.objects.filter(moto_id=moto_upd.id).first()

            if matricula_upd:
                letras_matricula = matricula_upd.matricula[0:3:1]
                num_matricula = matricula_upd.matricula[3:7:1]
                padron = matricula_upd.padron
            else:
                letras_matricula = None
                num_matricula = None
                padron = None

            if moto_upd.observaciones == None:
                moto_upd.observaciones = "Sin descripción"
            precio = int(moto_upd.precio) if moto_upd.precio else 0
            return render(req,"perfil_administrativo/cliente/modificacion_moto_vendida.html",{'datos_moto': moto_upd,
            "letras_matricula":letras_matricula,
            "num_matricula":num_matricula,
            "padron":padron,
            "precio":precio,
            "id_cliente":id_cliente}) 
    except Exception as e:
        return render(req,"perfil_administrativo/cliente/modificacion_moto_vendida.html",{'datos_moto': moto_upd,
            "letras_matricula":letras_matricula,
            "num_matricula":num_matricula,"id_cliente":id_cliente,"error_message":str(e)})

def modificar_accesorio_vendido(req,id_accesorio,id_cliente):
    # try:
        accesorio = Accesorio.objects.get(id=id_accesorio)
        venta = ClienteAccesorio.objects.filter(accesorio=accesorio,cliente_id=id_cliente).first()
        if req.method == "POST": 
            accesorio_upd = Accesorio.objects.get(id=id_accesorio)
            accesorio_upd.tipo = req.POST['tipo_accesorio']
            accesorio_upd.marca = req.POST['marca_accesorio'].upper()
            accesorio_upd.modelo = req.POST['modelo_accesorio'].upper()
            
            accesorio_upd.foto = req.FILES.get('foto_accesorio')
            accesorio_upd.moneda_precio = req.POST['moneda_accesorio']
            accesorio_upd.talle = req.POST['talle_accesorio']
            accesorio_upd.precio = req.POST['precio_accesorio']
            dolar = PrecioDolar.objects.get(id=1)
            precio_dolar = float(dolar.precio_dolar_tienda)

            if accesorio.moneda_precio == "Pesos":
                precio_pesos_bd = float(accesorio.precio)
                precio_dolares_bd = precio_pesos_bd / precio_dolar
            else:
                precio_dolares_bd = float(accesorio.precio)
                precio_pesos_bd = precio_dolares_bd * precio_dolar
            codigos_ventas = ClienteAccesorio.objects.filter(codigo_compra=venta.codigo_compra).latest('id')
            existen_cuotas = CuotasAccesorios.objects.filter(venta_id=codigos_ventas.id).first()
            if ((req.POST['moneda_accesorio'] == "Pesos" and precio_pesos_bd != float(req.POST['precio_accesorio'])) or (req.POST['moneda_accesorio'] == "Dolares" and precio_dolares_bd != float(req.POST['precio_accesorio']))):
                if existen_cuotas:
                    ult_cuota = CuotasAccesorios.objects.filter(venta=codigos_ventas.id).latest('id')
                    print("ENTRA AL SEGUNDO IF")
                    if req.POST['moneda_accesorio'] == "Pesos" and precio_pesos_bd < float(req.POST['precio_accesorio']):
                        diferencia_pesos = float(req.POST['precio_accesorio']) - precio_pesos_bd
                        diferencia_dolares = diferencia_pesos / precio_dolar
                        # mensaje = diferencia_pesos
                    elif req.POST['moneda_accesorio'] == "Pesos" and precio_pesos_bd > float(req.POST['precio_accesorio']):
                        diferencia_pesos = float(req.POST['precio_accesorio']) - precio_pesos_bd
                        diferencia_dolares = diferencia_pesos / precio_dolar
                        # mensaje = diferencia_pesos
                    elif req.POST['moneda_accesorio'] == "Dolares" and precio_dolares_bd < float(req.POST['precio_accesorio']):
                        diferencia_dolares = float(req.POST['precio_accesorio']) - precio_dolares_bd
                        diferencia_pesos = diferencia_dolares * precio_dolar
                        # mensaje = diferencia_dolares
                    else:
                        diferencia_dolares = float(req.POST['precio_accesorio']) - precio_dolares_bd
                        diferencia_pesos = diferencia_dolares * precio_dolar
                        # mensaje = diferencia_dolares
                    existen_cuotas.cant_restante_pesos = float(existen_cuotas.cant_restante_pesos) + diferencia_pesos
                    existen_cuotas.cant_restante_dolares = float(existen_cuotas.cant_restante_dolares) + diferencia_dolares
                    existen_cuotas.save()
            # print("PRECIO PESOS BD: " + str(precio_pesos_bd))
            # print("PRECIO PESOS FORM: " + str(float(req.POST['precio_accesorio'])))
            # print("PRECIO DOLARES BD: " + str(precio_dolares_bd))
            # print("PRECIO DOLARES FORM: " + str(float(req.POST['precio_accesorio'])))
            # messages.success(req, mensaje)
            # return redirect(reverse('ModificarAccesorioVendido', kwargs={'id_accesorio': id_accesorio,'id_cliente':id_cliente}))
            accesorio_upd.save()
            messages.success(req, "El accesorio ha sido modificado con éxito.")
            return redirect(reverse('DetallesCompraAccesorio', kwargs={'codigo_compra': venta.codigo_compra}))
        else:
            
            precio = int(accesorio.precio) if accesorio.precio else 0
            
            #
            return render(req,"perfil_administrativo/cliente/modificacion_accesorio_vendido.html",{"datos_accesorio":accesorio,"precio_accesorio":precio,"id_cliente":id_cliente,"id_venta":venta.codigo_compra})
    # except Exception as e:
    #     pass

def borrar_accesorio_vendido(req,codigo_compra,id_accesorio):
    # try:
        cantidad = ClienteAccesorio.objects.filter(codigo_compra=codigo_compra).count()
        venta = ClienteAccesorio.objects.filter(codigo_compra=codigo_compra,accesorio_id=id_accesorio).first()
        if cantidad == 1:
            id_venta = venta.id
        else:
            ult = ClienteAccesorio.objects.filter(codigo_compra=codigo_compra).latest('id')
            penult = ClienteAccesorio.objects.filter(codigo_compra=codigo_compra).order_by('-id')[1]
            if venta.id == ult.id:
                id_venta = penult.id
            else:
                id_venta = ult.id
        
        
        
        pagos_pesos = 0
        pagos_dolares = 0
        total_accesorios_pesos = 0
        total_accesorios_dolares = 0
        dolar = PrecioDolar.objects.get(id=1)
        precio_dolar = float(dolar.precio_dolar_tienda)
        prueba = ClienteAccesorio.objects.filter(codigo_compra=codigo_compra)
        for p in prueba:
            #SUMAR TOTAL DE LA SUMA DEL PRECIO DE LOS ACCESORIOS
            accesorio = Accesorio.objects.get(id=p.accesorio_id)
            if accesorio.moneda_precio == "Pesos":
                total_accesorios_pesos = total_accesorios_pesos + float(accesorio.precio)
                total_accesorios_dolares = total_accesorios_pesos / precio_dolar
            else:
                total_accesorios_dolares = total_accesorios_dolares + float(accesorio.precio)
                total_accesorios_pesos = total_accesorios_dolares * precio_dolar
            pagos = CuotasAccesorios.objects.filter(venta_id=p.id)
            if pagos.exists():
                #SUMAR EL TOTAL DE LOS PAGOS REALIZADOSS
                for pago in pagos:
                        id_venta_pagos = pago.venta_id
                        print(id_venta_pagos)
                        if pago.moneda == "Pesos":
                            pagos_pesos = pagos_pesos + float(pago.valor_pago_pesos)
                            pagos_dolares = pagos_pesos / precio_dolar
                        else:
                            pagos_dolares = pagos_dolares + float(pago.valor_pago_dolares)
                            pagos_pesos = pagos_dolares * precio_dolar
                
        #SI EXISTE AL MENOS UN PAGO QUE MUESTRE EL CHECKBOX DE REGISTRAR EGRESO O INGRESAR AL FONDO DEL CLIENTE
        if (pagos_pesos > 0) or (pagos_dolares > 0):
            mostrar_cbox = True
        else:
            mostrar_cbox = False

        pagos_pesos = round(pagos_pesos,2)
        pagos_dolares = round(pagos_dolares,2)
        
        info_pesos = "$" + str(pagos_pesos)
        info_dolares = "U$s" + str(pagos_dolares)
        accesorio = Accesorio.objects.get(id=venta.accesorio_id)
        if mostrar_cbox:
            opcion_elegida = req.POST.get('opcion_elegida')
            if opcion_elegida == "caja":
                rb_egresos = True
                rb_fondos = False
            else:
                rb_egresos = False
                rb_fondos = True

        if req.method == "POST":

            if mostrar_cbox and (rb_egresos or rb_fondos) and ((int(req.POST['monto_egreso']) > pagos_pesos and req.POST['moneda_devolucion'] == "Pesos") or (int(req.POST['monto_egreso']) > pagos_dolares and req.POST['moneda_devolucion'] == "Dolares")):
                return render(req,"perfil_administrativo/accesorios/baja_accesorio vendido.html",{"codigo_compra":codigo_compra,
                                                                                              "mostrar_cbox":mostrar_cbox,
                                                                                              "info_pesos":info_pesos,
                                                                                              "info_dolares":info_dolares,
                                                                                              "error_message":"El monto del Egreso no puede superar el total de los pagos realizados a la fecha"})
            else:
                # cantidad = ClienteAccesorio.objects.filter(codigo_compra=codigo_compra).count()
                venta = ClienteAccesorio.objects.filter(accesorio_id=id_accesorio,codigo_compra=codigo_compra).first()
                cliente = Cliente.objects.get(id=venta.cliente_id)
                nom_ape = cliente.nombre + " " + cliente.apellido
                venta.codigo_compra = 0
                venta.save()
                accesorio.activo = 1
                accesorio.save()

                

                if mostrar_cbox:
                    if rb_egresos:
                        caja = Caja.objects.latest('id')
                        usuario = req.user
                        personal = Personal.objects.filter(usuario=usuario.username).first()
                        if accesorio.moneda_precio == "Pesos":
                            monto_egreso = req.POST['monto_egreso']
                        else:
                            monto_egreso = req.POST['monto_egreso']
                        insert_movimientos_caja(f"Egreso por baja de accesorio vendido al cliente {nom_ape}","Egreso",monto_egreso,caja.id,personal.id,req.POST['moneda_devolucion'],None,req.POST['forma_devolucion'],0,1,None,None)
                    elif rb_fondos:
                        if req.POST['moneda_devolucion'] == "Pesos":
                            fondos_pesos = float(req.POST['monto_egreso'])
                            fondos_dolares = fondos_pesos / precio_dolar
                        else:
                            fondos_dolares = float(req.POST['monto_egreso'])
                            fondos_pesos = fondos_dolares * precio_dolar
                        fondos_cliente = ClienteFondos.objects.filter(cliente_id=venta.cliente_id).first()
                        if fondos_cliente:
                            fondos_cliente = ClienteFondos.objects.filter(cliente_id=venta.cliente_id).latest('id')
                            total_pesos = float(fondos_cliente.total_pesos) + fondos_pesos
                            total_dolares = float(fondos_cliente.total_dolares) + fondos_dolares
                        else:
                            total_pesos = fondos_pesos
                            total_dolares = fondos_dolares
                        nuevo_fondo = ClienteFondos(
                            cliente_id = venta.cliente_id,
                            moneda = req.POST['moneda_devolucion'],
                            ingreso_pesos = fondos_pesos,
                            ingreso_dolares = fondos_dolares,
                            fecha = datetime.now(),
                            total_pesos = total_pesos,
                            total_dolares = total_dolares,
                            tipo = "Ingreso",
                            metodo = req.POST['forma_devolucion'],
                            comprobante = None
                        )
                        nuevo_fondo.save()
                    # pagos = CuotasAccesorios.objects.filter(venta_id=p.id).first()
                    # if pagos:
                    #     pagos = CuotasAccesorios.objects.filter(venta_id=p.id).latest('id')
                    #     if req.POST['moneda_devolucion'] == "Pesos":
                    #         resto_pesos = float(pagos.cant_restante_pesos) + float(req.POST['monto_egreso'])
                    #         resto_dolares = resto_pesos / precio_dolar
                    #     else:
                    #         resto_dolares = float(pagos.cant_restante_dolares) + float(req.POST['monto_egreso'])
                    #         resto_pesos = resto_dolares * precio_dolar
                        
                    # prueba = ClienteAccesorio.objects.filter(codigo_compra=codigo_compra)
                    pagos = CuotasAccesorios.objects.filter(venta_id=id_venta_pagos)
                    if pagos.exists():
                        accesorio_baja = Accesorio.objects.get(id=id_accesorio)
                        if accesorio_baja.moneda_precio == "Pesos":
                            resta_pesos = total_accesorios_pesos - float(accesorio_baja.precio) - pagos_pesos
                            resta_dolares = resta_pesos / precio_dolar
                        else:
                            resta_dolares = total_accesorios_dolares - float(accesorio_baja.precio) - pagos_dolares
                            resta_pesos = resta_dolares * precio_dolar
                        
                        if req.POST['moneda_devolucion'] == "Pesos":
                            resta_pesos = resta_pesos + float(req.POST['monto_egreso'])
                            resta_dolares = resta_pesos / precio_dolar
                        else:
                            resta_dolares = resta_dolares + float(req.POST['monto_egreso'])
                            resta_pesos = resta_dolares * precio_dolar

                        for p in pagos:
                            p.venta_id = id_venta
                            p.save()
                        
                        alta_cuota_accesorio(req,id_venta,resta_dolares,resta_pesos,req.POST['moneda_devolucion'],"Reajuste por baja de accesorio",precio_dolar,0,0,None,req.POST['forma_devolucion'],0,0,None)
                messages.success(req, "Accesorio borrado de la ficha correctamente.")
                return redirect(reverse('ClienteFicha', kwargs={'id_cliente': cliente.id}))
        else:
            
            return render(req,"perfil_administrativo/accesorios/baja_accesorio vendido.html",{"codigo_compra":codigo_compra,
                                                                                              "mostrar_cbox":mostrar_cbox,
                                                                                              "info_pesos":info_pesos,
                                                                                              "info_dolares":info_dolares})
    # except Exception as e:
    #     pass

def fondos_cliente(req,id_cliente):
    # try:
        tipo = req.POST['accion_fondo']
        ult_fondo = ClienteFondos.objects.filter(cliente_id=id_cliente).first()
        validar_caja = validar_caja_abierta()
        if not ult_fondo and tipo == "Retiro":
            messages.error(req, "No se pueden realizar retiros ya que no existen fondos del cliente.")
            return redirect(reverse('ClienteFicha', kwargs={'id_cliente': id_cliente}))
        elif not validar_caja:
            messages.error(req, "No existe caja del día abierta.")
            return redirect(reverse('ClienteFicha', kwargs={'id_cliente': id_cliente}))
        else: 
            dolar = PrecioDolar.objects.get(id=1)
            precio_dolar = dolar.precio_dolar_tienda
            
            if req.POST['moneda_fondos'] == "Pesos":
                ingreso_pesos = int(req.POST['monto_fondos'])
                ingreso_dolares = int(ingreso_pesos) / float(precio_dolar)
            else:
                ingreso_dolares = int(req.POST['monto_fondos'])
                ingreso_pesos = int(ingreso_dolares) * float(precio_dolar)

            if ult_fondo:
                ult_fondo = ClienteFondos.objects.filter(cliente_id=id_cliente).latest('id')
                total_pesos = int(ult_fondo.total_pesos)
                total_dolares = int(ult_fondo.total_dolares)
                monto = int(req.POST['monto_fondos'])
                if (tipo == "Retiro" and req.POST['moneda_fondos'] == "Pesos" and monto > total_pesos) or (tipo == "Retiro" and req.POST['moneda_fondos'] == "Dolares" and monto > total_dolares):
                    messages.error(req, "El retiro no puede exceder el total acumulado en precio o dolares.")
                    return redirect(reverse('ClienteFicha', kwargs={'id_cliente': id_cliente}))
            else:
                total_pesos = 0
                total_dolares = 0
            
            if tipo == "Ingreso":
                total_pesos = total_pesos + int(ingreso_pesos)
                total_dolares = total_dolares + int(ingreso_dolares)
            else:
                total_pesos = total_pesos - int(ingreso_pesos)
                total_dolares = total_dolares - int(ingreso_dolares)
                ingreso_pesos = 0 - int(ingreso_pesos)
                ingreso_dolares = 0 - int(ingreso_dolares)
            
            comprobante = req.FILES.get('comprobante_fondo') if req.FILES.get('comprobante_fondo') else None
            
            nuevo_fondo = ClienteFondos(
                cliente_id = id_cliente,
                moneda = req.POST['moneda_fondos'],
                ingreso_pesos = int(ingreso_pesos),
                ingreso_dolares = int(ingreso_dolares),
                fecha = datetime.now(),
                total_pesos = total_pesos,
                total_dolares = total_dolares,
                tipo = tipo,
                metodo = req.POST['metodo_elegido'] if tipo == "Ingreso" else "----------",
                comprobante = comprobante 
            )
            #metodo_elegido
            nuevo_fondo.save()
            caja = Caja.objects.latest('id')
            cliente = Cliente.objects.get(id=id_cliente)
            usuario = req.user
            personal = Personal.objects.filter(usuario=usuario.username).first()
            if tipo == "Ingreso":
                movimiento = "Ingreso extra"
                descripcion = f"Ingreso de fondos del cliente {cliente.nombre} {cliente.apellido}"
            else:
                movimiento = "Egreso"
                descripcion = f"Engreso de fondos del cliente {cliente.nombre} {cliente.apellido}"
            insert_movimientos_caja(descripcion,movimiento,int(req.POST['monto_fondos']),caja.id,personal.id,req.POST['moneda_fondos'],None,req.POST['metodo_elegido'],0,0,None,None)

            #(movimiento_descripcion,tipo,monto,id_caja,id_personal,moneda,rubro,metodo,es_moto,es_accesorio,id_venta,producto)
            messages.success(req, "Fondos actualizados correctamente.")
            return redirect(reverse('ClienteFicha', kwargs={'id_cliente': id_cliente}))

    # except Exception as e:
    #     pass

@admin_required
def cargar_certificado(req,id_cv):
    try:
         venta = ComprasVentas.objects.get(id=id_cv)
         cliente = Cliente.objects.get(id=venta.cliente_id)
         moto = Moto.objects.get(id=venta.moto_id)
         tel_cliente = ClienteTelefono.objects.filter(cliente=cliente,principal=1).first()
         telefono = tel_cliente.telefono
         crear_certificado_bikeup(cliente,telefono,moto,id_cv)
        #  certificado = req.FILES.get('certificado_venta')
        #  venta.certificado_venta = certificado
        #  venta.save()
         messages.success(req, "Certificado actualizado con éxito")
         return redirect(f"{reverse('ClienteFicha',kwargs={'id_cliente':venta.cliente_id})}")
    except Exception as e:
        pass

@admin_required
def cargar_libreta(req,id_cv):
    try:
         libreta = req.FILES.get('libreta_venta')
         venta = ComprasVentas.objects.get(id=id_cv)
         venta.fotocopia_libreta = libreta
         venta.save()
         messages.success(req, "Libreta ingresada con éxito")
         return redirect(f"{reverse('ClienteFicha',kwargs={'id_cliente':venta.cliente_id})}")
    except Exception as e:
        pass

@admin_required
def cargar_factura(req,id_cv):
    try:
         factura = req.FILES.get('factura_venta')
         venta = ComprasVentas.objects.get(id=id_cv)
         venta.facturas = factura
         venta.save()
         messages.success(req, "Factura ingresada con éxito")
         return redirect(f"{reverse('ClienteFicha',kwargs={'id_cliente':venta.cliente_id})}")
    except Exception as e:
        pass

@admin_required
def cargar_ccv(req,id_cv):
    #generar_compromiso_compra_venta(req,id_moto,id_cliente)
    try:
        cv = ComprasVentas.objects.get(id=id_cv)
        id_cliente = cv.cliente_id
        ccv = req.FILES.get('ccv_venta')
        cv.compra_venta = ccv
        cv.save()
        messages.success(req, "Compromiso compraventa actualizado con éxito")
        return redirect(f"{reverse('ClienteFicha',kwargs={'id_cliente':id_cliente})}")
    except Exception as e:
        pass

def cargar_ccv_ma(req,id_cv,id_moto):
    #generar_compromiso_compra_venta(req,id_moto,id_cliente)
    try:
        cv = ComprasVentas.objects.get(id=id_cv)
        id_cliente = cv.cliente_id
        ccv = req.FILES.get('ccv_venta')
        cv.compra_venta = ccv
        cv.save()
        moto = Moto.objects.get(id=id_moto)
        moto.pertenece_tienda = 1
        moto.save()
        messages.success(req, "Compromiso compraventa actualizado con éxito. La moto ya se puede visualizar en el inventario.")
        return redirect(f"{reverse('ClienteFicha',kwargs={'id_cliente':id_cliente})}")
    except Exception as e:
        pass

def borrar_documentacion(req,id_cv,tipo):
    try:
        cv = ComprasVentas.objects.get(id=id_cv)
        if tipo == "Libreta":
            cv.fotocopia_libreta = None
        elif tipo == "Certificado":
            cv.certificado_venta = None
        else:
            cv.facturas = None
        
        cv.save()
        id_cliente = cv.cliente_id
        messages.success(req, "Datos actualizados con éxito")
        return redirect(f"{reverse('ClienteFicha',kwargs={'id_cliente':id_cliente})}")
    except Exception as e:
        pass

@admin_required
def vista_personal(req):
    administrativos = (Administrativo.objects
                       .filter(activo=True).exclude(usuario="adminapp")
                       .values('id', 'nombre', 'apellido', 'telefono', 'correo', 'activo','usuario')
                       .order_by('nombre'))
    page_obj = funcion_paginas_varias(req,administrativos)

    mecanicos = (Mecanico.objects
                       .filter(activo=True).exclude(usuario="adminapp")
                       .values('id', 'nombre', 'apellido', 'telefono', 'correo', 'activo')
                       .order_by('nombre'))
    page_objMec = funcion_paginas_varias(req,mecanicos)

    return render(req,"perfil_administrativo/personal/personal.html",{"page_obj":page_obj,"page_objMec":page_objMec})


@admin_required
def alta_personal(req):
    try:
        if req.method == "POST":
            documento = req.POST['tipo_doc'] + str(req.POST['doc'])
            telefono = req.POST['telefono'] 
            correo_nombre = req.POST['correo'] 
            correo_dominio = req.POST['dominio_correo'] 
            correo = correo_nombre + correo_dominio
            correo = correo.lower()

            valid_personal = validar_personal(documento,telefono,correo)

            if valid_personal == "existe_admin":
                return render(req,"perfil_administrativo/personal/alta_personal.html",{"error_message":"La persona que desea ingresar ya existe en el sistema"})
            elif valid_personal == "existe_telefono":
                return render(req,"perfil_administrativo/personal/alta_personal.html",{"error_message":"El teléfono ingresado ya existe en el sistema, el mismo no debe estar repetido"})
            elif valid_personal == "existe_correo":
                return render(req,"perfil_administrativo/personal/alta_personal.html",{"error_message":"El correo ingresado ya existe en el sistema, el mismo no debe estar repetido"})
            elif valid_personal == "admin_desactivado":    
                persona = Personal.objects.filter(documento=documento).first()
                id = persona.id
                return render(req,"perfil_administrativo/personal/alta_personal.html",{"reingresar":True,"id_personal":id})
            else:
                    f_nac_str = req.POST.get('f_nac')  # Cambiado a paréntesis
                    f_nac = datetime.strptime(f_nac_str, '%Y-%m-%d').date() if f_nac_str else None
                    usuario = nombre_usuario(req.POST['nombre'],req.POST['apellido'])
                    
                    personal = Personal.objects.create(
                        documento=documento,
                        nombre=req.POST['nombre'].capitalize(),
                        apellido=req.POST['apellido'].capitalize(),
                        fecha_nacimiento=f_nac,
                        usuario=usuario,
                        contrasena="Inicio1234",
                        correo=correo,
                        telefono=telefono,
                        primera_sesion = True
                    )
                    personal.save()
                    # Crear registro en Administrativo asociado al Personal
                    administrativo = Administrativo.objects.create(
                        personal_ptr=personal,  # Asociar al registro de Personal
                        activo=True,
                        documento=documento,
                        nombre=req.POST['nombre'].capitalize(),
                        apellido=req.POST['apellido'].capitalize(),
                        fecha_nacimiento=f_nac,
                        usuario=usuario,
                        contrasena="Inicio1234",
                        correo=correo,
                        telefono=telefono,
                        primera_sesion = True
                    )
                    administrativo.save()

                    # Crear registro en Mecanico asociado al Personal
                    mecanico = Mecanico.objects.create(
                        personal_ptr=personal,  # Asociar al registro de Personal
                        activo=False,
                        jefe=False,
                        documento=documento,
                        nombre=req.POST['nombre'].capitalize(),
                        apellido=req.POST['apellido'].capitalize(),
                        fecha_nacimiento=f_nac,
                        usuario=usuario,
                        contrasena="Inicio1234",
                        correo=correo,
                        telefono=telefono,
                        primera_sesion = True
                    )
                    mecanico.save()

                    user = User.objects.create_user(
                        username=usuario,
                        password="Inicio1234",  # La contraseña predeterminada
                        email=correo,
                        first_name=req.POST['nombre'].capitalize(),
                        last_name=req.POST['apellido'].capitalize()
                    )
                    
                    # Asignar permisos o grupos si es necesario
                    user.is_staff = False  # Si deseas que tenga acceso al admin de Django
                    user.save()

                    messages.success(req, "Personal ingresado con éxito")
                    return redirect('Personal')
                
        else:
            return render(req,"perfil_administrativo/personal/alta_personal.html",{})
    except Exception as e:
        return render(req,"perfil_administrativo/personal/alta_personal.html",{"error_message":e})


@admin_required
def ingresar_tienda(req,id_personal):
    try:
        administrativo = Administrativo.objects.get(personal_ptr_id=id_personal)
        administrativo.activo = 1
        administrativo.save()
        messages.success(req, "Personal reingresado con éxito")
        return redirect('Personal')
    except Exception as e:
        return render(req,"perfil_administrativo/personal/alta_personal.html",{"error_message":e})

@admin_required
def detalles_personal(req,id_personal):
    try:
        personal = Personal.objects.get(id=id_personal)
        administrativo = Administrativo.objects.filter(personal_ptr_id = id_personal,activo = 1).first()
        if administrativo:
            admin = True
        else:
            admin = False
        
        mecanico = Mecanico.objects.filter(personal_ptr_id = id_personal,activo = 1).first()
        if mecanico:
            mec = True
            if mecanico.jefe == 1:
                jefe = True
            else:
                jefe = False
        else:
            mec = False
            jefe = False
        contexto = {
            "personal":personal,
            "admin":admin,
            "mec":mec,
            "jefe":jefe
        }
        return render(req,"perfil_administrativo/personal/detalles_personal.html",contexto)
    except Exception as e:
         pass

def baja_personal(req,donde,id_personal):
    # try:
        personal = Personal.objects.get(id=id_personal)
        if req.method == "POST":
            if donde == "tienda":
                admin = Administrativo.objects.get(id=id_personal)
                admin.activo = 0
                admin.save()
            else:
                mec = Mecanico.objects.get(id=id_personal)
                mec.activo = 0
                mec.save()
            return render(req,"perfil_administrativo/personal/baja_personal.html",{"message":"Personal dado de baja con éxito"})
        else:
            return render(req,"perfil_administrativo/personal/baja_personal.html",{"personal":personal})
    # except Exception as e:
    #     pass

@admin_required
def form_venta_moto(req,id_moto):
    try:
        if req.method == "POST":
            tipo_doc = req.POST['tipo_documento']
            doc = req.POST['documento']
            documento = tipo_doc + str(doc)
            contexto = contexto_venta_moto(req,id_moto,None,documento)
            return render(req,"perfil_administrativo/motos/venta_moto.html",contexto)
        else:
            return render(req,"perfil_administrativo/motos/venta_moto.html",{})
    except Exception as e:
        return render(req,"perfil_administrativo/motos/venta_moto.html",{"error_message":e})





@admin_required
def venta_moto(req,id_moto,id_cliente):
    # try:
        cliente = Cliente.objects.get(id=id_cliente)
        fondos = int(req.POST['cantidad_destinada'])
        moto = Moto.objects.get(id=id_moto)
        fondos_cliente = ClienteFondos.objects.filter(cliente=cliente).first()
        precio_moto = int(moto.precio)
        cv = ComprasVentas.objects.filter(moto_id=id_moto,cliente_id=id_cliente).first()
        validar_caja = validar_caja_abierta()
        if fondos_cliente:
            fondos_cliente = ClienteFondos.objects.filter(cliente=cliente).latest('id')
            #PERMITE QUE NO SE INGRESEN MAS FONDOS DE LOS DISPONIBLES
            if moto.moneda_precio == "Pesos":
                mostrar_error = fondos > fondos_cliente.total_pesos
                mensaje = "La cantidad de fondos destinada a la compra de la moto supera la cantidad de fondos totales del cliente" if mostrar_error else None
            else:
                mostrar_error = fondos > fondos_cliente.total_dolares
                mensaje = "La cantidad de fondos destinada a la compra de la moto supera la cantidad de fondos totales del cliente" if mostrar_error else None
        elif (fondos > precio_moto and moto.moneda_precio == "Pesos") or (fondos > precio_moto and moto.moneda_precio == "Dolares"):
            #PERMITE QUE NO SE SUPERE EL PRECIO DE LA MOTO
            mostrar_error = True
            mensaje = "La cantidad de fondos destinada a la compra de la moto supera el precio de la moto"
        elif not fondos_cliente and fondos > 0:
            #PERMITE QUE NO SE INGRESEN FONDOS SI NO EXISTEN FONDOS A NOMBRE DEL CLIENTE
            mostrar_error = True
            mensaje = "El cliente no cuenta con fondos"
        elif fondos < 0:
            mostrar_error = True
            mensaje = "La cantidad de fondos destinada a la compra de la moto es menor a 0"
        elif cv:
            existen_pagos = CuotasMoto.objects.filter(venta=cv).first()
            if existen_pagos:
                ultimo_pago = CuotasMoto.objects.filter(venta=cv).latest('id')
                if moto.moneda_precio == "Pesos":
                    mostrar_error = fondos > ultimo_pago.cant_restante_pesos
                    mensaje = "La cantidad de fondos destinada a la compra de la moto supera la cantidad restante con los pagos realizados" if mostrar_error else None
                else:
                    mostrar_error = fondos > ultimo_pago.cant_restante_dolares
                    mensaje = "La cantidad de fondos destinada a la compra de la moto supera la cantidad restante con los pagos realizados" if mostrar_error else None
            else:
                mostrar_error = False
        # elif not validar_caja:
        #     mostrar_error = True
        #     mensaje = "No existe caja del día abierta."
        else:
            mostrar_error = False
        
        if not validar_caja:
            mostrar_error = True
            mensaje = "No existe caja del día abierta."

        if mostrar_error:
            contexto = contexto_venta_moto(req,id_moto,mensaje,cliente.documento)
            return render(req,"perfil_administrativo/motos/venta_moto.html",contexto)
        else:
            
            moto.pertenece_tienda = 0
            # moto.precio_final = req.POST['precio_recargo']
            moto.save()
            compra_venta = req.FILES.get('compra_venta_moto')
            
            existe_reserva = ComprasVentas.objects.filter(moto_id=id_moto,cliente_id=id_cliente,tipo="R").first()
            if existe_reserva:
                # existe_reserva.delete() NO BORRAR YA QUE EN CASO DE HABER UN PAGO EN CUOTASMOTOS (SEÑA) SE BORRARIA TAMBIEN
                # ADEMAS LA RESERVA NO ES UN DATO QUE SE NECESITE CONSERVAR UNA VEZ VENDIDA LA MOTO
                existe_reserva.tipo = "V"
                existe_reserva.fecha_compra = datetime.now()
                existe_reserva.compra_venta = compra_venta
                existe_reserva.forma_de_pago = req.POST['forma_pago']
                existe_reserva.save()
                id_cv = existe_reserva.id
            else:
                id_cv = insert_compras_ventas("V",None,id_cliente,id_moto,compra_venta,None,req.POST['forma_pago'])
            
            checkbox_incluir_fondos = 'incluir_fondos' in req.POST
            if checkbox_incluir_fondos:
                dolar = PrecioDolar.objects.get(id=1)
                precio_dolar = dolar.precio_dolar_tienda
                

                #INSERT EN ClienteFondos COMO tipo = 'Retiro'
                #TOMAR ULTIMO VALOR DE CUOTASMOTO
                #CANTIDAD RESTANTE - FONDOS
                #SI NO EXISTE CUOTASMOTO, RESTAR PRECIO MOTO - FONDOS
                #INSERT EN CuotasMoto
                fondos_cliente = ClienteFondos.objects.filter(cliente=cliente).first()
                if fondos_cliente and fondos > 0:
                    if moto.moneda_precio == "Pesos":
                        ingreso_pesos = fondos
                        ingreso_dolares = fondos / float(precio_dolar)
                        ingreso_pesos = 0 - ingreso_pesos
                        ingreso_dolares = 0 - ingreso_dolares

                        fondos_cliente = ClienteFondos.objects.filter(cliente=cliente).latest('id')
                        total_pesos = int(fondos_cliente.total_pesos) - fondos
                        total_dolares = total_pesos / float(precio_dolar)
                    else:
                        ingreso_dolares = fondos
                        ingreso_pesos = fondos * float(precio_dolar)
                        ingreso_pesos = 0 - ingreso_pesos
                        ingreso_dolares = 0 - ingreso_dolares
                        
                        fondos_cliente = ClienteFondos.objects.filter(cliente=cliente).latest('id')
                        total_dolares = int(fondos_cliente.total_dolares) - fondos
                        total_pesos = total_dolares * float(precio_dolar)

                

                    nuevo_fondo = ClienteFondos(
                                cliente_id = id_cliente,
                                moneda = moto.moneda_precio,
                                ingreso_pesos = int(ingreso_pesos),
                                ingreso_dolares = int(ingreso_dolares),
                                fecha = datetime.now(),
                                total_pesos = int(total_pesos),
                                total_dolares = int(total_dolares),
                                tipo = "Retiro",
                                metodo = "----------",
                                comprobante = None 
                            )
                    nuevo_fondo.save()

                    ult_cuota = CuotasMoto.objects.filter(venta_id=id_cv).first()
                    if ult_cuota:
                        #SI EXISTE UN PAGO, SE LE RESTAN LOS FONDOS DEL CLIENTE A LA CANTIDAD RESTANTE
                        ult_cuota = CuotasMoto.objects.filter(venta_id=id_cv).latest('id')
                        if moto.moneda_precio == "Pesos":
                            resto_pesos = int(ult_cuota.cant_restante_pesos) - fondos
                            resto_dolares = float(resto_pesos) / float(precio_dolar)
                        else:
                            resto_dolares = int(ult_cuota.cant_restante_dolares) - fondos
                            resto_pesos = resto_dolares * float(precio_dolar)
                        # alta_cuota_funcion(req,None,id_cv,int(resto_dolares),int(resto_pesos),moto.moneda_precio,None,float(precio_dolar),abs(ingreso_dolares),abs(ingreso_pesos),None,"Fondos",False,"Entrega")

                    else:
                        #SI NO EXISTEN PAGOS, SE LE RESTAN LOS FONDOS AL PRECIO DE LA MOTO Y SE INGRESA COMO PAGO DE LA MOTO EN CUOTASMOTO
                        if moto.moneda_precio == "Pesos":
                            resto_pesos = int(moto.precio) - fondos
                            resto_dolares = float(resto_pesos) / float(precio_dolar)
                        else:
                            resto_dolares = int(moto.precio) - fondos
                            resto_pesos = resto_dolares * float(precio_dolar)
                    
                    alta_cuota_funcion(req,None,id_cv,int(resto_dolares),int(resto_pesos),moto.moneda_precio,"Fondos",float(precio_dolar),abs(ingreso_dolares),abs(ingreso_pesos),None,None,False,"Entrega")
                                    #(req,fecha_prox_pago,id_cv,resto_dolares,resto_pesos,moneda,observaciones_pago,precio_dolar,entrega_dolares,entrega_pesos,comprobante,forma_pago,ingresar_fin,tipo_pago)

            telefono = ClienteTelefono.objects.filter(cliente_id=id_cliente,principal=1).first()
            # alta_financiamientos(req.POST['recargo'],req.POST['cant_cuotas'],req.POST['valor_cuota'],moto.moneda_precio,1,id_cv,1)
            checkbox_generar_comprobante_venta = 'comprobante_venta' in req.POST
            if checkbox_generar_comprobante_venta:
                crear_certificado_bikeup(cliente,telefono.telefono,moto,id_cv)
            # #REDIRIGIR A LA FICHA DEL CLIENTE
            messages.success(req, "Venta generada con éxito")
            return redirect(f"{reverse('ClienteFicha',kwargs={'id_cliente':id_cliente})}")
    # except Exception as e:
    #     return render(req,"perfil_administrativo/motos/venta_moto.html",{"error_message":e})

@admin_required
def vista_ventas(req):
    try:
            resultados_motos = (
                ComprasVentas.objects
                .filter(tipo='V')
                .select_related('moto','cliente')
                .values(
                    'id',
                    'moto__marca', 
                    'moto__modelo', 
                    'fecha_compra', 
                    'cliente__nombre',
                    'cliente__apellido',
                    'cliente__id'
                ).order_by('-fecha_compra')
            )

            res_documentacion = []
            for resultado in resultados_motos:
                cv = ComprasVentas.objects.get(id=resultado['id'])
                res_documentacion.append({
                'moto': resultado
            })

            page_obj = funcion_paginas_varias(req,res_documentacion)
            
            resultados_accesorios = (
                ClienteAccesorio.objects
                # .filter(cliente__id=1)
                .select_related('accesorio', 'cliente')
                .values(
                    'id',
                    'accesorio__tipo',
                    'accesorio__marca',
                    'accesorio__modelo',
                    'fecha_compra',
                    'cliente__nombre',
                    'cliente__apellido',
                    'cliente__id'
                
                ).order_by('-fecha_compra')
            )
            res_facturas = []
            for resultado_accesorio in resultados_accesorios:
                    ca = ClienteAccesorio.objects.get(id=resultado_accesorio['id'])
                    res_facturas.append({
                    'accesorio': resultado_accesorio,
                    'factura_documento': ca.factura_documento.url if ca.factura_documento else None
                })
            page_obj_accesorio = funcion_paginas_varias(req,res_facturas)

            return render(req,"perfil_administrativo/ventas/ventas.html",{"page_obj":page_obj,"page_objAccs":page_obj_accesorio})
    except Exception as e:
        pass

def buscar_venta_cliente(req):
    # try:
            nombre = req.GET.get('nombre').capitalize()
            apellido = req.GET.get('apellido').capitalize()

            clientes = Cliente.objects.filter(
            nombre__icontains = nombre,
            apellido__icontains = apellido
            )

            for cliente in clientes:
                resultados_motos = (
                    ComprasVentas.objects
                    .filter(tipo='V',cliente=cliente)
                    .select_related('moto','cliente')
                    .values(
                        'id',
                        'moto__marca', 
                        'moto__modelo', 
                        'fecha_compra', 
                        'cliente__nombre',
                        'cliente__apellido',
                        'cliente__id'
                    ).order_by('-fecha_compra')
                )

                res_documentacion = []
                for resultado in resultados_motos:
                    cv = ComprasVentas.objects.get(id=resultado['id'])
                    res_documentacion.append({
                    'moto': resultado
                })
                    
                resultados_accesorios = (
                    ClienteAccesorio.objects
                    .filter(cliente=cliente)
                    .select_related('accesorio', 'cliente')
                    .values(
                        'id',
                        'accesorio__tipo',
                        'accesorio__marca',
                        'accesorio__modelo',
                        'fecha_compra',
                        'cliente__nombre',
                        'cliente__apellido',
                        'cliente__id'
                    
                    ).order_by('-fecha_compra')
                )
                res_facturas = []
                for resultado_accesorio in resultados_accesorios:
                        ca = ClienteAccesorio.objects.get(id=resultado_accesorio['id'])
                        res_facturas.append({
                        'accesorio': resultado_accesorio,
                        'factura_documento': ca.factura_documento.url if ca.factura_documento else None
                    })

            page_obj = funcion_paginas_varias(req,res_documentacion)
            page_obj_accesorio = funcion_paginas_varias(req,res_facturas)

            return render(req,"perfil_administrativo/ventas/ventas.html",{"page_obj":page_obj,"page_objAccs":page_obj_accesorio})
    # except Exception as e:
    #     pass

def buscar_venta_fecha(req):
    try:
            f_venta_str = req.GET.get('f_venta')  # Cambiado a paréntesis
            f_venta = datetime.strptime(f_venta_str, '%Y-%m-%d').date() if f_venta_str else None
            resultados_motos = (
                ComprasVentas.objects
                .filter(tipo='V',fecha_compra=f_venta)
                .select_related('moto','cliente')
                .values(
                    'id',
                    'moto__marca', 
                    'moto__modelo', 
                    'fecha_compra', 
                    'cliente__nombre',
                    'cliente__apellido',
                    'cliente__id'
                ).order_by('-fecha_compra')
            )

            res_documentacion = []
            for resultado in resultados_motos:
                cv = ComprasVentas.objects.get(id=resultado['id'])
                res_documentacion.append({
                'moto': resultado
            })

            page_obj = funcion_paginas_varias(req,res_documentacion)
            
            resultados_accesorios = (
                ClienteAccesorio.objects
                .filter(fecha_compra=f_venta)
                .select_related('accesorio', 'cliente')
                .values(
                    'id',
                    'accesorio__tipo',
                    'accesorio__marca',
                    'accesorio__modelo',
                    'fecha_compra',
                    'cliente__nombre',
                    'cliente__apellido',
                    'cliente__id'
                
                ).order_by('-fecha_compra')
            )
            res_facturas = []
            for resultado_accesorio in resultados_accesorios:
                    ca = ClienteAccesorio.objects.get(id=resultado_accesorio['id'])
                    res_facturas.append({
                    'accesorio': resultado_accesorio,
                    'factura_documento': ca.factura_documento.url if ca.factura_documento else None
                })
            page_obj_accesorio = funcion_paginas_varias(req,res_facturas)

            return render(req,"perfil_administrativo/ventas/ventas.html",{"page_obj":page_obj,"page_objAccs":page_obj_accesorio})
    
    except Exception as e:
        pass


def buscar_venta_marca_modelo_moto(req):
    # try:
            marca = req.GET.get('marca_moto', '')  # Cambiado a paréntesis
            modelo = req.GET.get('modelo_moto', '')
            resultados_motos = (
                ComprasVentas.objects
                .filter(tipo='V', moto__marca__contains=marca, moto__modelo__contains=modelo)
                .select_related('moto','cliente')
                .values(
                    'id',
                    'moto__marca', 
                    'moto__modelo', 
                    'fecha_compra', 
                    'cliente__nombre',
                    'cliente__apellido',
                    'cliente__id'
                ).order_by('-fecha_compra')
            )

            res_documentacion = []
            for resultado in resultados_motos:
                cv = ComprasVentas.objects.get(id=resultado['id'])
                res_documentacion.append({
                'moto': resultado
            })

            page_obj = funcion_paginas_varias(req,res_documentacion)
            
            resultados_accesorios = (
                ClienteAccesorio.objects
                .all()
                .select_related('accesorio', 'cliente')
                .values(
                    'id',
                    'accesorio__tipo',
                    'accesorio__marca',
                    'accesorio__modelo',
                    'fecha_compra',
                    'cliente__nombre',
                    'cliente__apellido',
                    'cliente__id'
                
                ).order_by('-fecha_compra')
            )
            res_facturas = []
            for resultado_accesorio in resultados_accesorios:
                    ca = ClienteAccesorio.objects.get(id=resultado_accesorio['id'])
                    res_facturas.append({
                    'accesorio': resultado_accesorio,
                    'factura_documento': ca.factura_documento.url if ca.factura_documento else None
                })
            page_obj_accesorio = funcion_paginas_varias(req,res_facturas)

            return render(req,"perfil_administrativo/ventas/ventas.html",{"page_obj":page_obj,"page_objAccs":page_obj_accesorio})
    
    # except Exception as e:
    #     pass

def buscar_venta_accesorio(req):
    # try:
            
            resultados_motos = (
                ComprasVentas.objects
                .all()
                .select_related('moto','cliente')
                .values(
                    'id',
                    'moto__marca', 
                    'moto__modelo', 
                    'fecha_compra', 
                    'cliente__nombre',
                    'cliente__apellido',
                    'cliente__id'
                ).order_by('-fecha_compra')
            )

            res_documentacion = []
            for resultado in resultados_motos:
                cv = ComprasVentas.objects.get(id=resultado['id'])
                res_documentacion.append({
                'moto': resultado
            })

            page_obj = funcion_paginas_varias(req,res_documentacion)

            tipo = req.GET.get('tipo_accesorio', '') or ''
            marca = req.GET.get('marca_accesorio', '') or ''
            modelo = req.GET.get('modelo_accesorio', '') or ''

            # Filtrar accesorios
            resultados_accesorios = (
                ClienteAccesorio.objects
                .filter(
                    accesorio__tipo__icontains=tipo,
                    accesorio__marca__icontains=marca,
                    accesorio__modelo__icontains=modelo
                )
                .select_related('accesorio', 'cliente')
                .values(
                    'id',
                    'accesorio__tipo',
                    'accesorio__marca',
                    'accesorio__modelo',
                    'fecha_compra',
                    'cliente__nombre',
                    'cliente__apellido',
                    'cliente__id'
                ).order_by('-fecha_compra')
            )

            # Procesar facturas y documentos
            res_facturas = [
                {
                    'accesorio': resultado,
                    'factura_documento': ClienteAccesorio.objects.get(id=resultado['id']).factura_documento.url if ClienteAccesorio.objects.get(id=resultado['id']).factura_documento else None
                }
                for resultado in resultados_accesorios
            ]

            # Paginación
            paginator = Paginator(res_facturas, 10)  # 10 registros por página
            page_number = req.GET.get('page')
            page_obj_accesorio = paginator.get_page(page_number)

            return render(req,"perfil_administrativo/ventas/ventas.html",{"page_obj":page_obj,"page_objAccs":page_obj_accesorio})
    
    # except Exception as e:
    #     pass


def buscar_venta_num_motor_moto(req):
    # try:
            num_motor = req.GET.get('num_motor', '')  # Cambiado a paréntesis
            # modelo = req.GET.get('modelo_moto', '')
            resultados_motos = (
                ComprasVentas.objects
                .filter(tipo='V', moto__num_motor=num_motor)
                .select_related('moto','cliente')
                .values(
                    'id',
                    'moto__marca', 
                    'moto__modelo', 
                    'fecha_compra', 
                    'cliente__nombre',
                    'cliente__apellido',
                    'cliente__id'
                ).order_by('-fecha_compra')
            )

            res_documentacion = []
            for resultado in resultados_motos:
                cv = ComprasVentas.objects.get(id=resultado['id'])
                res_documentacion.append({
                'moto': resultado
            })

            page_obj = funcion_paginas_varias(req,res_documentacion)
            
            resultados_accesorios = (
                ClienteAccesorio.objects
                .all()
                .select_related('accesorio', 'cliente')
                .values(
                    'id',
                    'accesorio__tipo',
                    'accesorio__marca',
                    'accesorio__modelo',
                    'fecha_compra',
                    'cliente__nombre',
                    'cliente__apellido',
                    'cliente__id'
                
                ).order_by('-fecha_compra')
            )
            res_facturas = []
            for resultado_accesorio in resultados_accesorios:
                    ca = ClienteAccesorio.objects.get(id=resultado_accesorio['id'])
                    res_facturas.append({
                    'accesorio': resultado_accesorio,
                    'factura_documento': ca.factura_documento.url if ca.factura_documento else None
                })
            page_obj_accesorio = funcion_paginas_varias(req,res_facturas)

            return render(req,"perfil_administrativo/ventas/ventas.html",{"page_obj":page_obj,"page_objAccs":page_obj_accesorio})
    
    # except Exception as e:
    #     pass

def buscar_venta_num_chasis_moto(req):
    # try:
            marca = req.GET.get('marca', '')  # Cambiado a paréntesis
            
            resultados_motos = (
                ComprasVentas.objects
                .filter(tipo='V', moto__marca=marca)
                .select_related('moto','cliente')
                .values(
                    'id',
                    'moto__marca', 
                    'moto__modelo', 
                    'fecha_compra', 
                    'cliente__nombre',
                    'cliente__apellido',
                    'cliente__id'
                ).order_by('-fecha_compra')
            )

            res_documentacion = []
            for resultado in resultados_motos:
                cv = ComprasVentas.objects.get(id=resultado['id'])
                res_documentacion.append({
                'moto': resultado
            })

            page_obj = funcion_paginas_varias(req,res_documentacion)
            
            resultados_accesorios = (
                ClienteAccesorio.objects
                .all()
                .select_related('accesorio', 'cliente')
                .values(
                    'id',
                    'accesorio__tipo',
                    'accesorio__marca',
                    'accesorio__modelo',
                    'fecha_compra',
                    'cliente__nombre',
                    'cliente__apellido',
                    'cliente__id'
                
                ).order_by('-fecha_compra')
            )
            res_facturas = []
            for resultado_accesorio in resultados_accesorios:
                    ca = ClienteAccesorio.objects.get(id=resultado_accesorio['id'])
                    res_facturas.append({
                    'accesorio': resultado_accesorio,
                    'factura_documento': ca.factura_documento.url if ca.factura_documento else None
                })
            page_obj_accesorio = funcion_paginas_varias(req,res_facturas)

            return render(req,"perfil_administrativo/ventas/ventas.html",{"page_obj":page_obj,"page_objAccs":page_obj_accesorio})
    
    # except Exception as e:
    #     pass



def obtener_cuotas_json(req, id_cv):
    try:
        # c_v = ComprasVentas.objects.get(id=id_cv)
        cuotas = CuotasMoto.objects.filter(venta_id=id_cv)
        
        # Serializar los datos
        cuotas_data = [
            {
                "fecha_vencimiento": cuota.fecha_prox_pago.strftime('%Y-%m-%d'),
                "monto": float(cuota.valor_pago_pesos),
            }
            for cuota in cuotas
        ]

        return JsonResponse({"cuotas": cuotas_data}, status=200)
    except ComprasVentas.DoesNotExist:
        return JsonResponse({"error": "Compra/Venta no encontrada"}, status=404)
    

# @admin_required
# def detalles_cuotas(req,id_cv):
#     # try: 
#         c_v = ComprasVentas.objects.get(id=id_cv)
#         page_obj = obtener_compras_motos(req,id_cv)
#         cuotas = CuotasMoto.objects.filter(venta_id=id_cv)
#         if cuotas:
#             ult_cuota = CuotasMoto.objects.filter(venta_id=id_cv).latest('id')
#         else: 
#             ult_cuota = None
#         # Serializar los datos
#         cuotas_data = [
#             {
#                 "fecha_vencimiento": cuota.fecha_prox_pago.strftime('%Y-%m-%d'),
#                 "moneda":cuota.moneda,
#                 "monto": float(cuota.valor_pago_pesos) if cuota.moneda == "pesos" else float(cuota.valor_pago_dolares),
#             }
#             for cuota in cuotas
#         ]
#         cuotas_json = json.dumps(cuotas_data)
#         moto = Moto.objects.get(id=c_v.moto_id)

#         data_fin = []
#         for cuota in cuotas:
#             # Obtiene el primer financiamiento relacionado con la cuota
#             financiamiento = Financiamientos.objects.filter(id=cuota.financiamiento_id).first()
#             if financiamiento:  # Verifica si existe al menos un registro
#                 # Convierte la fecha a un formato legible
#                 fecha_str = financiamiento.fecha.strftime("%d/%m/%Y")  # Ajusta el formato si es necesario
#                 actual = "Actual" if financiamiento.actual else ""
#                 data_fin.append({
#                     "financiamientos": f"{fecha_str} {actual}".strip(),  # Quita espacios en blanco si no hay "Actual"
#                     "id": financiamiento.id
#                 })
#             else:
#                 data_fin.append({
#                     "financiamientos": None
#                 })


#         # financiamiento = Financiamientos.objects.filter(venta_id=id_cv).first()
#         # if financiamiento:
#         #     if financiamiento.moneda_cuota == "Pesos":
#         #         moneda = "$"
#         #     else:
#         #         moneda = "U$S"
#         #     txt_financiamiento = f"{str(financiamiento.cantidad_cuotas)} x {moneda} {str(financiamiento.valor_cuota)}"
#         # else:
#         #     txt_financiamiento = None
        
#         producto = f"{moto.marca} {moto.modelo}"
#         # precio_dolar = PrecioDolar.objects.get(id=1)
#         # if moto.moneda_precio == "Pesos":
#         #     precio_final_pesos = moto.precio_final
#         #     precio_final_dolares = (moto.precio_final / precio_dolar.precio_dolar_tienda)
#         # else:
#         #     precio_final_pesos = (moto.precio_final * precio_dolar.precio_dolar_tienda)
#         #     precio_final_dolares = moto.precio_final
#         return render(req,"perfil_administrativo/ventas/detalles_cuotas.html",{'cuotas_json': cuotas_json,
#                                                                                "page_obj":page_obj,
#                                                                                "id_cv":id_cv,
#                                                                                "pago_acordado":c_v.forma_de_pago,
#                                                                                "id_cliente":c_v.cliente_id,
#                                                                                "producto":producto,
#                                                                                "precio_inicial":moto.precio,
#                                                                                "precio_final":moto.precio_final,
#                                                                             "financiamientos":data_fin
#                                                                             #    "cant_restante_pesos":int(ult_cuota.cant_restante_pesos) if ult_cuota else int(precio_final_pesos),
#                                                                             #    "cant_restante_dolares":int(ult_cuota.cant_restante_dolares) if ult_cuota else int(precio_final_dolares),
#                                                                             #    "financiamiento":txt_financiamiento if txt_financiamiento else "Sin financiamiento"
#                                                                                })
#     # except Exception as e:
#     #     return render(req,"perfil_administrativo/ventas/detalles_cuotas.html",{"error_message":e})
def detalles_cuotas(req,id_cv):
    contexto = funcion_detalles_cuotas(req,id_cv,False,None)
    return render(req,"perfil_administrativo/ventas/detalles_cuotas.html",contexto)

def buscar_pagos_por_refinanciamiento(req):
    # try:
        # id = int(req.POST['ref_id'])
        # fin = Financiamientos.objects.filter(id=id).first()
        # datos = obtener_detalles_cuotas_comunes(fin.venta_id)
        # fin_inicial = Financiamientos.objects.filter(venta_id=fin.venta_id,inicial=1).first()
        # moneda_ini = "$" if fin_inicial.moneda_cuota == "Pesos" else "U$S"
        # fin_ini = f"{str(fin_inicial.cantidad_cuotas)} x {moneda_ini} {str(fin_inicial.valor_cuota)}"#
        
        # page_obj = obtener_detalles_cuotas_financiamiento(req,id)
        # moneda = "$" if fin.moneda_cuota == "Pesos" else "U$S"
        # financiamiento = f"{str(fin.cantidad_cuotas)} x {moneda} {str(fin.valor_cuota)}"
        # ult_cuota = CuotasMoto.objects.filter(venta_id=fin.venta_id).latest('id')
        # precio_dolar = PrecioDolar.objects.get(id=1)
        
        # cv = ComprasVentas.objects.get(id=fin.venta_id)
        # if ult_cuota:
        #     cant_restante_pesos = ult_cuota.cant_restante_pesos
        #     cant_restante_dolares = ult_cuota.cant_restante_dolares
        # else:
        #     moto = ComprasVentas.objects.get(id=cv.moto_id)
            
        #     if moto.moneda_precio == "Pesos":
        #         cant_restante_pesos = moto.precio_final
        #         cant_restante_dolares = (moto.precio_final / precio_dolar.precio_dolar_tienda)
        #     else:
        #         cant_restante_pesos = (moto.precio_final * precio_dolar.precio_dolar_tienda)
        #         cant_restante_dolares = moto.precio_final
        
        # precio = float(precio_dolar.precio_dolar_tienda) if precio_dolar.precio_dolar_tienda else 0
        # data_jason = json_para_resumen_pagos(moto,cv.id)
        # contexto = {"id_cv":fin.venta_id,
        #             "producto":datos[0],
        #             "precio_inicial":datos[1],
        #             "precio_final":datos[2],
        #             "pago_acordado":datos[3],
        #             "primeros_pagos":datos[4],
        #             "financiamiento":datos[5],
        #             "financiamientos":datos[6],
        #             "page_obj":page_obj,
        #             "financiacion_info":financiamiento,
        #             "financiacion_inicial":fin_ini,
        #             "fecha_financiacion":fin.fecha,
        #             "boton_pagar":True if fin.actual else False,
        #             "fin_actual":True,
        #             "cant_restante_dolares":int(cant_restante_dolares),
        #             "cant_restante_pesos":int(cant_restante_pesos),
        #             "precio_dolar":precio,
        #             "id_cliente":cv.cliente_id,
        #             'cuotas_json': data_jason[0],
        #             'cliente_json': data_jason[1],
        #             "detalle_json":data_jason[2],
        #             "p_pagos_json":data_jason[3],
        #             # "id_f":id_f, 
        #             # "mon_pago":mon_pago, MONEDA DEL FINANCIAMIENTO ACTUAL PARA CARGAR AUTOMATICAMENTE LA MONEDA EN EL PAGO
        #             "fin_pagos_json":data_jason[4]}
        id = int(req.POST['ref_id'])
        fin = Financiamientos.objects.filter(id=id).first()
        contexto = funcion_detalles_cuotas(req,fin.venta_id,True,id)
        return render(req,"perfil_administrativo/ventas/detalles_cuotas.html",contexto)

    # except Exception as e:
    #     pass


def refinanciar_pagos(req,id_cv):
    try:
        if req.POST['moneda_refinanciacion'] == "Pesos":
            valor_cuota = req.POST['valor_cuota_pesos']
        else:
            valor_cuota = req.POST['valor_cuota_dolares']
        
        quincena = 1 if req.POST['quincena_mensual'] == "Quincena" else 0
        alta_financiamientos(req.POST['recargo_porcentaje_coma'],req.POST['cant_cuotas'],valor_cuota,req.POST['moneda_refinanciacion'],1,id_cv,0,quincena)
        messages.success(req, "Financiamiento ingresado con éxito")
        return redirect(f"{reverse('DetallesCuotas',kwargs={'id_cv':id_cv})}?comprobante_url={None}")
    except Exception as e:
        return render(req,"perfil_administrativo/ventas/detalles_cuotas.html",{'id_cv':id_cv})

def baja_financiamiento(req,id_f,id_cv):
    # try:
        if req.method == "POST":
            validar_caja = validar_caja_abierta()
            if not validar_caja:
                return render(req, "perfil_administrativo/ventas/baja_financiamiento.html", {"error_message":"No existe caja del día abierta.","id_cv":id_cv})
            else:
                fin_cuotas = CuotasFinanciacion.objects.filter(financiamiento_id=id_f)
                total_dinero = 0
                dolar = PrecioDolar.objects.get(id=1)
                precio_dolar = float(dolar.precio_dolar_tienda)
                for fin in fin_cuotas:
                    cuota = CuotasMoto.objects.filter(id=fin.cuota_id).first()
                    if cuota.moneda == "Pesos":
                        total_dinero = total_dinero + float(cuota.valor_pago_pesos)
                    else:
                        valor_pesos = float(cuota.valor_pago_dolares) * precio_dolar 
                        total_dinero = total_dinero + valor_pesos
                    if cuota.metodo_pago == "Fondos":
                        dolar = PrecioDolar.objects.get(id=1)
                        precio_dolar = float(dolar.precio_dolar_tienda)
                        venta = ComprasVentas.objects.get(id=id_cv)
                        fondos_cliente = ClienteFondos.objects.filter(cliente=venta.cliente).latest('id')
                        if cuota.moneda == "Pesos":
                            entrega_pesos = float(cuota.valor_pago_pesos)
                            entrega_dolares = entrega_pesos / precio_dolar
                        else:
                            entrega_dolares = float(cuota.valor_pago_dolares)
                            entrega_pesos = entrega_dolares * precio_dolar
                        
                        total_pesos = float(fondos_cliente.total_pesos) + entrega_pesos
                        total_dolares = float(fondos_cliente.total_dolares) + entrega_dolares
                        nuevo_fondo = ClienteFondos(
                            moneda = cuota.moneda,
                            fecha = datetime.now().date(),
                            cliente = fondos_cliente.cliente,
                            ingreso_pesos = entrega_pesos,
                            ingreso_dolares = entrega_dolares,
                            total_pesos = total_pesos,
                            total_dolares = total_dolares,
                            comprobante = None,
                            metodo = "----------",
                            tipo = "Ingreso"
                        )   
                        nuevo_fondo.save()
                    mov_pago = MovimientoPagoMoto.objects.filter(pago=cuota).first()
                    if mov_pago:
                        mov = Movimientos.objects.get(id=mov_pago.movimiento_id)
                        if mov.metodo == "Efectivo":
                            caja = Caja.objects.latest('id')
                            if mov.moneda == "Pesos":
                                monto_a_quitar = float(mov.monto)
                            else:
                                dolar = PrecioDolar.objects.get(id=1)
                                precio_dolar = float(dolar.precio_dolar_tienda)
                                monto_a_quitar = float(mov.monto) * precio_dolar
                            caja.depositos = float(caja.depositos) - monto_a_quitar
                            # if mov.tipo == "Ingreso" or mov.tipo == "Ingreso extra":
                            #     nueva_diferencia = float(caja.diferencia) - monto_a_quitar
                            # else:
                            #     nueva_diferencia = float(caja.diferencia) + monto_a_quitar
                            
                            # caja.diferencia = nueva_diferencia
                            caja.save()
                        
                        mov_pago.delete()
                        mov.delete()

                    fin.delete()
                    cuota.delete()
                financiamiento = Financiamientos.objects.get(id=id_f)
                financiamiento.delete()
            # fin_actual = Financiamientos.objects.filter(venta_id=id_cv,inicial=0).first()
            # if fin_actual:
            #     fin_actual = Financiamientos.objects.filter(venta_id=id_cv,inicial=0).latest('id')
            #     fin_actual.actual = 1
            #     fin_actual.save()
            return render(req, "perfil_administrativo/ventas/baja_financiamiento.html", {"message":"Financiamiento borrado con éxito","id_cv":id_cv})
        else:
            return render(req,"perfil_administrativo/ventas/baja_financiamiento.html",{"id_cv":id_cv})
    # except Exception as e:
    #     return render(req, "perfil_administrativo/ventas/baja_financiamiento.html", {"error_message":str(e),"id_cv":id_cv})


def alta_pago_cuota(req,id_cv):
    
    # page_obj = obtener_detalles_cuotas_financiamiento(req,id_cv)
    try:
        comprobante = req.FILES.get('comprobante_pago')
        moneda = req.POST['moneda_entrega']
        # caja = Caja.objects.filter(estado="Abierto").first()
        forma_pago = req.POST['forma_pago']
        fecha_proximo_pago = req.POST['f_prox_pago']
        observaciones_pago = req.POST['observaciones_pago']
        cv = ComprasVentas.objects.get(id=id_cv)
        dolar = PrecioDolar.objects.get(id=1)
        precio_dolar = dolar.precio_dolar_tienda
        # recargo = req.POST['recargo']
        total = req.POST['valor_a_pagar']
        existe_cuota = CuotasMoto.objects.filter(venta_id=id_cv).exists()
        moto = Moto.objects.get(id=cv.moto_id)
        valores = valores_compras(existe_cuota,moneda,req.POST['valor_a_pagar'],id_cv,moto,"Moto",precio_dolar)
        fin_actual = Financiamientos.objects.filter(venta_id=id_cv,actual=1).first()
        
        validar_precio = validar_entrega_menor_precio(moneda,req.POST['valor_a_pagar'],moto,precio_dolar,"Moto",existe_cuota,id_cv)
        validar_fecha_proximo_pago = datetime.strptime(fecha_proximo_pago, '%Y-%m-%d')
        fecha_actual = datetime.now().date()
        if forma_pago == "Fondos":
            fondos_cliente = ClienteFondos.objects.filter(cliente=cv.cliente).first()
            if fondos_cliente:
                ult_fondo = ClienteFondos.objects.filter(cliente=cv.cliente).latest('id')
                if moneda == "Pesos":
                    sin_fondos = float(total) > float(ult_fondo.total_pesos)
                else:
                    sin_fondos = float(total) > float(ult_fondo.total_dolares)
            else:
                sin_fondos = True
        else:
            sin_fondos = False
        validar_caja = validar_caja_abierta()
        if validar_precio:
            #return render(req,"perfil_administrativo/ventas/detalles_cuotas.html",{"error_message":validar_precio,"id_cv":id_cv,"page_obj": page_obj,"id_cliente":cv.cliente_id})
            messages.error(req, validar_precio)
            return redirect(f"{reverse('DetallesCuotas', kwargs={'id_cv': id_cv})}?comprobante_url={None}")
        elif validar_fecha_proximo_pago.date() < fecha_actual:
            #return render(req,"perfil_administrativo/ventas/detalles_cuotas.html",{"error_message":"La fecha del próximo pago no debe ser anterior a la actual","id_cv":id_cv,"page_obj": page_obj,"id_cliente":cv.cliente_id})
            messages.error(req, "La fecha del próximo pago no debe ser anterior a la actual")
            return redirect(f"{reverse('DetallesCuotas', kwargs={'id_cv': id_cv})}?comprobante_url={None}")
        # elif int(fin_actual.valor_cuota) != int(req.POST['valor_a_pagar']):
        #     #return render(req,"perfil_administrativo/ventas/detalles_cuotas.html",{"error_message":"El valor ingresado de la cantidad a pagar es incorrecto","id_cv":id_cv,"page_obj": page_obj,"id_cliente":cv.cliente_id})
        #     messages.error(req, "El valor ingresado de la cantidad a pagar es incorrecto")
        #     return redirect(f"{reverse('DetallesCuotas', kwargs={'id_cv': id_cv})}?comprobante_url={None}")
        elif not validar_caja:
            messages.error(req, "No existe caja del día abierta.")
            return redirect(f"{reverse('DetallesCuotas', kwargs={'id_cv': id_cv})}?comprobante_url={None}")
        elif sin_fondos:
            messages.error(req, "El cliente no dispone de suficientes fondos para realizar el pago correspondiente.")
            return redirect(f"{reverse('DetallesCuotas',kwargs={'id_cv':id_cv})}?comprobante_url={None}")
        else:
            
            
            cant_cuotas = CuotasFinanciacion.objects.filter(financiamiento_id=fin_actual.id).count() 
            num_cuota_actual = cant_cuotas + 1
            if fin_actual.quincena:
                if num_cuota_actual % 2 == 0:
                    num_cuota_actual = int(num_cuota_actual / 2)
                    cuota = "Cuota " + str(num_cuota_actual) + " Segunda Quincena"
                else:
                    num_cuota_actual = int((num_cuota_actual + 1) / 2)
                    cuota = "Cuota " + str(num_cuota_actual) + " Primer Quincena"
            else:
                cuota = "Cuota " + str(num_cuota_actual)   
            if cant_cuotas == 0:
                if fin_actual.moneda_cuota == "Pesos": #DUDA
                    resto_pesos = (int(fin_actual.cantidad_cuotas) * int(fin_actual.valor_cuota)) - int(req.POST['valor_a_pagar'])
                    resto_dolares = int(int(resto_pesos) / float(precio_dolar))
                else:
                    resto_dolares = (int(fin_actual.cantidad_cuotas) * int(fin_actual.valor_cuota)) - int(req.POST['valor_a_pagar'])
                    resto_pesos = int(int(resto_dolares) * float(precio_dolar))
                alta = alta_cuota_funcion(req,fecha_proximo_pago,id_cv,resto_dolares,resto_pesos,moneda,observaciones_pago,precio_dolar,valores[3],valores[2],comprobante,forma_pago,True,cuota)
            else:
                alta = alta_cuota_funcion(req,fecha_proximo_pago,id_cv,valores[0],valores[1],moneda,observaciones_pago,precio_dolar,valores[3],valores[2],comprobante,forma_pago,True,cuota)
            if alta.comprobante_pago:
                comprobante_url = alta
            else:
                comprobante_url = None
            
            if forma_pago == "Fondos":
                ult_fondo = ClienteFondos.objects.filter(cliente=cv.cliente).latest('id')
                total_pesos = float(ult_fondo.total_pesos) - float(valores[2])
                total_dolares = float(ult_fondo.total_dolares) - float(valores[3]) 
                ingreso_pesos = 0 - float(valores[2])
                ingreso_dolares = 0 - float(valores[3]) 
                retiro_fondos = ClienteFondos(
                    moneda = moneda,
                    fecha = fecha_actual,
                    cliente = cv.cliente,
                    ingreso_pesos = ingreso_pesos,
                    ingreso_dolares = ingreso_dolares,
                    total_pesos = total_pesos,
                    total_dolares = total_dolares,
                    comprobante = None,
                    metodo = "----------",
                    tipo = "Retiro"
                )
                retiro_fondos.save()
            # if caja:
            if forma_pago != "Fondos":
                if cv.fecha_compra == fecha_actual:
                    tipo = "Ingreso"
                else:
                    tipo = "Ingreso extra"
                movimiento_caja_por_pago(req,float(total),id_cv,moneda,forma_pago,tipo,alta,"moto")
            messages.success(req, "Pago ingresado con éxito.")
            return redirect(f"{reverse('DetallesCuotas',kwargs={'id_cv':id_cv})}?comprobante_url={comprobante_url}")
    except Exception as e:
        messages.error(req, "Algo salió mal: " +str(e))
        return redirect(f"{reverse('DetallesCuotas', kwargs={'id_cv': id_cv})}?comprobante_url={None}")


@admin_required
def alta_pago(req,id_cv):
    #
    # page_obj = obtener_compras_motos(req,id_cv)
    # try:
        comprobante = req.FILES.get('comprobante_pago')
        moneda = req.POST['moneda_entrega']
        # caja = Caja.objects.filter(estado="Abierto").first()
        forma_pago = req.POST['forma_pago']
        fecha_proximo_pago = req.POST['f_prox_pago']
        observaciones_pago = req.POST['observaciones_pago']
        cv = ComprasVentas.objects.get(id=id_cv)
        dolar = PrecioDolar.objects.get(id=1)
        precio_dolar = dolar.precio_dolar_tienda
        # recargo = req.POST['recargo']
        total = req.POST['valor_a_pagar']
        existe_cuota = CuotasMoto.objects.filter(venta_id=id_cv).exists()
        if existe_cuota:
            cuota = CuotasMoto.objects.filter(venta_id=id_cv).latest('id')
        else:
            cuota = None
        moto = Moto.objects.get(id=cv.moto_id)
        valores = valores_compras(existe_cuota,moneda,req.POST['valor_a_pagar'],id_cv,moto,"Moto",precio_dolar)
        
        validar_precio = validar_entrega_menor_precio(moneda,req.POST['valor_a_pagar'],moto,precio_dolar,"Moto",cuota,id_cv)
        # validar_fecha_proximo_pago = datetime.strptime(fecha_proximo_pago, '%Y-%m-%d')
        fecha_actual = datetime.today().date()
        if fecha_proximo_pago:
            validar_fecha_proximo_pago = datetime.strptime(fecha_proximo_pago, '%Y-%m-%d').date()
        else:
            validar_fecha_proximo_pago = None

        if forma_pago == "Fondos":
            fondos_cliente = ClienteFondos.objects.filter(cliente=cv.cliente).first()
            if fondos_cliente:
                ult_fondo = ClienteFondos.objects.filter(cliente=cv.cliente).latest('id')
                if moneda == "Pesos":
                    sin_fondos = float(total) > float(ult_fondo.total_pesos)
                else:
                    sin_fondos = float(total) > float(ult_fondo.total_dolares)
            else:
                sin_fondos = True
        else:
            sin_fondos = False
        validar_caja = validar_caja_abierta()
        if validar_precio:
            # return render(req,"perfil_administrativo/ventas/detalles_cuotas.html",{"error_message":validar_precio,"id_cv":id_cv,"page_obj": page_obj,"id_cliente":cv.cliente_id})
            messages.error(req, f"{validar_precio}")
            return redirect(f"{reverse('DetallesCuotas',kwargs={'id_cv':id_cv})}?comprobante_url={None}")
        elif fecha_proximo_pago and validar_fecha_proximo_pago < fecha_actual:
            # print(fecha_actual)
            # print(validar_fecha_proximo_pago)
            # if validar_fecha_proximo_pago < fecha_actual:
            # return render(req,"perfil_administrativo/ventas/detalles_cuotas.html",{"error_message":"La fecha del próximo pago no debe ser anterior a la actual","id_cv":id_cv,"page_obj": page_obj,"id_cliente":cv.cliente_id})
                messages.error(req, "La fecha del próximo pago no debe ser anterior a la actual.")
                return redirect(f"{reverse('DetallesCuotas',kwargs={'id_cv':id_cv})}?comprobante_url={None}")
        elif not validar_caja:
            messages.error(req, "No existe caja del día abierta.")
            return redirect(f"{reverse('DetallesCuotas',kwargs={'id_cv':id_cv})}?comprobante_url={None}")
        elif sin_fondos:
            messages.error(req, "El cliente no dispone de suficientes fondos para realizar el pago correspondiente.")
            return redirect(f"{reverse('DetallesCuotas',kwargs={'id_cv':id_cv})}?comprobante_url={None}")
        else:
            
            alta = alta_cuota_funcion(req,fecha_proximo_pago,id_cv,valores[0],valores[1],moneda,observaciones_pago,precio_dolar,valores[3],valores[2],comprobante,forma_pago,False,req.POST['pago_a_realizar'])
            if alta.comprobante_pago:
                comprobante_url = alta
            else:
                comprobante_url = None
            ult_financiamiento = Financiamientos.objects.filter(venta_id=id_cv,actual=1).first()
            if ult_financiamiento:
                #LA ULTIMA REFINANCIACION LA DESACTIVA PARA QUE NO PUEDAN AGREGARSE MAS PAGOS EN LA MISMA
                ult_financiamiento.actual = 0
                ult_financiamiento.save()
            
            if forma_pago == "Fondos":
                ult_fondo = ClienteFondos.objects.filter(cliente=cv.cliente).latest('id')
                total_pesos = float(ult_fondo.total_pesos) - float(valores[2])
                total_dolares = float(ult_fondo.total_dolares) - float(valores[3]) 
                ingreso_pesos = 0 - float(valores[2])
                ingreso_dolares = 0 - float(valores[3]) 
                retiro_fondos = ClienteFondos(
                    moneda = moneda,
                    fecha = fecha_actual,
                    cliente = cv.cliente,
                    ingreso_pesos = ingreso_pesos,
                    ingreso_dolares = ingreso_dolares,
                    total_pesos = total_pesos,
                    total_dolares = total_dolares,
                    comprobante = None,
                    metodo = "----------",
                    tipo = "Retiro"
                )
                retiro_fondos.save()
            # if caja:
            if forma_pago != "Fondos":
                if cv.fecha_compra == fecha_actual:
                    tipo = "Ingreso"
                else:
                    tipo = "Ingreso extra"
                #REGISTRA EL MOVIMIENTO UNICAMENTE SI NO SE PAGA CON LOS FONDOS YA QUE ESTOS FUERON REGISTRADOS EN EL BALANCE/CAJA AL MOMENTO DE INGRESARLOS
                movimiento_caja_por_pago(req,float(total),id_cv,moneda,forma_pago,tipo,alta,"moto") 
            messages.success(req, "Pago ingresado con éxito, se requiere refinanciar.")
            return redirect(f"{reverse('DetallesCuotas',kwargs={'id_cv':id_cv})}?comprobante_url={comprobante_url}")
    # except Exception as e:
    #     # return render(req,"perfil_administrativo/ventas/detalles_cuotas.html",{"error_message":e,"id_cliente":cv.cliente_id,"page_obj":page_obj})
    #     print("ERROR")
    #     messages.error(req, f"Algo salió mal: " + str(e))
    #     return redirect(f"{reverse('DetallesCuotas',kwargs={'id_cv':id_cv})}?comprobante_url={None}")

@admin_required
def baja_pago(req,id_cm):
    try:
        cuota = CuotasMoto.objects.get(id=id_cm)
        id_cv = cuota.venta_id
        # print(id_cv)
        if req.method == "POST":
            # if cuota.metodo_pago == "Efectivo":
            validar_caja = validar_caja_abierta()
            if not validar_caja:
                return render(req, "perfil_administrativo/ventas/baja_pago.html", {"error_message":"No existe caja del día abierta.","id_cv":id_cv})
            else:
                if cuota.metodo_pago == "Fondos":
                    dolar = PrecioDolar.objects.get(id=1)
                    precio_dolar = float(dolar.precio_dolar_tienda)
                    venta = ComprasVentas.objects.get(id=id_cv)
                    fondos_cliente = ClienteFondos.objects.filter(cliente=venta.cliente).latest('id')
                    if cuota.moneda == "Pesos":
                        entrega_pesos = float(cuota.valor_pago_pesos)
                        entrega_dolares = entrega_pesos / precio_dolar
                    else:
                        entrega_dolares = float(cuota.valor_pago_dolares)
                        entrega_pesos = entrega_dolares * precio_dolar
                    
                    total_pesos = float(fondos_cliente.total_pesos) + entrega_pesos
                    total_dolares = float(fondos_cliente.total_dolares) + entrega_dolares
                    nuevo_fondo = ClienteFondos(
                        moneda = cuota.moneda,
                        fecha = datetime.now().date(),
                        cliente = fondos_cliente.cliente,
                        ingreso_pesos = entrega_pesos,
                        ingreso_dolares = entrega_dolares,
                        total_pesos = total_pesos,
                        total_dolares = total_dolares,
                        comprobante = None,
                        metodo = "----------",
                        tipo = "Ingreso"
                    )   
                    nuevo_fondo.save()
                
                
                
                    # insert_movimientos_caja("Se borra pago de moto ingresado por error","Egreso",quitar_deposito,caja.id,personal.id,0,0,None,None)
                    
                fin_actual = Financiamientos.objects.filter(venta_id=cuota.venta_id).first()
                c_f = CuotasFinanciacion.objects.filter(financiamiento_id=fin_actual.id,cuota_id=id_cm).first()
                if c_f:
                    c_f.delete()
                    
                mov_pago = MovimientoPagoMoto.objects.filter(pago=cuota).first()
                if mov_pago:
                    if cuota.metodo_pago == "Efectivo":
                        if cuota.moneda == "Pesos":
                            quitar_deposito = cuota.valor_pago_pesos
                        else:
                            dolar = PrecioDolar.objects.get(id=1)
                            precio_dolar = dolar.precio_dolar_tienda
                            quitar_deposito = cuota.valor_pago_dolares * precio_dolar
                        caja = Caja.objects.latest('id')
                        caja.depositos = caja.depositos - quitar_deposito
                        caja.save()
                    mov = Movimientos.objects.get(id=mov_pago.movimiento_id)
                    # caja = Caja.objects.latest('id')
                    # if mov.metodo == "Efectivo" and int(caja.diferencia) != 0:
                    #     if mov.moneda == "Pesos":
                    #         monto_a_quitar = float(mov.monto)
                    #     else:
                    #         dolar = PrecioDolar.objects.get(id=1)
                    #         precio_dolar = float(dolar.precio_dolar_tienda)
                    #         monto_a_quitar = float(mov.monto) * precio_dolar
                    #     if mov.tipo == "Ingreso" or mov.tipo == "Ingreso extra":
                    #         nueva_diferencia = float(caja.diferencia) - monto_a_quitar
                    #     else:
                    #         nueva_diferencia = float(caja.diferencia) + monto_a_quitar
                        
                    #     caja.diferencia = nueva_diferencia
                    #     caja.save()
                    mov_pago.delete()
                    mov.delete()
                cuota.delete()
                print("ID: " +str(mov_pago.movimiento_id)) 
                    
                return render(req, "perfil_administrativo/ventas/baja_pago.html", {"message":"Pago borrado con éxito","id_cv":id_cv})
            # messages.success(req, "Pago borrado con éxito")
            # return redirect(f"{reverse('DetallesCuotas',kwargs={'id_cv':id_cv})}?comprobante_url={None}")
        else:
            return render(req,"perfil_administrativo/ventas/baja_pago.html",{"id_cv":id_cv})
    except Exception as e:
        return render(req,"perfil_administrativo/ventas/baja_pago.html",{"error_message":e})

def baja_primeros_pagos(req,id_cm):
    try:
        cuota = CuotasMoto.objects.get(id=id_cm)
        id_cv = cuota.venta_id
        if req.method == "POST":
            validar_caja = validar_caja_abierta()
            if not validar_caja:
                return render(req,"perfil_administrativo/ventas/baja_pago.html",{"error_message":"No existe caja del día abierta."})
            else:
                if cuota.metodo_pago == "Fondos":
                    dolar = PrecioDolar.objects.get(id=1)
                    precio_dolar = float(dolar.precio_dolar_tienda)
                    venta = ComprasVentas.objects.get(id=id_cv)
                    fondos_cliente = ClienteFondos.objects.filter(cliente=venta.cliente).latest('id')
                    if cuota.moneda == "Pesos":
                        entrega_pesos = float(cuota.valor_pago_pesos)
                        entrega_dolares = entrega_pesos / precio_dolar
                    else:
                        entrega_dolares = float(cuota.valor_pago_dolares)
                        entrega_pesos = entrega_dolares * precio_dolar
                    
                    total_pesos = float(fondos_cliente.total_pesos) + entrega_pesos
                    total_dolares = float(fondos_cliente.total_dolares) + entrega_dolares
                    nuevo_fondo = ClienteFondos(
                        moneda = cuota.moneda,
                        fecha = datetime.now().date(),
                        cliente = fondos_cliente.cliente,
                        ingreso_pesos = entrega_pesos,
                        ingreso_dolares = entrega_dolares,
                        total_pesos = total_pesos,
                        total_dolares = total_dolares,
                        comprobante = None,
                        metodo = "----------",
                        tipo = "Ingreso"
                    )   
                    nuevo_fondo.save()
                     
                mov_pago = MovimientoPagoMoto.objects.filter(pago=cuota).first()
                if mov_pago:
                    mov = Movimientos.objects.get(id=mov_pago.movimiento_id)
                    caja = Caja.objects.latest('id')
                    if mov.metodo == "Efectivo":
                        if cuota.moneda == "Pesos":
                            quitar_deposito = cuota.valor_pago_pesos
                        else:
                            dolar = PrecioDolar.objects.get(id=1)
                            precio_dolar = dolar.precio_dolar_tienda
                            quitar_deposito = cuota.valor_pago_dolares * precio_dolar
                        caja.depositos = caja.depositos - quitar_deposito
                        caja.save()
                        # if mov.moneda == "Pesos":
                        #     monto_a_quitar = float(mov.monto)
                        # else:
                        #     dolar = PrecioDolar.objects.get(id=1)
                        #     precio_dolar = float(dolar.precio_dolar_tienda)
                        #     monto_a_quitar = float(mov.monto) * precio_dolar
                        # if mov.tipo == "Ingreso" or mov.tipo == "Ingreso extra":
                        #     nueva_diferencia = float(caja.diferencia) - monto_a_quitar
                        # else:
                        #     nueva_diferencia = float(caja.diferencia) + monto_a_quitar
                        
                        # caja.diferencia = nueva_diferencia
                        # caja.save()
                    mov_pago.delete()
                    mov.delete()
                cuota.delete()    
                return render(req, "perfil_administrativo/ventas/baja_pago.html", {"message":"Pago borrado con éxito","id_cv":id_cv})
        else:
            return render(req,"perfil_administrativo/ventas/baja_pago.html",{"id_cv":id_cv})
    except Exception as e:
        return render(req,"perfil_administrativo/ventas/baja_pago.html",{"error_message":e})
    
@admin_required
def reservas(req):
    try:
        resultados_motos = (
                ComprasVentas.objects
                .filter(tipo='R')
                .select_related('moto','cliente')
                .values(
                    'id',
                    'moto__id',
                    'moto__marca', 
                    'moto__modelo', 
                    'fecha_compra', 
                    'cliente__nombre',
                    'cliente__apellido'
                ).order_by('-fecha_compra')
            )

        res_documentacion = []
        for resultado in resultados_motos:
                cv = ComprasVentas.objects.get(id=resultado['id'])
                res_documentacion.append({
                'moto': resultado
            })


        page_obj = funcion_paginas_varias(req,res_documentacion)
        return render(req,"perfil_administrativo/motos/reservas.html",{"page_obj":page_obj})
    except Exception as e:
        pass

@admin_required
def form_reservar_moto(req,id_moto):
    try:
        if req.method == "POST":
            tipo_doc = req.POST['tipo_documento']
            doc = req.POST['documento']
            documento = tipo_doc + str(doc)
            cliente = Cliente.objects.filter(documento=documento).first()
            if cliente:
                tel1 = ClienteTelefono.objects.filter(principal=1,cliente_id=cliente.id).first()
                tel_1 = tel1.telefono
                tel2 = ClienteTelefono.objects.filter(principal=0,cliente_id=cliente.id).first()

                correo1 = ClienteCorreo.objects.filter(principal=1,cliente_id=cliente.id).first()
                correo2 = ClienteCorreo.objects.filter(principal=0,cliente_id=cliente.id).first()
                if tel2:
                    tel_2 = tel2.telefono
                else:
                    tel_2 = None

                if correo1:
                    c_1 = correo1.correo
                else:
                    c_1 = None
                
                if correo2:
                    c_2 = correo1.correo
                else:
                    c_2 = None
                moto = Moto.objects.get(id=id_moto)
                
                #RENDERIZAR PAPEL COMPRA-VENTA

                return render(req,"perfil_administrativo/motos/reservar_moto.html",{"datos_moto":True,
                                                                                "cliente":cliente,
                                                                                "moto":moto,
                                                                                "tel1":tel_1,
                                                                                "tel2":tel_2,
                                                                                "correo1":c_1,
                                                                                "correo2":c_2
                                                                                })
            else:
                return render(req,"perfil_administrativo/motos/reservar_moto.html",{"datos_moto":False,"error_message_cliente":"El cliente no se encuentra registrado en el sistema, para ingresarlo haga clic "})
        else:
            return render(req,"perfil_administrativo/motos/reservar_moto.html",{})
    except Exception as e:
        return render(req,"perfil_administrativo/motos/reservar_moto.html",{"error_message":e})

@admin_required
def reservar_moto(req,id_moto,id_cliente):
    try:
        forma_pago = req.POST['forma_pago_senia']
        moneda = req.POST['moneda_senia']
        cliente = Cliente.objects.get(id=id_cliente)
        moto = Moto.objects.get(id=id_moto)
        if forma_pago == "Fondos":
            fondos_cliente = ClienteFondos.objects.filter(cliente=id_cliente).first()
            if fondos_cliente:
                ult_fondo = ClienteFondos.objects.filter(cliente=id_cliente).latest('id')
                if moneda == "Pesos":
                    sin_fondos = float(req.POST['senia']) > float(ult_fondo.total_pesos)
                else:
                    sin_fondos = float(req.POST['senia']) > float(ult_fondo.total_dolares)
            else:
                sin_fondos = True
        else:
            sin_fondos = False
        validar_caja = validar_caja_abierta()
        cliente = Cliente.objects.get(id=id_cliente)
        tel1 = ClienteTelefono.objects.filter(principal=1,cliente_id=cliente.id).first()
        tel_1 = tel1.telefono
        tel2 = ClienteTelefono.objects.filter(principal=0,cliente_id=cliente.id).first()

        correo1 = ClienteCorreo.objects.filter(principal=1,cliente_id=cliente.id).first()
        correo2 = ClienteCorreo.objects.filter(principal=0,cliente_id=cliente.id).first()
        if tel2:
            tel_2 = tel2.telefono
        else:
            tel_2 = None

        if correo1:
            c_1 = correo1.correo
        else:
            c_1 = None
        
        if correo2:
            c_2 = correo1.correo
        else:
            c_2 = None
        if not validar_caja:
            # return render(req,"perfil_administrativo/motos/reservar_moto.html",{})
            return render(req,"perfil_administrativo/motos/reservar_moto.html",{"error_message":"No existe caja del día abierta.",
                                                                                "active_page":"Motos",
                                                                                    "datos_moto":True,
                                                                                    "cliente":cliente,
                                                                                    "moto":moto,
                                                                                    "tel1":tel_1,
                                                                                    "tel2":tel_2,
                                                                                    "correo1":c_1,
                                                                                    "correo2":c_2
                                                                                    })
    
        else:
            if sin_fondos:
            # return render(req,"perfil_administrativo/motos/reservar_moto.html",{"error_message":"El cliente no dispone de suficientes fondos para realizar el pago correspondiente.","active_page":"Motos"})
                return render(req,"perfil_administrativo/motos/reservar_moto.html",{"error_message":"El cliente no dispone de suficientes fondos para realizar el pago correspondiente.",
                                                                                "active_page":"Motos",
                                                                                "datos_moto":True,
                                                                                "cliente":cliente,
                                                                                "moto":moto,
                                                                                "tel1":tel_1,
                                                                                "tel2":tel_2,
                                                                                "correo1":c_1,
                                                                                "correo2":c_2
                                                                                
                                                                                })
            else:
                compras_ventas = ComprasVentas(
                    fecha_compra = datetime.now(),
                    tipo = "R",
                    cliente_id = id_cliente,
                    moto_id = id_moto
                )
                compras_ventas.save()
                dolar = PrecioDolar.objects.get(id=1)
                precio_dolar = dolar.precio_dolar_tienda
                id_cv = compras_ventas.id

                if moneda == "Pesos":
                        entrega_pesos = req.POST['senia']
                        entrega_dolares = 0
                        if moto.moneda_precio == "Pesos":
                            resto_pesos = int(moto.precio) - int(entrega_pesos)
                            resto_dolares = resto_pesos / precio_dolar
                        else:
                            resto_pesos = int((moto.precio * precio_dolar)) - int(entrega_pesos)
                            resto_dolares = resto_pesos / precio_dolar
                else:
                        entrega_pesos = 0
                        entrega_dolares = req.POST['senia']
                        if moto.moneda_precio == "Pesos":
                            resto_dolares = int((moto.precio / precio_dolar)) - int(entrega_dolares)
                            resto_pesos = resto_dolares * precio_dolar
                        else:
                            resto_dolares = int(moto.precio) - int(entrega_dolares)
                            resto_pesos = resto_dolares * precio_dolar
                entrega = req.POST['senia']
                # movimiento_caja_por_pago(req,entrega,id_cv,moneda,None,"Ingreso extra",None,"moto")
                # fecha_prox_pago = datetime.now() + relativedelta(months=1)
                pago_reserva = insert_cuotas_moto(None,id_cv,resto_dolares,resto_pesos,moneda,precio_dolar,entrega_dolares,entrega_pesos,None,"Seña",forma_pago)
                caja = Caja.objects.latest('id')
                if forma_pago == "Fondos":
                    ult_fondo = ClienteFondos.objects.filter(cliente=cliente).latest('id')
                    total_pesos = float(ult_fondo.total_pesos) - float(entrega_pesos)
                    total_dolares = float(ult_fondo.total_dolares) - float(entrega_dolares) 
                    ingreso_pesos = 0 - float(entrega_pesos)
                    ingreso_dolares = 0 - float(entrega_dolares) 
                    retiro_fondos = ClienteFondos(
                        moneda = moneda,
                        fecha = datetime.now().date(),
                        cliente = cliente,
                        ingreso_pesos = ingreso_pesos,
                        ingreso_dolares = ingreso_dolares,
                        total_pesos = total_pesos,
                        total_dolares = total_dolares,
                        comprobante = None,
                        metodo = "----------",
                        tipo = "Retiro"
                    )
                    retiro_fondos.save()
                # if caja:
                if forma_pago != "Fondos":
                    moto_datos = f"{moto.tipo} {moto.marca} {moto.modelo}"
                    cliente_datos = f"{cliente.nombre} {cliente.apellido}"
                    usuario = req.user
                    personal = Personal.objects.filter(usuario=usuario.username).first()
                    insert_movimientos_caja(f"Reserva de {moto_datos}, cliente: {cliente_datos}","Ingreso extra",entrega,caja.id,personal.id,moneda,None,forma_pago,1,0,pago_reserva,"moto")
                moto.save()
                messages.success(req, "Moto reservada con éxito.")
                return redirect('Motos')
            
    except Exception as e:
        return render(req,"perfil_administrativo/motos/reservar_moto.html",{"error_message":e,"active_page":"Motos"})


def baja_reserva_moto(req,id_reserva):
    try:
        if req.method == "POST":
            validar_caja = validar_caja_abierta()
            if not validar_caja:
                return render(req,"perfil_administrativo/motos/baja_reserva.html",{"error_message":"No existe caja del día abierta."})
            else:
                reserva = ComprasVentas.objects.get(id=id_reserva)
                senias = CuotasMoto.objects.filter(venta_id=id_reserva)
                if senias.exists:
                    for senia in senias:
                        if senia.metodo_pago == "Fondos":
                            dolar = PrecioDolar.objects.get(id=1)
                            precio_dolar = float(dolar.precio_dolar_tienda)
                            # venta = ClienteAccesorio.objects.get(id=id_cv)
                            fondos_cliente = ClienteFondos.objects.filter(cliente=reserva.cliente).latest('id')
                            if senia.moneda == "Pesos":
                                entrega_pesos = float(senia.valor_pago_pesos)
                                entrega_dolares = entrega_pesos / precio_dolar
                            else:
                                entrega_dolares = float(senia.valor_pago_dolares)
                                entrega_pesos = entrega_dolares * precio_dolar
                            
                            total_pesos = float(fondos_cliente.total_pesos) + entrega_pesos
                            total_dolares = float(fondos_cliente.total_dolares) + entrega_dolares
                            nuevo_fondo = ClienteFondos(
                                moneda = senia.moneda,
                                fecha = datetime.now().date(),
                                cliente = fondos_cliente.cliente,
                                ingreso_pesos = entrega_pesos,
                                ingreso_dolares = entrega_dolares,
                                total_pesos = total_pesos,
                                total_dolares = total_dolares,
                                comprobante = None,
                                metodo = senia.metodo_pago,
                                tipo = "Ingreso"
                            )   
                            nuevo_fondo.save()
                        mov_pago = MovimientoPagoMoto.objects.filter(pago_id=senia.id).first()
                        if mov_pago:
                            id_pago = mov_pago.movimiento_id
                            mov = Movimientos.objects.get(id=id_pago)
                            if senia.metodo_pago == "Efectivo":
                                if senia.moneda == "Pesos":
                                    quitar_deposito = senia.valor_pago_pesos
                                else:
                                    dolar = PrecioDolar.objects.get(id=1)
                                    precio_dolar = dolar.precio_dolar_tienda
                                    quitar_deposito = senia.valor_pago_dolares * precio_dolar
                                caja = Caja.objects.latest('id')
                                caja.depositos = caja.depositos - quitar_deposito
                                caja.save()
                            mov_pago.delete()
                            mov.delete()
                            senia.delete()
                reserva.delete()
                messages.success(req, "Reserva borrada con éxito.")
                return redirect('Reservas')
        else:
            return render(req,"perfil_administrativo/motos/baja_reserva.html",{"active_page":"Reservas"})
    except Exception as e:
        pass

def baja_venta_moto(req,id_venta):
    try:
        venta = ComprasVentas.objects.get(id=id_venta)
        id_cliente = venta.cliente_id
        id_moto = venta.moto_id
        venta.delete()
        moto = Moto.objects.get(id=id_moto)
        moto.pertenece_tienda = 1
        moto.save()
        messages.success(req, "Venta borrada con éxito.")
        return redirect(f"{reverse('ClienteFicha',kwargs={'id_cliente':id_cliente})}")      
    except Exception as e:
        pass

@admin_required
def estadisticas(req):
    # try:
        anio_actual = req.GET.get("anio", now().year)  
        anio_actual = int(anio_actual)  # Asegurar que sea un entero
        # years_available = range(2023,2026)
        
        # prueba_anio = []
        # # Ventas de motos por mes
        # for anio in years_available:
        #     mes_actual = 12  # Si es el año actual, hasta el mes actual
        #     print(anio)
        #     ventas_mensuales = (
        #         ComprasVentas.objects
        #         .filter(fecha_compra__year=anio, tipo="V")
        #         .values('fecha_compra__month','fecha_compra__year')
        #         .annotate(total=Count('id'))
        #         .order_by('fecha_compra__month')
        #     )
        #     prueba_anio.append({
        #         "prueba":ventas_mensuales
        #     })
        #     datos_ventas = [0] * mes_actual  
        #     for venta in ventas_mensuales:
        #         datos_ventas[venta['fecha_compra__month'] - 1] = venta['total']
        years_available = range(2023, now().year + 1)  # Rango de años a comparar
        meses_disponibles = range(1, 13)  # Meses del 1 al 12

        ventas_anuales = {anio: [0] * 12 for anio in years_available}  # Diccionario para almacenar ventas por año y mes

        ventas_anuales_accs = {anio: [0] * 12 for anio in years_available}
        # Obtener ventas por mes y año
        for anio in years_available:
            ventas_mensuales = (
                ComprasVentas.objects
                .filter(fecha_compra__year=anio, tipo="V")
                .values('fecha_compra__month')
                .annotate(total=Count('id'))
            )
            
            # Llenar la lista con los datos obtenidos
            for venta in ventas_mensuales:
                ventas_anuales[anio][venta['fecha_compra__month'] - 1] = venta['total']
            
            ventas_mensuales_accs = (
                ClienteAccesorio.objects
                .filter(fecha_compra__year=anio)
                .values('fecha_compra__month')
                .annotate(total=Count('id'))
            )
            
            # Llenar la lista con los datos obtenidos
            for venta in ventas_mensuales_accs:
                ventas_anuales_accs[anio][venta['fecha_compra__month'] - 1] = venta['total']



        # 5 Marcas más vendidas
        marcas_mas_vendidas = (
            ComprasVentas.objects
            .filter(tipo='V', fecha_compra__year=anio_actual)
            .values('moto__marca')
            .annotate(total_vendidas=Count('id'))
            .order_by('-total_vendidas')[:5]
        )

        datos_marcas = [
            {'marca': resultado['moto__marca'], 'total_vendidas': resultado['total_vendidas']}
            for resultado in marcas_mas_vendidas
        ]

        tipo_accesorio_mas_vendidos = (
            ClienteAccesorio.objects
            .all()
            .values('accesorio__tipo')
            .annotate(total_vendidos=Count('id'))
            .order_by('-total_vendidos')[:5]
        )

        tipo_accesorio_vendidos = [
            {'tipo': resultado['accesorio__tipo'], 'total_vendidos': resultado['total_vendidos']}
            for resultado in tipo_accesorio_mas_vendidos
        ]

        # 5 Motos más vendidas
        motos_mas_vendidas = (
            ComprasVentas.objects
            .filter(tipo='V', fecha_compra__year=anio_actual)
            .values('moto__marca', 'moto__modelo')
            .annotate(total_motos_vendidas=Count('id'))
            .order_by('-total_motos_vendidas')[:5]
        )

        datos_motos = [
            {'marca': resultado['moto__marca'], 'modelo': resultado['moto__modelo'], 'total_motos_vendidas': resultado['total_motos_vendidas']}
            for resultado in motos_mas_vendidas
        ]

        # Ventas de accesorios por mes
        accesorios_mensuales = (
            ClienteAccesorio.objects
            .filter(fecha_compra__year=anio_actual)
            .values('fecha_compra__month')
            .annotate(total=Count('id'))
            .order_by('fecha_compra__month')
        )
        mes_actual_accesorios = now().month if anio_actual == now().year else 12
        datos_accesorios_ventas = [0] * mes_actual_accesorios
        for accesorio in accesorios_mensuales:
            datos_accesorios_ventas[accesorio['fecha_compra__month'] - 1] = accesorio['total']
        
        
        return render(req, "perfil_administrativo/estadisticas/estadisticas.html", {
            "anio": anio,  # Enviar el año seleccionado
            
            "marcas": datos_marcas,
            "motos": datos_motos,
            "tipo_accesorio_vendidos":tipo_accesorio_vendidos,
            "accesorios": datos_accesorios_ventas,
            "active_page": "Estadisticas",
            "years_available":years_available,
            "years_available": years_available,  # Lista de años
            "meses_disponibles": meses_disponibles,  # Lista de meses
            "ventas_anuales": ventas_anuales,
            "ventas_anuales_accs": ventas_anuales_accs,
            "years_available": list(years_available),  # Convertimos a lista para evitar problemas en el template
        "ventas_anuales_json": json.dumps(ventas_anuales),
        "ventas_anuales_accs_json": json.dumps(ventas_anuales_accs)
            
            
        })

    # except Exception as e:
    #     return render(req, "perfil_administrativo/estadisticas/estadisticas.html", {
    #         "error_message": str(e),
    #         "active_page": "Estadisticas"
    #     })

@admin_required
def datos_tienda(req):
    try:
        dolar = PrecioDolar.objects.get(id=1)
        logo_tienda = Logos.objects.get(id=1)
        logo_cv = Logos.objects.get(id=2)
        #"foto_moto":moto.foto.url if moto.foto else None
        imagen_logo_tienda = logo_tienda.logo_UM.url if logo_tienda.logo_UM else None
        imagen_logo_cv = logo_cv.logo_UM.url if logo_cv.logo_UM else None
        return render(req,"perfil_administrativo/tienda.html",{"dolar_precio":dolar.precio_dolar_tienda,
                                                               "imagen_logo_tienda":imagen_logo_tienda,
                                                               "imagen_logo_cv":imagen_logo_cv})
    except Exception as e:
        return render(req,"perfil_administrativo/tienda.html",{"error_message":e})

@admin_required
def modificar_precio_dolar(req):
    try:
        precio = req.POST['precio_dolar']
        dolar = PrecioDolar.objects.get(id=1)
        dolar.precio_dolar_tienda = precio
        dolar.save()
        messages.success(req, "Precio del dólar modificado con éxito.")
        return redirect('Tienda')

    except Exception as e:
        return render(req,"perfil_administrativo/tienda.html",{"error_message":e})

@admin_required
def modificar_logo_tienda(req):
    try:
        logo_tienda = Logos.objects.get(id=1)
        nuevo_logo = req.FILES.get('nuevo_logo')
        logo_tienda.logo_UM = nuevo_logo
        logo_tienda.save()
        messages.success(req, "Logo modificado modificado con éxito.")
        return redirect('Tienda')
    except Exception as e:
        return render(req,"perfil_administrativo/tienda.html",{"error_message":e})

@admin_required
def modificar_logo_cv(req):
    try:
        logo_cv = Logos.objects.get(id=2)
        nuevo_logo = req.FILES.get('nuevo_logo_cv')
        logo_cv.logo_UM = nuevo_logo
        logo_cv.save()
        messages.success(req, "Logo del papel compraventa modificado modificado con éxito.")
        return redirect('Tienda')
    except Exception as e:
        return render(req,"perfil_administrativo/tienda.html",{"error_message":e})

def notificaciones_cumples():

    #EJECUTAR ESTA FUNCION TODOS LOS DIAS A LAS 6AM
    today = datetime.now()
    cumples_cliente = Cliente.objects.filter(fecha_nacimiento__month=today.month,fecha_nacimiento__day=today.day)
    for cliente in cumples_cliente:
        tiene_correo = ClienteCorreo.objects.filter(cliente_id=cliente.id,principal=1).first()
        if tiene_correo:
            correo_enviado = True
            send_mail(
            subject='¡Feliz cumpleaños!',
            message=f'Hola {cliente.nombre}, ¡de parte de Umpierrez Motos y Motos Daniel te deseamos un feliz cumpleaños!',
            #CAMBIAR POR CORREO DE UMPIERREZ MOTOS
            from_email='lumacajuanmanuel@gmail.com',
            #CAMBIAR POR CORREO DE CLIENTE (tiene_correo.correo)
            recipient_list=[tiene_correo.correo],
        )
        else:
            correo_enviado = False
            
        if correo_enviado:
            mensaje = "Se le ha enviado un correo de saludos."
        else:
            mensaje = ""
        nueva_notificacion = Notificaciones(
            descripcion = f"¡Hoy es el cumpleaños de {cliente.nombre} {cliente.apellido}! {mensaje}",
            fecha = datetime.now(),
            tipo = "Cumpleaños"
        )
        #nueva_notificacion.save()

def notificaciones_pagos_atrasados():
    #EJECUTAR ESTA FUNCION TODOS LOS DIAS A LAS 6AM
    pass

def notificaciones_administrativo(req):
    # try:
        #notificaciones_cumples()
        # notificaciones = [
        # {
        # "titulo": "Pago pendiente",
        # "fecha": "2024-11-28",
        # "descripcion": "El cliente Juan Pérez tiene un pago atrasado desde hace 7 días.",
        # "acciones": [
        #     {"nombre": "Ver detalle", "url": "/detalle_pago/123/"},
        #     {"nombre": "Contactar cliente", "url": "/contacto_cliente/123/"}
        # ]
        # },
        # {
        # "titulo": "Cumpleaños del cliente",
        # "fecha": "2024-11-29",
        # "descripcion": "Hoy es el cumpleaños de María López. Envíale un saludo especial.",
        # "acciones": [{"nombre": "Enviar saludo", "url": "/enviar_saludo/456/"}]
        #     }
        #     ]
        usuario = req.user
        usuario_actual = Personal.objects.filter(usuario=usuario.username).first()
        filter_option = req.GET.get('filter', 'all')
        if filter_option == 'leidas':
            # notificaciones = Notificaciones.objects.filter(leido=True).order_by('-id')  # Filtrar solo las leídas
            notificaciones = (
                NotificacionPersonal.objects
                .filter(leido=True,personal=usuario_actual)
                .values(
                    'id',
                    'notificacion__tipo', 
                    'notificacion__fecha', 
                    'notificacion__descripcion',
                    'notificacion__id' 
                     
                )
            ).order_by('-id')
        else:
            # notificaciones = Notificaciones.objects.all().order_by('-id')  # Mostrar todas las notificaciones
            notificaciones = (
                NotificacionPersonal.objects
                .filter(leido=False,personal=usuario_actual)
                .values(
                    'id',
                    'notificacion__id',
                    'notificacion__tipo', 
                    'notificacion__fecha', 
                    'notificacion__descripcion', 
                     
                )
            ).order_by('-id')
        
        #notificaciones = Notificaciones.objects.all().order_by('-id')
        data = []
        for notificacion in notificaciones:
            if notificacion['notificacion__tipo'] == "Atraso en pago":
                acciones = {"nombre": "Enviar correo", "url": ""}
            elif notificacion['notificacion__tipo'] == "Cumpleaños":
                acciones = [{"nombre": "Ver detalle", "url": ""},]
            else:
                acciones = [{"nombre": "Ver detalle", "url": ""},]
            data.append({
                "notificacion":notificacion,
                "acciones":acciones
            })
        
        
        notif_no_leidas = NotificacionPersonal.objects.filter(leido=0,personal=usuario_actual)
        for notificacion in notif_no_leidas:
            notificacion.leido = 1
            notificacion.save()
        
        return render(req,"perfil_administrativo/notificaciones/notificaciones.html",{"data":data}) 
    # except Exception as e:
    #     return render(req,"perfil_administrativo/notificaciones/notificaciones.html",{"error_message":e})

def borrar_notificacion_tienda(req,id_notificacion):
    # try:
        usuario = req.user
        usuario_actual = Personal.objects.filter(usuario=usuario.username).first()
        
        notificacion_usuario = NotificacionPersonal.objects.filter(notificacion_id=id_notificacion,personal=usuario_actual).first()
        notificacion_usuario.delete()
        return redirect('NotificacionesAdministrativo')

    # except Exception as e:
    #     pass

@admin_required
def arqueos(req):
    try:


        arqueos = Caja.objects.all().order_by('-apertura')
        data = []
        dolar = PrecioDolar.objects.get(id=1)
        precio_dolar = float(dolar.precio_dolar_tienda)
        for arqueo in arqueos:
            usuario = Personal.objects.get(id=arqueo.usuario_id)
            diferencia = float(arqueo.diferencia) if arqueo.diferencia is not None else 0
            data.append({
                "arqueo": arqueo,
                "monto_inicial_dol":round(float(arqueo.monto_inicial) / precio_dolar, 2),
                "cierre": arqueo.cierre,  # Permite que sea None si no está definido
                "depositos": arqueo.depositos if arqueo.depositos is not None else 0,
                "depositos_dolares":round(float(arqueo.depositos) / precio_dolar, 2) if arqueo.depositos is not None else 0,
                "egresos": arqueo.egresos if arqueo.egresos is not None else 0,
                "egresos_dolares": round(float(arqueo.egresos) / precio_dolar, 2) if arqueo.egresos is not None else 0,
                "saldo_caja": arqueo.saldo_caja if arqueo.saldo_caja is not None else 0,
                "saldo_sistema": arqueo.saldo_sistema if arqueo.saldo_sistema is not None else 0,
                "diferencia": arqueo.diferencia if arqueo.diferencia is not None else 0,
                "diferencia_dolares":round(diferencia / precio_dolar,2),
                "usuario":usuario.nombre + " " + usuario.apellido
            })

        page_obj = funcion_paginas_varias(req,data)
        return render(req, "perfil_administrativo/arqueos/arqueos.html", {"page_obj": page_obj})

    except Exception as e:
        return render(req,"perfil_administrativo/arqueos/arqueos.html",{"error_message":e})

def buscar_x_fecha_apertura(req):
    try:
        apertura = req.GET.get('f_apertura')
        arqueos = Caja.objects.filter(apertura__date=apertura)
        data = []
        dolar = PrecioDolar.objects.get(id=1)
        precio_dolar = float(dolar.precio_dolar_tienda)
        for arqueo in arqueos:
            usuario = Personal.objects.get(id=arqueo.usuario_id)
            data.append({
                "arqueo": arqueo,
                "monto_inicial_dol":round(float(arqueo.monto_inicial) / precio_dolar, 2),
                "cierre": arqueo.cierre,  # Permite que sea None si no está definido
                "depositos": arqueo.depositos if arqueo.depositos is not None else 0,
                "depositos_dolares":round(float(arqueo.depositos) / precio_dolar, 2) if arqueo.depositos is not None else 0,
                "egresos": arqueo.egresos if arqueo.egresos is not None else 0,
                "egresos_dolares": round(float(arqueo.egresos) / precio_dolar, 2) if arqueo.egresos is not None else 0,
                "saldo_caja": arqueo.saldo_caja if arqueo.saldo_caja is not None else 0,
                "saldo_sistema": arqueo.saldo_sistema if arqueo.saldo_sistema is not None else 0,
                "diferencia": arqueo.diferencia if arqueo.diferencia is not None else 0,
                "usuario":usuario.nombre + " " + usuario.apellido
            })

        page_obj = funcion_paginas_varias(req,data)
        return render(req, "perfil_administrativo/arqueos/arqueos.html", {"page_obj": page_obj})

    except Exception as e:
        return render(req,"perfil_administrativo/notificaciones/notificaciones.html",{"error_message":e})

def buscar_x_fecha_cierre(req):
    try:
        cierre = req.GET.get('fecha_de_cierre')
        arqueos = Caja.objects.filter(cierre__date=cierre)
        data = []
        dolar = PrecioDolar.objects.get(id=1)
        precio_dolar = float(dolar.precio_dolar_tienda)
        for arqueo in arqueos:
            usuario = Personal.objects.get(id=arqueo.usuario_id)
            data.append({
                "arqueo": arqueo,
                "monto_inicial_dol":round(float(arqueo.monto_inicial) / precio_dolar, 2),
                "cierre": arqueo.cierre,  # Permite que sea None si no está definido
                "depositos": arqueo.depositos if arqueo.depositos is not None else 0,
                "depositos_dolares":round(float(arqueo.depositos) / precio_dolar, 2) if arqueo.depositos is not None else 0,
                "egresos": arqueo.egresos if arqueo.egresos is not None else 0,
                "egresos_dolares": round(float(arqueo.egresos) / precio_dolar, 2) if arqueo.egresos is not None else 0,
                "saldo_caja": arqueo.saldo_caja if arqueo.saldo_caja is not None else 0,
                "saldo_sistema": arqueo.saldo_sistema if arqueo.saldo_sistema is not None else 0,
                "diferencia": arqueo.diferencia if arqueo.diferencia is not None else 0,
                "usuario":usuario.nombre + " " + usuario.apellido
            })

        page_obj = funcion_paginas_varias(req,data)
        return render(req, "perfil_administrativo/arqueos/arqueos.html", {"page_obj": page_obj})

    except Exception as e:
        return render(req,"perfil_administrativo/notificaciones/notificaciones.html",{"error_message":e})

@admin_required
def abrir_caja(req):
    try:
        if req.method == "POST":
            caja_abierta = Caja.objects.filter(estado="Abierto").first()
            caja_abierta_2 = Caja.objects.filter(estado="Cuadre de caja").first()
            total_billetes_2000 = (int(req.POST['billetes_2000']) * 2000) 
            total_billetes_1000 = (int(req.POST['billetes_1000']) * 1000)
            total_billetes_500 = (int(req.POST['billetes_500']) * 500)
            total_billetes_200 = (int(req.POST['billetes_200']) * 200)
            total_billetes_100 = (int(req.POST['billetes_100']) * 100)
            total_billetes_50 = (int(req.POST['billetes_50']) * 50)
            total_billetes_20 = (int(req.POST['billetes_20']) * 20)

            total_monedas_50 = (int(req.POST['monedas_50']) * 50)
            total_monedas_10 = (int(req.POST['monedas_10']) * 10)
            total_monedas_5 = (int(req.POST['monedas_5']) * 5)
            total_monedas_2 = (int(req.POST['monedas_2']) * 2)
            total_monedas_1 = int(req.POST['monedas_1'])


            total_billetes_100_dolares = (int(req.POST['billetes_100_dolares']) * 100) 
            total_billetes_50_dolares = (int(req.POST['billetes_50_dolares']) * 50)
            total_billetes_20_dolares = (int(req.POST['billetes_20_dolares']) * 20)
            total_billetes_10_dolares = (int(req.POST['billetes_10_dolares']) * 10)
            total_billetes_5_dolares = (int(req.POST['billetes_5_dolares']) * 5)
            total_billetes_2_dolares = (int(req.POST['billetes_2_dolares']) * 2)
            total_billetes_1_dolares = (int(req.POST['billetes_1_dolares']) * 1)

            error_dinero = validar_billetes(total_billetes_2000,total_billetes_1000,total_billetes_500,total_billetes_200,total_billetes_100,total_billetes_50,total_billetes_20)
            error_monedas = validar_monedas(total_monedas_50,total_monedas_10,total_monedas_5,total_monedas_2,total_monedas_1)
            if error_dinero:
                return render(req,"perfil_administrativo/arqueos/alta_caja.html",{"error_message":error_dinero})
            elif error_monedas:
                return render(req,"perfil_administrativo/arqueos/alta_caja.html",{"error_message":error_monedas})
            elif caja_abierta or caja_abierta_2:
                return render(req,"perfil_administrativo/arqueos/alta_caja.html",{"error_message":"Ya existe una caja abierta."})  
            else:
                dolar = PrecioDolar.objects.get(id=1)
                precio_dolar = float(dolar.precio_dolar_tienda)
                total_dolares = (total_billetes_100_dolares + total_billetes_50_dolares + total_billetes_20_dolares + total_billetes_10_dolares 
                                 + total_billetes_5_dolares + total_billetes_2_dolares + total_billetes_1_dolares)
                total_dolares_a_pesos = total_dolares * precio_dolar
                total_dolares_a_pesos = round(total_dolares_a_pesos,2)
                saldo_inicial = (
                    total_billetes_2000 + total_billetes_1000 + total_billetes_500 + total_billetes_200 + total_billetes_100 + total_billetes_50
                    +total_billetes_20 + total_monedas_50 + total_monedas_10 + total_monedas_5 + total_monedas_2 + total_monedas_1 + total_dolares_a_pesos
                   
                )

                print("SALDO INICIAL ES: " +str(saldo_inicial))
                usuario = req.user
                persona = Personal.objects.filter(usuario=usuario.username).first()
                

                nueva_caja = Caja(
                    apertura = datetime.now(),
                    cierre = None,
                    moneda = "Pesos",
                    monto_inicial = saldo_inicial,
                    depositos = 0,
                    egresos = 0,
                    saldo_caja = 0,
                    saldo_sistema = 0,
                    diferencia = 0,
                    estado = "Abierto",
                    usuario_id = persona.id
                )
                nueva_caja.save()

                messages.success(req, "Caja abierta con éxito.")
                return redirect('Arqueos')
        else:
            return render(req,"perfil_administrativo/arqueos/alta_caja.html",{})
    except Exception as e:
        return render(req,"perfil_administrativo/arqueos/alta_caja.html",{"error_message":e})

@admin_required
def ingresos_caja(req,id_caja):
    # try:
        if int(req.POST['ingresos']) < 0 or int(req.POST['ingresos']) == 0:
            # return render(req,"perfil_administrativo/arqueos/arqueos.html",{"error_message":"Ingrese un valor correcto"})
            messages.error(req,"El monto del ingreso es incorrecto.")
            return redirect('Arqueos')
        else:
            moneda = req.POST['moneda_monto_ingreso']
            if moneda == "Pesos":
                monto = int(req.POST['ingresos'])
            else:
                dolar = PrecioDolar.objects.get(id=1)
                precio_dolar = float(dolar.precio_dolar_tienda)
                monto = float(precio_dolar * float(req.POST['ingresos']))
            caja = Caja.objects.get(id=id_caja)
            if int(caja.diferencia) != 0:
                sumar_diferencia = float(caja.diferencia) + monto
                caja.diferencia = sumar_diferencia
            ingresos = float(caja.depositos) + monto
            caja.depositos = ingresos
            caja.save()
            usuario = req.user
            personal = Personal.objects.filter(usuario=usuario.username).first()
            moneda = req.POST['moneda_monto_ingreso']
            
            insert_movimientos_caja(req.POST['descripcion_ingreso'],"Ingreso extra",req.POST['ingresos'],id_caja,personal.id,moneda,None,"Efectivo",0,0,None,None)
            #insert_movimientos_caja(movimiento_descripcion,tipo,monto,id_caja,id_personal,moneda,rubro,metodo,es_moto,es_accesorio,id_venta,producto)
            messages.success(req, "Rubro ingresado con éxito")
            return redirect(f"{reverse('MovimientosCaja',kwargs={'id_caja':id_caja})}")
    # except Exception as e:
    #     return render(req,"perfil_administrativo/arqueos/arqueos.html",{"error_message":e})

@admin_required
def egresos_caja(req,id_caja):
    try:
        if int(req.POST['egresos']) < 0 or int(req.POST['egresos']) == 0:
            # return render(req,"perfil_administrativo/arqueos/arqueos.html",{"error_message":"Ingrese un valor correcto"})
            messages.error(req,"El monto del egreso es incorrecto.")
            return redirect('Arqueos')
        else:
            moneda = req.POST['moneda_monto_egreso']
            if moneda == "Pesos":
                monto = int(req.POST['egresos'])
            else:
                dolar = PrecioDolar.objects.get(id=1)
                precio_dolar = float(dolar.precio_dolar_tienda)
                monto = float(precio_dolar * float(req.POST['egresos']))
            caja = Caja.objects.get(id=id_caja)
            if int(caja.diferencia) != 0:
                restar_diferencia = float(caja.diferencia) - monto
                caja.diferencia = restar_diferencia
            egresos = float(caja.egresos) + monto
            caja.egresos = egresos
            caja.save()
            usuario = req.user
            personal = Personal.objects.filter(usuario=usuario.username).first()
            id_rubro = req.POST['rubro_egreso']
            rubro = Rubros.objects.get(id=id_rubro)
            nombre_rubro = rubro.rubro
            rubro.cantidad_usos = int(rubro.cantidad_usos) + 1
            rubro.save()
            moneda = req.POST['moneda_monto_egreso']
            insert_movimientos_caja(req.POST['descripcion_egreso'],"Egreso",req.POST['egresos'],id_caja,personal.id,moneda,nombre_rubro,"Efectivo",0,0,None,None)
            #insert_movimientos_caja(movimiento_descripcion,tipo,monto,id_caja,id_personal,moneda,rubro,metodo,es_moto,es_accesorio,id_venta,producto)
            messages.success(req, "Rubro ingresado con éxito")
            return redirect(f"{reverse('MovimientosCaja',kwargs={'id_caja':id_caja})}")
    except Exception as e:
        return render(req,"perfil_administrativo/arqueos/arqueos.html",{"error_message":e})

@admin_required
def saldo_final_caja(req,id_caja):
    # try:
            if req.method == "POST": 
                total_billetes_2000 = (int(req.POST['billetes_2000']) * 2000) 
                total_billetes_1000 = (int(req.POST['billetes_1000']) * 1000)
                total_billetes_500 = (int(req.POST['billetes_500']) * 500)
                total_billetes_200 = (int(req.POST['billetes_200']) * 200)
                total_billetes_100 = (int(req.POST['billetes_100']) * 100)
                total_billetes_50 = (int(req.POST['billetes_50']) * 50)
                total_billetes_20 = (int(req.POST['billetes_20']) * 20)

                total_monedas_50 = (int(req.POST['monedas_50']) * 50)
                total_monedas_10 = (int(req.POST['monedas_10']) * 10)
                total_monedas_5 = (int(req.POST['monedas_5']) * 5)
                total_monedas_2 = (int(req.POST['monedas_2']) * 2)
                total_monedas_1 = int(req.POST['monedas_1'])

                total_billetes_100_dolares = (int(req.POST['billetes_100_dolares']) * 100) 
                total_billetes_50_dolares = (int(req.POST['billetes_50_dolares']) * 50)
                total_billetes_20_dolares = (int(req.POST['billetes_20_dolares']) * 20)
                total_billetes_10_dolares = (int(req.POST['billetes_10_dolares']) * 10)
                total_billetes_5_dolares = (int(req.POST['billetes_5_dolares']) * 5)
                total_billetes_2_dolares = (int(req.POST['billetes_2_dolares']) * 2)
                total_billetes_1_dolares = (int(req.POST['billetes_1_dolares']) * 1)
                error_dinero = validar_billetes(total_billetes_2000,total_billetes_1000,total_billetes_500,total_billetes_200,total_billetes_100,total_billetes_50,total_billetes_20)
                error_monedas = validar_monedas(total_monedas_50,total_monedas_10,total_monedas_5,total_monedas_2,total_monedas_1)
                if error_dinero:
                    return render(req,"perfil_administrativo/arqueos/alta_caja.html",{"error_message":error_dinero})
                elif error_monedas:
                    return render(req,"perfil_administrativo/arqueos/alta_caja.html",{"error_message":error_monedas})
                else:
                    dolar = PrecioDolar.objects.get(id=1)
                    precio_dolar = float(dolar.precio_dolar_tienda)
                    total_dolares = (total_billetes_100_dolares + total_billetes_50_dolares + total_billetes_20_dolares + total_billetes_10_dolares 
                                    + total_billetes_5_dolares + total_billetes_2_dolares + total_billetes_1_dolares)
                    total_dolares_a_pesos = total_dolares * precio_dolar
                    total_dolares_a_pesos = round(total_dolares_a_pesos,2)
                    saldo_final_declarado = (
                        total_billetes_2000 + total_billetes_1000 + total_billetes_500 + total_billetes_200 + total_billetes_100 + total_billetes_50
                        +total_billetes_20 + total_monedas_50 + total_monedas_10 + total_monedas_5 + total_monedas_2 + total_monedas_1 + total_dolares_a_pesos
                    )

                    caja = Caja.objects.get(id=id_caja)

                    movimientos = Movimientos.objects.filter(caja_id=id_caja,metodo="Efectivo")
                    ingresos_pesos = 0
                    ingresos_dolares = 0
                    egresos_pesos = 0
                    egresos_dolares = 0
                    for mov in movimientos:
                        if mov.moneda == "Pesos":
                            if mov.tipo == "Ingreso" or mov.tipo == "Ingreso extra":
                                ingresos_pesos = ingresos_pesos + float(mov.monto)
                            else:
                                egresos_pesos = egresos_pesos + float(mov.monto)
                        else:
                            if mov.tipo == "Ingreso" or mov.tipo == "Ingreso extra":
                                ingresos_dolares = ingresos_dolares + float(mov.monto)
                            else:
                                egresos_dolares = egresos_dolares + float(mov.monto)
                    
                    total_movimientos_pesos = ingresos_pesos - egresos_pesos
                    total_movimientos_dolares = ingresos_dolares - egresos_dolares
                    total = total_movimientos_pesos + (total_movimientos_dolares * precio_dolar)
                    total = round(total,2)
                    # saldo_sistema = caja.monto_inicial + caja.depositos - caja.egresos
                    # diferencia = float(saldo_final_declarado) - float(caja.monto_inicial)
                    saldo_final = float(caja.monto_inicial) + total
                    # saldo_sistema = abs(saldo_sistema)
                    
                    #DECLARADO POR EL USUARIO, APARECERA EN "EFECTIVO FINAL"
                    caja.saldo_caja = saldo_final_declarado
                    #SALDO CALCULADO POR EL PROGRAMA, APARECERA EN "SALDO CAJA"
                    caja.saldo_sistema = saldo_final
                    #EL TOTAL DE LO QUE TENGO AL FINAL MENOS LA CANTIDAD CALCULADA POR EL SISTEMA, APARECERA EN COLUMNA DIERENCIA
                    diferencia_caja = saldo_final_declarado - saldo_final
                    caja.diferencia = diferencia_caja
                    caja.estado = "Cuadre de caja"
                    caja.save()
                    # print("SALDO FINAL ES : " + str(saldo_final))
                    # print("SALDO SISTEMA ES : " + str(saldo_sistema))
                    # print("DIFERENCIA ES : " + str(diferencia))
                    messages.success(req, "Caja declarada con éxito")
                    return redirect('Arqueos')
            else:
                return render(req,"perfil_administrativo/arqueos/saldo_final.html",{})
    # except Exception as e:
    #     return render(req,"perfil_administrativo/arqueos/alta_caja.html",{"error_message":e})

@admin_required
def editar_saldo_inicial(req,id_caja):
    # try:
            if req.method == "POST": 
                total_billetes_2000 = (int(req.POST['billetes_2000']) * 2000) 
                total_billetes_1000 = (int(req.POST['billetes_1000']) * 1000)
                total_billetes_500 = (int(req.POST['billetes_500']) * 500)
                total_billetes_200 = (int(req.POST['billetes_200']) * 200)
                total_billetes_100 = (int(req.POST['billetes_100']) * 100)
                total_billetes_50 = (int(req.POST['billetes_50']) * 50)
                total_billetes_20 = (int(req.POST['billetes_20']) * 20)

                total_monedas_50 = (int(req.POST['monedas_50']) * 50)
                total_monedas_10 = (int(req.POST['monedas_10']) * 10)
                total_monedas_5 = (int(req.POST['monedas_5']) * 5)
                total_monedas_2 = (int(req.POST['monedas_2']) * 2)
                total_monedas_1 = int(req.POST['monedas_1'])

                total_billetes_100_dolares = (int(req.POST['billetes_100_dolares']) * 100) 
                total_billetes_50_dolares = (int(req.POST['billetes_50_dolares']) * 50)
                total_billetes_20_dolares = (int(req.POST['billetes_20_dolares']) * 20)
                total_billetes_10_dolares = (int(req.POST['billetes_10_dolares']) * 10)
                total_billetes_5_dolares = (int(req.POST['billetes_5_dolares']) * 5)
                total_billetes_2_dolares = (int(req.POST['billetes_2_dolares']) * 2)
                total_billetes_1_dolares = (int(req.POST['billetes_1_dolares']) * 1)
                error_dinero = validar_billetes(total_billetes_2000,total_billetes_1000,total_billetes_500,total_billetes_200,total_billetes_100,total_billetes_50,total_billetes_20)
                error_monedas = validar_monedas(total_monedas_50,total_monedas_10,total_monedas_5,total_monedas_2,total_monedas_1)
                if error_dinero:
                    return render(req,"perfil_administrativo/arqueos/editar_monto_inicial.html",{"error_message":error_dinero})
                elif error_monedas:
                    return render(req,"perfil_administrativo/arqueos/editar_monto_inicial.html",{"error_message":error_monedas})
                else:
                    dolar = PrecioDolar.objects.get(id=1)
                    precio_dolar = float(dolar.precio_dolar_tienda)
                    total_dolares = (total_billetes_100_dolares + total_billetes_50_dolares + total_billetes_20_dolares + total_billetes_10_dolares 
                                    + total_billetes_5_dolares + total_billetes_2_dolares + total_billetes_1_dolares)
                    total_dolares_a_pesos = total_dolares * precio_dolar
                    total_dolares_a_pesos = round(total_dolares_a_pesos,2)
                    monto_inicial_declarado = (
                        total_billetes_2000 + total_billetes_1000 + total_billetes_500 + total_billetes_200 + total_billetes_100 + total_billetes_50
                        +total_billetes_20 + total_monedas_50 + total_monedas_10 + total_monedas_5 + total_monedas_2 + total_monedas_1 + total_dolares_a_pesos
                    )

                    caja = Caja.objects.get(id=id_caja)
                    if caja.estado == "Cuadre de caja":
                        movimientos = Movimientos.objects.filter(caja_id=id_caja,metodo="Efectivo")
                        ingresos_pesos = 0
                        ingresos_dolares = 0
                        egresos_pesos = 0
                        egresos_dolares = 0
                        for mov in movimientos:
                            if mov.moneda == "Pesos":
                                if mov.tipo == "Ingreso" or mov.tipo == "Ingreso extra":
                                    ingresos_pesos = ingresos_pesos + float(mov.monto)
                                else:
                                    egresos_pesos = egresos_pesos + float(mov.monto)
                            else:
                                if mov.tipo == "Ingreso" or mov.tipo == "Ingreso extra":
                                    ingresos_dolares = ingresos_dolares + float(mov.monto)
                                else:
                                    egresos_dolares = egresos_dolares + float(mov.monto)
                        
                        total_movimientos_pesos = ingresos_pesos - egresos_pesos
                        total_movimientos_dolares = ingresos_dolares - egresos_dolares
                        total = total_movimientos_pesos + (total_movimientos_dolares * precio_dolar)
                        total = round(total,2)
                        # saldo_sistema = caja.monto_inicial + caja.depositos - caja.egresos
                        # diferencia = float(saldo_final_declarado) - float(caja.monto_inicial)
                        saldo_final = monto_inicial_declarado + total
                        # saldo_sistema = abs(saldo_sistema)
                        
                        #SALDO CALCULADO POR EL PROGRAMA, APARECERA EN "SALDO CAJA"
                        caja.saldo_sistema = saldo_final
                        #EL TOTAL DE LO QUE TENGO AL FINAL MENOS LA CANTIDAD CALCULADA POR EL SISTEMA, APARECERA EN COLUMNA DIERENCIA
                        diferencia_caja = float(caja.saldo_caja) - saldo_final
                        caja.diferencia = diferencia_caja
                    
                    caja.monto_inicial = monto_inicial_declarado
                    caja.save()
                    messages.success(req, "Monto inicial modificado con éxito")
                    return redirect(f"{reverse('MovimientosCaja',kwargs={'id_caja':id_caja})}")
                    # return redirect('Arqueos')
            else:
                return render(req,"perfil_administrativo/arqueos/editar_monto_inicial.html",{})
    # except Exception as e:
    #     return render(req,"perfil_administrativo/arqueos/alta_caja.html",{"error_message":e})

@admin_required
def cerrar_caja(req,id_caja):
    try:
        if req.method == "POST":
            caja = Caja.objects.get(id=id_caja)
            caja.estado = "Cerrado"
            caja.cierre = datetime.now()
            caja.save()
            return render(req,"perfil_administrativo/arqueos/cierre_caja.html",{"message":"Caja cerrada con éxito."})
        else:
             return render(req,"perfil_administrativo/arqueos/cierre_caja.html",{})
    except Exception as e:
        return render(req,"perfil_administrativo/arqueos/alta_caja.html",{"error_message":e})

@admin_required
def movimientos_caja(req,id_caja):
    try:
        print("ID CAJA ES: " + str(id_caja))
        req.session["id_caja_movimiento"] = id_caja
        movimientos = Movimientos.objects.filter(caja_id=id_caja,metodo="Efectivo").order_by('-fecha')
        caja = Caja.objects.get(id=id_caja)
        monto_inicial = caja.monto_inicial
        
        
        dolar = PrecioDolar.objects.get(id=1)
        precio_dolar = float(dolar.precio_dolar_tienda)
        if caja.estado == "Cuadre de caja" or caja.estado == "Cerrado": 
            efectivo_final_pesos = float(caja.saldo_caja)
            saldo_caja_pesos = float(caja.saldo_sistema)
            efectivo_final_dolares = round(efectivo_final_pesos / precio_dolar,2)
            saldo_caja_dolares = round(saldo_caja_pesos / precio_dolar,2)
            texto_efectivo_final = f"${efectivo_final_pesos}/U$s{efectivo_final_dolares}"
            texto_saldo_caja = f"${saldo_caja_pesos}/U$s{saldo_caja_dolares}"
        else:
            texto_efectivo_final = ""
            texto_saldo_caja = ""
        # print(mes_actual)
        data = []
        for movimiento in movimientos:
            usuario = Personal.objects.get(id=movimiento.usuario_id)
            es_pago_moto = MovimientoPagoMoto.objects.filter(movimiento_id=movimiento.id).first()
            es_pago_accesorio = MovimientoPagoAccesorio.objects.filter(movimiento_id=movimiento.id).first()

            if es_pago_moto or es_pago_accesorio:
                mostrar_boton = False
            else: 
                mostrar_boton = True
            data.append({
                "movimiento":movimiento,
                "descripcion":movimiento.movimiento if movimiento.movimiento else "Sin descripción",
                "usuario":usuario.nombre + " " + usuario.apellido,
                "mostrar_boton":mostrar_boton
            })
        
        
        mes_actual = datetime.now().month
        movimientos_mes_actual = Movimientos.objects.filter(caja_id=id_caja,fecha__month=mes_actual,metodo="Efectivo")
        ganancias_mes = 0
        egresos_mes = 0 
        ganancias_netas_mes = 0
        ingresos_extras_mes = 0
        
        for movimiento in movimientos_mes_actual:
            if movimiento.tipo == "Ingreso":
                if movimiento.moneda == "Pesos":
                    ganancias_mes = ganancias_mes + float(movimiento.monto)
                else:
                    ganancias_mes = ganancias_mes + (float(movimiento.monto) * precio_dolar)
            elif movimiento.tipo == "Egreso":
                if movimiento.moneda == "Pesos":
                    egresos_mes = egresos_mes + float(movimiento.monto)
                else:
                    egresos_mes = egresos_mes + (float(movimiento.monto) * precio_dolar)
            elif movimiento.tipo == "Ingreso extra":
                if movimiento.moneda == "Pesos":
                    ingresos_extras_mes = ingresos_extras_mes + float(movimiento.monto)
                else:
                    ingresos_extras_mes = ingresos_extras_mes + (float(movimiento.monto) * precio_dolar)
            else:
                pass
            

        ganancias_netas_mes = ganancias_mes + ingresos_extras_mes - egresos_mes

        # print("GANANCIAS DEL MES: " + str(ganancias_mes))
        # print("EGRESOS DEL MES: -" + str(egresos_mes))
        # print("INGRESOS EXTRAS DEL MES: " + str(ingresos_extras_mes))
        # print("GANANCIAS NETAS DEL MES: " + str(ganancias_netas_mes))
        
        
        
        ganancias_dia = 0
        egresos_dia = 0
        ganancias_netas_dia = 0
        ingresos_extras_dia = 0

        fecha_actual = datetime.now().date()
        movimientos_fecha_actual = Movimientos.objects.filter(caja_id=id_caja,fecha__date=fecha_actual,metodo="Efectivo")

        for movimiento in movimientos_fecha_actual:
            if movimiento.tipo == "Ingreso":
                if movimiento.moneda == "Pesos":
                    ganancias_dia = ganancias_dia + float(movimiento.monto)
                else:
                    ganancias_dia = ganancias_dia + (float(movimiento.monto) * precio_dolar)
            elif movimiento.tipo == "Egreso":
                if movimiento.moneda == "Pesos":
                    egresos_dia = egresos_dia + float(movimiento.monto)
                else:
                    egresos_dia = egresos_dia + (float(movimiento.monto) * precio_dolar)
            elif movimiento.tipo == "Ingreso extra":
                if movimiento.moneda == "Pesos":
                    ingresos_extras_dia = ingresos_extras_dia + float(movimiento.monto)
                else:
                    ingresos_extras_dia = ingresos_extras_dia + (float(movimiento.monto) * precio_dolar)
            else:
                pass
            

        ganancias_netas_dia = ganancias_dia + ingresos_extras_dia - egresos_dia
        # print("GANANCIAS DEL DIA: " + str(ganancias_dia))
        # print("EGRESOS DEL DIA: -" + str(egresos_dia))
        # print("INGRESOS EXTRAS DEL DIA: " + str(ingresos_extras_dia))
        # print("GANANCIAS NETAS DEL DIA: " + str(ganancias_netas_dia))

        

        todos_movimientos_pesos = Movimientos.objects.filter(caja_id=id_caja,metodo="Efectivo")
        ingresos_efectivo_pesos = 0
        # ingresos_transferencia_pesos = 0
        # ingresos_debito_pesos = 0
        egresos_efectivo_pesos = 0
        # egresos_transferencia_pesos = 0
        # egresos_debito_pesos = 0

        ingresos_efectivo_dolares = 0
        # ingresos_transferencia_dolares = 0
        # ingresos_debito_dolares = 0
        egresos_efectivo_dolares = 0
        # egresos_transferencia_dolares = 0
        # egresos_debito_dolares = 0

        for mov in todos_movimientos_pesos:
            if mov.moneda == "Pesos":
                if mov.tipo == "Egreso":
                    # if mov.metodo == "Efectivo":
                        egresos_efectivo_pesos = egresos_efectivo_pesos + float(mov.monto)
                    # elif mov.metodo == "Transferencia":
                    #     egresos_transferencia_pesos = egresos_transferencia_pesos + float(mov.monto) 
                    # else:
                    #     egresos_debito_pesos = egresos_debito_pesos + float(mov.monto)
                else:
                    # if mov.metodo == "Efectivo":
                        ingresos_efectivo_pesos = ingresos_efectivo_pesos + float(mov.monto)
                    # elif mov.metodo == "Transferencia":
                    #     ingresos_transferencia_pesos = ingresos_transferencia_pesos + float(mov.monto)
                    # else:
                    #     ingresos_debito_pesos = ingresos_debito_pesos + float(mov.monto)
            else:
                if mov.tipo == "Egreso":
                    # if mov.metodo == "Efectivo":
                        egresos_efectivo_dolares = egresos_efectivo_dolares + float(mov.monto)
                    # elif mov.metodo == "Transferencia":
                    #     egresos_transferencia_dolares = egresos_transferencia_dolares + float(mov.monto) 
                    # else:
                    #     egresos_debito_dolares = egresos_debito_dolares + float(mov.monto)
                else:
                    # if mov.metodo == "Efectivo":
                        ingresos_efectivo_dolares = ingresos_efectivo_dolares + float(mov.monto)
                    # elif mov.metodo == "Transferencia":
                    #     ingresos_transferencia_dolares = ingresos_transferencia_dolares + float(mov.monto)
                    # else:
                    #     ingresos_debito_dolares = ingresos_debito_dolares + float(mov.monto)


        # total_ingresos_dolares = ingresos_efectivo_dolares + ingresos_transferencia_dolares + ingresos_debito_dolares
        # total_egresos_dolares = egresos_efectivo_dolares + egresos_transferencia_dolares + egresos_debito_dolares
        
        # total_ingresos_pesos = ingresos_efectivo_pesos + ingresos_transferencia_pesos + ingresos_debito_pesos
        # total_egresos_pesos = egresos_efectivo_pesos + egresos_transferencia_pesos + egresos_debito_pesos

        total_ingresos_efectivo_pesos = round(ingresos_efectivo_pesos + float(ingresos_efectivo_dolares * precio_dolar),2)
        total_ingresos_efectivo_dolares = round(ingresos_efectivo_dolares + float(ingresos_efectivo_pesos / precio_dolar),2)

        # total_ingresos_transferencia_pesos = round(ingresos_transferencia_pesos + float(ingresos_transferencia_dolares * precio_dolar),2)
        # total_ingresos_transferencia_dolares = round(ingresos_transferencia_dolares + float(ingresos_transferencia_pesos / precio_dolar),2)

        # total_ingresos_debito_pesos = round(ingresos_debito_pesos + float(ingresos_debito_dolares * precio_dolar),2)
        # total_ingresos_debito_dolares = round(ingresos_debito_dolares + float(ingresos_debito_pesos / precio_dolar),2)


        total_egresos_efectivo_pesos = round(egresos_efectivo_pesos + float(egresos_efectivo_dolares * precio_dolar),2)
        total_egresos_efectivo_dolares = round(egresos_efectivo_dolares + float(egresos_efectivo_pesos / precio_dolar),2)

        # total_egresos_transferencia_pesos = round(egresos_transferencia_pesos + float(egresos_transferencia_dolares * precio_dolar),2)
        # total_egresos_transferencia_dolares = round(egresos_transferencia_dolares + float(egresos_transferencia_pesos / precio_dolar),2)

        # total_egresos_debito_pesos = round(egresos_debito_pesos + float(egresos_debito_dolares * precio_dolar),2)
        # total_egresos_debito_dolares = round(egresos_debito_dolares + float(egresos_debito_pesos / precio_dolar),2)

        


        
        rubros = Rubros.objects.filter(habilitado=1).order_by('-cantidad_usos')
        page_obj = funcion_paginas_varias(req,data)
        return render(req, "perfil_administrativo/arqueos/movimientos.html", {"page_obj": page_obj,
                                                                              
                                                                              "estado_caja":caja.estado,
                                                                              "ganancias_mes":ganancias_mes,
                                                                              "ganancias_dia":ganancias_dia,
                                                                              "egresos_dia":egresos_dia,
                                                                              "egresos_mes":egresos_mes,
                                                                              "ganancias_netas_mes":ganancias_netas_mes,
                                                                              "ganancias_netas_dia":ganancias_netas_dia,
                                                                              "ingresos_extra_dia":ingresos_extras_dia,
                                                                              "ingresos_extra_mes":ingresos_extras_mes,

                                                                              "ganancias_mes_dolares":round(ganancias_mes / precio_dolar,2),
                                                                              "ganancias_dia_dolares":round(ganancias_dia / precio_dolar,2),
                                                                              "egresos_dia_dolares":round(egresos_dia / precio_dolar,2),
                                                                              "egresos_mes_dolares":round(egresos_mes / precio_dolar,2),
                                                                              "ganancias_netas_mes_dolares":round(ganancias_netas_mes / precio_dolar,2),
                                                                              "ganancias_netas_dia_dolares":round(ganancias_netas_dia / precio_dolar,2),
                                                                              "ingresos_extra_dia_dolares":round(ingresos_extras_dia / precio_dolar,2),
                                                                              "ingresos_extra_mes_dolares":round(ingresos_extras_mes / precio_dolar,2),
                                                                              
                                                                              
                                                                            
                                                                              "id_caja":id_caja,
                                                                              "rubros":rubros,
                                                                              "total_ingresos_efectivo_pesos":total_ingresos_efectivo_pesos,
                                                                              "total_ingresos_efectivo_dolares":total_ingresos_efectivo_dolares,
                                                                    
                                                                              "total_egresos_efectivo_pesos":total_egresos_efectivo_pesos,
                                                                              "total_egresos_efectivo_dolares":total_egresos_efectivo_dolares,
                                                            
                                                                              "efectivo_inicial":texto_efectivo_final,
                                                                              "saldo_caja":texto_saldo_caja,
                                                                              "monto_inicial":monto_inicial
                                                                              
                                                                              
                                                                              })
    except Exception as e:
        return render(req,"perfil_administrativo/arqueos/movimientos.html",{"error_message":e})

def buscar_detalles_movimientos_x_fecha(req):
    # try:
        # fecha = req.POST['']
        print("Datos recibidos en GET:", req.GET)
        id_caja = req.session["id_caja_movimiento"]
        f_detalle_str = req.POST['f_detalles']  # Cambiado a paréntesis
        fecha = datetime.strptime(f_detalle_str, '%Y-%m-%d').date() if f_detalle_str else None
        print(fecha)
        print(f_detalle_str)
        dolar = PrecioDolar.objects.get(id=1)
        precio_dolar = float(dolar.precio_dolar_tienda)

        #CANTIDAD MOTOS VENDIDAS EN LA FECHA INGRESADA, TOTAL DEL INGRESO EN PESOS Y DOLARES 
        movs_motos_vendidas_dia = Movimientos.objects.filter(caja_id=id_caja,fecha__date=fecha, es_moto=1,metodo="Efectivo")
        cantidad_vendidas_dia = ComprasVentas.objects.filter(fecha_compra=fecha,tipo="V").count()
        total_moto_dia_pesos = 0
        total_moto_dia_dolares = 0
        for movimiento in movs_motos_vendidas_dia:
            # moto = Moto.objects.get(id=movimiento.moto_id)
            if movimiento.tipo == "Ingreso":
                if movimiento.moneda == "Pesos":
                    total_moto_dia_pesos = total_moto_dia_pesos + float(movimiento.monto)
                    total_moto_dia_dolares = total_moto_dia_pesos / precio_dolar 
                elif movimiento.moneda == "Dolares":
                    total_moto_dia_dolares = total_moto_dia_dolares + float(movimiento.monto)
                    total_moto_dia_pesos = total_moto_dia_dolares * precio_dolar


        #CANTIDAD ACCESORIOS VENDIDOS EN LA FECHA INGRESADA, TOTAL DEL INGRESO EN PESOS Y DOLARES 
        movs_accesorios_vendidos_dia = Movimientos.objects.filter(caja_id=id_caja,fecha__date=fecha, es_accesorio=1,metodo="Efectivo")
        cantidad_vendidos_dia = ClienteAccesorio.objects.filter(fecha_compra=fecha).count()
        total_accesorio_dia_pesos = 0
        total_accesorio_dia_dolares = 0
        for movimiento in movs_accesorios_vendidos_dia:
            if movimiento.tipo == "Ingreso":
                if movimiento.moneda == "Pesos":
                    total_accesorio_dia_pesos = total_accesorio_dia_pesos + float(movimiento.monto)
                    total_accesorio_dia_dolares = total_accesorio_dia_pesos / precio_dolar 
                elif movimiento.moneda == "Dolares":
                    total_accesorio_dia_dolares = total_accesorio_dia_dolares + float(movimiento.monto)
                    total_accesorio_dia_pesos = total_accesorio_dia_dolares * precio_dolar


        movimientos_dia = Movimientos.objects.filter(caja_id=id_caja,fecha__date=fecha,metodo="Efectivo")
        ingresos_extra = 0
        egresos = 0 
        ingresos_extra_dolares = 0
        egresos_dolares = 0 
        data_egresos = []
        data_ingresos_extra = []
        for mov in movimientos_dia:
            if mov.tipo == "Ingreso extra":
                if mov.moneda == "Pesos":
                    ingresos_extra = ingresos_extra + float(mov.monto)
                    ingresos_extra_dolares = ingresos_extra / precio_dolar
                else:
                    ingresos_extra_dolares = ingresos_extra_dolares + float(mov.monto)
                    ingresos_extra = ingresos_extra_dolares * precio_dolar
                data_ingresos_extra.append({
                    "valor":mov.monto,
                    "motivo":mov.movimiento
                })
            elif mov.tipo == "Egreso":
                if mov.moneda == "Pesos":
                    egresos = egresos + float(mov.monto)
                    egresos_dolares = egresos / precio_dolar
                else:
                    egresos_dolares = egresos_dolares + float(mov.monto)
                    egresos = egresos_dolares * precio_dolar
                data_egresos.append({
                    "rubro":mov.rubro,
                    "valor":mov.monto,
                    "motivo":mov.movimiento
                })
            
        
        total_moto_dia_pesos = round(total_moto_dia_pesos, 2)
        total_moto_dia_dolares = round(total_moto_dia_dolares, 2)
        total_accesorio_dia_pesos = round(total_accesorio_dia_pesos, 2)
        total_accesorio_dia_dolares = round(total_accesorio_dia_dolares, 2)
        subtotal_pesos = total_moto_dia_pesos + total_accesorio_dia_pesos
        subtotal_pesos = round(subtotal_pesos, 2)
        subtotal_dolares = total_moto_dia_dolares + total_accesorio_dia_dolares
        subtotal_dolares = round(subtotal_dolares, 2)
        ingresos_extra = round(ingresos_extra, 2)
        egresos = round(egresos, 2)
        ingresos_extra_dolares = round(ingresos_extra_dolares, 2)
        egresos_dolares = round(egresos_dolares, 2)
        total_pesos = subtotal_pesos + ingresos_extra - egresos
        total_pesos = int(total_pesos)
        total_dolares = subtotal_dolares + ingresos_extra_dolares - egresos_dolares
        total_dolares = int(total_dolares)
        # fecha_json = datetime.strptime(f_detalle_str, '%d-%m-%Y').date() if f_detalle_str else None
        data = [
        {"tipo": "Motos", 
         "venta": cantidad_vendidas_dia, 
         "total_pesos": total_moto_dia_pesos, 
         "total_dolares": total_moto_dia_dolares},
        {"tipo": "Accesorios", 
         "venta": cantidad_vendidos_dia, 
         "total_pesos": total_accesorio_dia_pesos, 
         "total_dolares": total_accesorio_dia_dolares},
         {"egresos":egresos,
          "ingresos_extra":ingresos_extra,
          "egresos_dolares":egresos_dolares,
          "ingresos_extra_dolares":ingresos_extra_dolares,
            "subtotal_pesos": subtotal_pesos,
          "subtotal_dolares":subtotal_dolares,
          "total_general_pesos":total_pesos,
          "total_general_dolares":total_dolares},
          {"egresos_detalle": data_egresos},  # Pasamos la lista completa
            {"ingresos_extra_detalle": data_ingresos_extra},
            {"fecha":f_detalle_str}

        ]
        print("TOTAL ACCESORIOS DEL DIA EN PESOS: $" + str(total_accesorio_dia_pesos))
        print("TOTAL ACCESORIOS DEL DIA EN DOLARES: U$s" + str(total_accesorio_dia_dolares))
        print("CANTIDAD ACCESORIOS DE MOTOS VENDIDAS: " + str(cantidad_vendidos_dia))
        print("TOTAL EN PESOS: " + str(total_pesos))
        print("TOTAL EN DOLARES: " + str(total_dolares))
        print("EGRESOS: " + str(egresos))
        print("INGRESOS EXTRA: " + str(ingresos_extra))
        
        
        return JsonResponse(data, safe=False)
        
        # motos_vendidas_mes = ComprasVentas.objects.filter(fecha_compra__month=mes_actual,tipo="V")
        # cantidad_vendidas_mes = 0
        # for venta in motos_vendidas_mes:
        #     pass
    # except Exception as e:
    #     pass

def buscar_detalles_movimientos_x_fecha_excel(req):
        print("Datos recibidos en GET:", req.GET)
        f_detalle_str = req.POST['f_detalles']  # Cambiado a paréntesis
        fecha = datetime.strptime(f_detalle_str, '%Y-%m-%d').date() if f_detalle_str else None
        print(fecha)
        print(f_detalle_str)
        dolar = PrecioDolar.objects.get(id=1)
        precio_dolar = float(dolar.precio_dolar_tienda)
        id_caja = req.session["id_caja_movimiento"]
        #CANTIDAD MOTOS VENDIDAS EN LA FECHA INGRESADA, TOTAL DEL INGRESO EN PESOS Y DOLARES 
        movs_motos_vendidas_dia = Movimientos.objects.filter(caja_id=id_caja,fecha__date=fecha, es_moto=1,metodo="Efectivo")
        cantidad_vendidas_dia = ComprasVentas.objects.filter(fecha_compra=fecha,tipo="V").count()
        total_moto_dia_pesos = 0
        total_moto_dia_dolares = 0
        for movimiento in movs_motos_vendidas_dia:
            # moto = Moto.objects.get(id=movimiento.moto_id)
            if movimiento.tipo == "Ingreso":
                if movimiento.moneda == "Pesos":
                    total_moto_dia_pesos = total_moto_dia_pesos + float(movimiento.monto)
                    total_moto_dia_dolares = total_moto_dia_pesos / precio_dolar 
                elif movimiento.moneda == "Dolares":
                    total_moto_dia_dolares = total_moto_dia_dolares + float(movimiento.monto)
                    total_moto_dia_pesos = total_moto_dia_dolares * precio_dolar


        #CANTIDAD ACCESORIOS VENDIDOS EN LA FECHA INGRESADA, TOTAL DEL INGRESO EN PESOS Y DOLARES 
        movs_accesorios_vendidos_dia = Movimientos.objects.filter(caja_id=id_caja,fecha__date=fecha, es_accesorio=1,metodo="Efectivo")
        cantidad_vendidos_dia = ClienteAccesorio.objects.filter(fecha_compra=fecha).count()
        total_accesorio_dia_pesos = 0
        total_accesorio_dia_dolares = 0
        for movimiento in movs_accesorios_vendidos_dia:
            if movimiento.tipo == "Ingreso":
                if movimiento.moneda == "Pesos":
                    total_accesorio_dia_pesos = total_accesorio_dia_pesos + float(movimiento.monto)
                    total_accesorio_dia_dolares = total_accesorio_dia_pesos / precio_dolar 
                elif movimiento.moneda == "Dolares":
                    total_accesorio_dia_dolares = total_accesorio_dia_dolares + float(movimiento.monto)
                    total_accesorio_dia_pesos = total_accesorio_dia_dolares * precio_dolar


        movimientos_dia = Movimientos.objects.filter(caja_id=id_caja,fecha__date=fecha,metodo="Efectivo")
        ingresos_extra = 0
        egresos = 0 
        ingresos_extra_dolares = 0
        egresos_dolares = 0 
        data_egresos = []
        data_ingresos_extra = []
        for mov in movimientos_dia:
            if mov.tipo == "Ingreso extra":
                if mov.moneda == "Pesos":
                    ingresos_extra = ingresos_extra + float(mov.monto)
                    ingresos_extra_dolares = ingresos_extra / precio_dolar
                else:
                    ingresos_extra_dolares = ingresos_extra_dolares + float(mov.monto)
                    ingresos_extra = ingresos_extra_dolares * precio_dolar
                data_ingresos_extra.append({
                    "valor":mov.monto,
                    "motivo":mov.movimiento
                })
            elif mov.tipo == "Egreso":
                if mov.moneda == "Pesos":
                    egresos = egresos + float(mov.monto)
                    egresos_dolares = egresos / precio_dolar
                else:
                    egresos_dolares = egresos_dolares + float(mov.monto)
                    egresos = egresos_dolares * precio_dolar
                data_egresos.append({
                    "rubro":mov.rubro,
                    "valor":mov.monto,
                    "motivo":mov.movimiento
                })
            
        
        total_moto_dia_pesos = round(total_moto_dia_pesos, 2)
        total_moto_dia_dolares = round(total_moto_dia_dolares, 2)
        total_accesorio_dia_pesos = round(total_accesorio_dia_pesos, 2)
        total_accesorio_dia_dolares = round(total_accesorio_dia_dolares, 2)
        subtotal_pesos = total_moto_dia_pesos + total_accesorio_dia_pesos
        subtotal_pesos = round(subtotal_pesos, 2)
        subtotal_dolares = total_moto_dia_dolares + total_accesorio_dia_dolares
        subtotal_dolares = round(subtotal_dolares, 2)
        ingresos_extra = round(ingresos_extra, 2)
        egresos = round(egresos, 2)
        ingresos_extra_dolares = round(ingresos_extra_dolares, 2)
        egresos_dolares = round(egresos_dolares, 2)
        total_pesos = subtotal_pesos + ingresos_extra - egresos
        total_pesos = int(total_pesos)
        total_dolares = subtotal_dolares + ingresos_extra_dolares - egresos_dolares
        total_dolares = int(total_dolares)


        data = [
            {"tipo": "Motos", "venta": cantidad_vendidas_dia, "total_pesos": total_moto_dia_pesos, "total_dolares": total_moto_dia_dolares},
            {"tipo": "Accesorios", "venta": cantidad_vendidos_dia, "total_pesos": total_accesorio_dia_pesos, "total_dolares": total_accesorio_dia_dolares},
            # Datos adicionales como egresos, ingresos extra, totales y subtotales
            {"egresos":egresos,
          "ingresos_extra":ingresos_extra,
          "egresos_dolares":egresos_dolares,
          "ingresos_extra_dolares":ingresos_extra_dolares,
            "subtotal_pesos": subtotal_pesos,
          "subtotal_dolares":subtotal_dolares,
          "total_general_pesos":total_pesos,
          "total_general_dolares":total_dolares},
            {"egresos_detalle": data_egresos},  # Detalles de los egresos
            {"ingresos_extra_detalle": data_ingresos_extra},  # Detalles de los ingresos extra
            {"fecha": f_detalle_str}  # Fecha del reporte
        ]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Movimientos del Día"

        # Estilo para los encabezados
        header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")  # Amarillo
        header_font = Font(bold=True, color="000000")

        # Encabezados de la tabla de ventas
        headers = ["Tipo", "Cantidad Vendida", "Total Pesos", "Total Dólares"]
        ws.append(headers)

        # Aplicar estilo al encabezado
        for col_num, cell in enumerate(ws[1], 1):
            cell.fill = header_fill
            cell.font = header_font

        # Rellenar las filas con los datos de ventas
        for item in data:
            if "tipo" in item and "venta" in item and "total_pesos" in item and "total_dolares" in item:
                ws.append([item['tipo'], item['venta'], item['total_pesos'], item['total_dolares']])

        # Estilos para las celdas de la tabla de ventas
        for row in ws.iter_rows(min_row=2, max_row=len(data)+1, min_col=1, max_col=4):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if cell.row % 2 == 0:
                    cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Gris claro

        # Agregar los subtotales y totales generales debajo de la tabla de ventas

        ws.append([])  # Fila vacía para separar las secciones
        ws.append(["Datos en Pesos"])
        # ws.append([])
        ws.append(["Ingresos extra","Egresos","Subtotal", "Total"])
        # for item in data:
        # if "ingresos_extra" in item and "egresos" in item and "subtotal_pesos" in item and "total_pesos" in item:
        ws.append([ingresos_extra, egresos,subtotal_pesos,total_pesos])
                # ws.append([item['ingresos_extra'], item['egresos'],item['subtotal_pesos'],item['total_pesos']])

        ws.append([])  # Fila vacía para separar las secciones
        ws.append(["Datos en Dolares"])
        # ws.append([])
        ws.append(["Ingresos extra","Egresos","Subtotal", "Total"])
        # for item in data:
        # if "ingresos_extra_dolares" in item and "egresos_dolares" in item and "subtotal_dolares" in item and "total_dolares" in item:
        ws.append([ingresos_extra_dolares, egresos_dolares,subtotal_dolares,total_dolares])
                # ws.append(item['ingresos_extra_dolares'], item['egresos_dolares'],item['subtotal_dolares'],item['total_dolares']

        # Detalles de los egresos
        ws.append([])  # Fila vacía para separar las secciones
        ws.append(["Egresos Detalle"])
        # ws.append([])
        ws.append(["Monto","Motivo","Rubro"])
        for egreso in data_egresos:
            ws.append([egreso['valor'], egreso['motivo'],egreso['rubro']])

        # Detalles de los ingresos extra
        ws.append([])  # Fila vacía para separar las secciones
        ws.append(["Ingresos Extra Detalles"])
        # ws.append([])
        ws.append(["Monto","Motivo"])
        for ingreso in data_ingresos_extra:
            ws.append([ingreso['valor'], ingreso['motivo']])

        # Configuración de la respuesta para la descarga del archivo
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=Movimientos.xlsx'
        wb.save(response)
        return response

def buscar_detalles_movimientos_x_mes_anio(req):
    # try:
        print("PDF")
        mes = int(req.POST['mes_reporte'])
        anio = int(req.POST['anio_reporte'])
        dolar = PrecioDolar.objects.get(id=1)
        precio_dolar = float(dolar.precio_dolar_tienda)
        id_caja = req.session["id_caja_movimiento"]
        #CANTIDAD MOTOS VENDIDAS EN LA FECHA INGRESADA, TOTAL DEL INGRESO EN PESOS Y DOLARES 
        movs_motos_vendidas_dia = Movimientos.objects.filter(caja_id=id_caja,fecha__month=mes,fecha__year=anio, es_moto=1,metodo="Efectivo")
        cantidad_vendidas_dia = ComprasVentas.objects.filter(fecha_compra__month=mes,fecha_compra__year=anio,tipo="V").count()
        total_moto_dia_pesos = 0
        total_moto_dia_dolares = 0
        for movimiento in movs_motos_vendidas_dia:
            # moto = Moto.objects.get(id=movimiento.moto_id)
            if movimiento.tipo == "Ingreso":
                if movimiento.moneda == "Pesos":
                    total_moto_dia_pesos = total_moto_dia_pesos + float(movimiento.monto)
                    total_moto_dia_dolares = total_moto_dia_pesos / precio_dolar 
                elif movimiento.moneda == "Dolares":
                    total_moto_dia_dolares = total_moto_dia_dolares + float(movimiento.monto)
                    total_moto_dia_pesos = total_moto_dia_dolares * precio_dolar


        #CANTIDAD ACCESORIOS VENDIDOS EN LA FECHA INGRESADA, TOTAL DEL INGRESO EN PESOS Y DOLARES 
        movs_accesorios_vendidos_dia = Movimientos.objects.filter(caja_id=id_caja,fecha__month=mes,fecha__year=anio, es_accesorio=1,metodo="Efectivo")
        cantidad_vendidos_dia = ClienteAccesorio.objects.filter(fecha_compra__month=mes,fecha_compra__year=anio).count()
        total_accesorio_dia_pesos = 0
        total_accesorio_dia_dolares = 0
        for movimiento in movs_accesorios_vendidos_dia:
            if movimiento.tipo == "Ingreso":
                if movimiento.moneda == "Pesos":
                    total_accesorio_dia_pesos = total_accesorio_dia_pesos + float(movimiento.monto)
                    total_accesorio_dia_dolares = total_accesorio_dia_pesos / precio_dolar 
                elif movimiento.moneda == "Dolares":
                    total_accesorio_dia_dolares = total_accesorio_dia_dolares + float(movimiento.monto)
                    total_accesorio_dia_pesos = total_accesorio_dia_dolares * precio_dolar


        movimientos_dia = Movimientos.objects.filter(caja_id=id_caja,fecha__month=mes,fecha__year=anio)
        ingresos_extra = 0
        egresos = 0 
        ingresos_extra_dolares = 0
        egresos_dolares = 0 
        data_egresos = []
        data_ingresos_extra = []
        for mov in movimientos_dia:
            if mov.tipo == "Ingreso extra":
                if mov.moneda == "Pesos":
                    ingresos_extra = ingresos_extra + float(mov.monto)
                    ingresos_extra_dolares = ingresos_extra / precio_dolar
                else:
                    ingresos_extra_dolares = ingresos_extra_dolares + float(mov.monto)
                    ingresos_extra = ingresos_extra_dolares * precio_dolar
                data_ingresos_extra.append({
                    "valor":mov.monto,
                    "motivo":mov.movimiento
                })
            elif mov.tipo == "Egreso":
                if mov.moneda == "Pesos":
                    egresos = egresos + float(mov.monto)
                    egresos_dolares = egresos / precio_dolar
                else:
                    egresos_dolares = egresos_dolares + float(mov.monto)
                    egresos = egresos_dolares * precio_dolar
                data_egresos.append({
                    "rubro":mov.rubro,
                    "valor":mov.monto,
                    "motivo":mov.movimiento
                })
            
        meses = ["","Enero","Febrero","Marzo","Abril","Mayo","Junio","Julio","Agosto","Setiembre","Octubre","Noviembre","Diciembre"]
        total_moto_dia_pesos = round(total_moto_dia_pesos, 2)
        total_moto_dia_dolares = round(total_moto_dia_dolares, 2)
        total_accesorio_dia_pesos = round(total_accesorio_dia_pesos, 2)
        total_accesorio_dia_dolares = round(total_accesorio_dia_dolares, 2)
        subtotal_pesos = total_moto_dia_pesos + total_accesorio_dia_pesos
        subtotal_pesos = round(subtotal_pesos, 2)
        subtotal_dolares = total_moto_dia_dolares + total_accesorio_dia_dolares
        subtotal_dolares = round(subtotal_dolares, 2)
        ingresos_extra = round(ingresos_extra, 2)
        egresos = round(egresos, 2)
        ingresos_extra_dolares = round(ingresos_extra_dolares, 2)
        egresos_dolares = round(egresos_dolares, 2)
        total_pesos = subtotal_pesos + ingresos_extra - egresos
        total_pesos = int(total_pesos)
        total_dolares = subtotal_dolares + ingresos_extra_dolares - egresos_dolares
        total_dolares = int(total_dolares)

        print("TOTAL MOTOS DEL DIA EN PESOS: $" + str(total_moto_dia_pesos))
        print("TOTAL MOTOS DEL DIA EN DOLARES: U$s" + str(total_moto_dia_dolares))
        print("CANTIDAD MOTOS VENDIDOS: " + str(cantidad_vendidas_dia))
        print("TOTAL ACCESORIOS DEL DIA EN PESOS: $" + str(total_accesorio_dia_pesos))
        print("TOTAL ACCESORIOS DEL DIA EN DOLARES: U$s" + str(total_accesorio_dia_dolares))
        print("CANTIDAD ACCESORIOS VENDIDOS: " + str(cantidad_vendidos_dia))
        print("TOTAL EN PESOS: " + str(total_pesos))
        print("TOTAL EN DOLARES: " + str(total_dolares))
        print("EGRESOS: " + str(egresos))
        print("INGRESOS EXTRA: " + str(ingresos_extra))
        # fecha_json = datetime.strptime(f_detalle_str, '%d-%m-%Y').date() if f_detalle_str else None
        data = [
        {"tipo": "Motos", 
         "venta": cantidad_vendidas_dia, 
         "total_pesos": total_moto_dia_pesos, 
         "total_dolares": total_moto_dia_dolares},
        {"tipo": "Accesorios", 
         "venta": cantidad_vendidos_dia, 
         "total_pesos": total_accesorio_dia_pesos, 
         "total_dolares": total_accesorio_dia_dolares},
         {"egresos":egresos,
          "ingresos_extra":ingresos_extra,
          "egresos_dolares":egresos_dolares,
          "ingresos_extra_dolares":ingresos_extra_dolares,
            "subtotal_pesos": subtotal_pesos,
          "subtotal_dolares":subtotal_dolares,
          "total_general_pesos":total_pesos,
          "total_general_dolares":total_dolares},
          {"egresos_detalle": data_egresos},  # Pasamos la lista completa
            {"ingresos_extra_detalle": data_ingresos_extra},
            {"mes":meses[mes],
             "anio":anio}

        ]
        
        return JsonResponse(data, safe=False)

def buscar_detalles_movimientos_x_mes_anio_excel(req):
        print("EXCEL")
        mes = int(req.POST['mes_reporte'])
        anio = int(req.POST['anio_reporte'])
        dolar = PrecioDolar.objects.get(id=1)
        precio_dolar = float(dolar.precio_dolar_tienda)
        id_caja = req.session["id_caja_movimiento"]
        #CANTIDAD MOTOS VENDIDAS EN LA FECHA INGRESADA, TOTAL DEL INGRESO EN PESOS Y DOLARES 
        movs_motos_vendidas_dia = Movimientos.objects.filter(caja_id=id_caja,fecha__month=mes,fecha__year=anio, es_moto=1,metodo="Efectivo")
        cantidad_vendidas_dia = ComprasVentas.objects.filter(fecha_compra__month=mes,fecha_compra__year=anio,tipo="V").count()
        total_moto_dia_pesos = 0
        total_moto_dia_dolares = 0
        for movimiento in movs_motos_vendidas_dia:
            # moto = Moto.objects.get(id=movimiento.moto_id)
            if movimiento.tipo == "Ingreso":
                if movimiento.moneda == "Pesos":
                    total_moto_dia_pesos = total_moto_dia_pesos + float(movimiento.monto)
                    total_moto_dia_dolares = total_moto_dia_pesos / precio_dolar 
                elif movimiento.moneda == "Dolares":
                    total_moto_dia_dolares = total_moto_dia_dolares + float(movimiento.monto)
                    total_moto_dia_pesos = total_moto_dia_dolares * precio_dolar


        #CANTIDAD ACCESORIOS VENDIDOS EN LA FECHA INGRESADA, TOTAL DEL INGRESO EN PESOS Y DOLARES 
        movs_accesorios_vendidos_dia = Movimientos.objects.filter(caja_id=id_caja,fecha__month=mes,fecha__year=anio, es_accesorio=1,metodo="Efectivo")
        cantidad_vendidos_dia = ClienteAccesorio.objects.filter(fecha_compra__month=mes,fecha_compra__year=anio).count()
        total_accesorio_dia_pesos = 0
        total_accesorio_dia_dolares = 0
        for movimiento in movs_accesorios_vendidos_dia:
            if movimiento.tipo == "Ingreso":
                if movimiento.moneda == "Pesos":
                    total_accesorio_dia_pesos = total_accesorio_dia_pesos + float(movimiento.monto)
                    total_accesorio_dia_dolares = total_accesorio_dia_pesos / precio_dolar 
                elif movimiento.moneda == "Dolares":
                    total_accesorio_dia_dolares = total_accesorio_dia_dolares + float(movimiento.monto)
                    total_accesorio_dia_pesos = total_accesorio_dia_dolares * precio_dolar


        movimientos_dia = Movimientos.objects.filter(caja_id=id_caja,fecha__month=mes,fecha__year=anio,metodo="Efectivo")
        ingresos_extra = 0
        egresos = 0 
        ingresos_extra_dolares = 0
        egresos_dolares = 0 
        data_egresos = []
        data_ingresos_extra = []
        for mov in movimientos_dia:
            if mov.tipo == "Ingreso extra":
                if mov.moneda == "Pesos":
                    ingresos_extra = ingresos_extra + float(mov.monto)
                    ingresos_extra_dolares = ingresos_extra / precio_dolar
                else:
                    ingresos_extra_dolares = ingresos_extra_dolares + float(mov.monto)
                    ingresos_extra = ingresos_extra_dolares * precio_dolar
                data_ingresos_extra.append({
                    "valor":mov.monto,
                    "motivo":mov.movimiento
                })
            elif mov.tipo == "Egreso":
                if mov.moneda == "Pesos":
                    egresos = egresos + float(mov.monto)
                    egresos_dolares = egresos / precio_dolar
                else:
                    egresos_dolares = egresos_dolares + float(mov.monto)
                    egresos = egresos_dolares * precio_dolar
                data_egresos.append({
                    "rubro":mov.rubro,
                    "valor":mov.monto,
                    "motivo":mov.movimiento
                })
            
        
        total_moto_dia_pesos = round(total_moto_dia_pesos, 2)
        total_moto_dia_dolares = round(total_moto_dia_dolares, 2)
        total_accesorio_dia_pesos = round(total_accesorio_dia_pesos, 2)
        total_accesorio_dia_dolares = round(total_accesorio_dia_dolares, 2)
        subtotal_pesos = total_moto_dia_pesos + total_accesorio_dia_pesos
        subtotal_pesos = round(subtotal_pesos, 2)
        subtotal_dolares = total_moto_dia_dolares + total_accesorio_dia_dolares
        subtotal_dolares = round(subtotal_dolares, 2)
        ingresos_extra = round(ingresos_extra, 2)
        egresos = round(egresos, 2)
        ingresos_extra_dolares = round(ingresos_extra_dolares, 2)
        egresos_dolares = round(egresos_dolares, 2)
        total_pesos = subtotal_pesos + ingresos_extra - egresos
        total_pesos = int(total_pesos)
        total_dolares = subtotal_dolares + ingresos_extra_dolares - egresos_dolares
        total_dolares = int(total_dolares)

        print("TOTAL MOTOS DEL DIA EN PESOS: $" + str(total_moto_dia_pesos))
        print("TOTAL MOTOS DEL DIA EN DOLARES: U$s" + str(total_moto_dia_dolares))
        print("CANTIDAD MOTOS VENDIDOS: " + str(cantidad_vendidas_dia))
        print("TOTAL ACCESORIOS DEL DIA EN PESOS: $" + str(total_accesorio_dia_pesos))
        print("TOTAL ACCESORIOS DEL DIA EN DOLARES: U$s" + str(total_accesorio_dia_dolares))
        print("CANTIDAD ACCESORIOS VENDIDOS: " + str(cantidad_vendidos_dia))
        print("TOTAL EN PESOS: " + str(total_pesos))
        print("TOTAL EN DOLARES: " + str(total_dolares))
        print("EGRESOS: " + str(egresos))
        print("INGRESOS EXTRA: " + str(ingresos_extra))


        data = [
            {"tipo": "Motos", "venta": cantidad_vendidas_dia, "total_pesos": total_moto_dia_pesos, "total_dolares": total_moto_dia_dolares},
            {"tipo": "Accesorios", "venta": cantidad_vendidos_dia, "total_pesos": total_accesorio_dia_pesos, "total_dolares": total_accesorio_dia_dolares},
            # Datos adicionales como egresos, ingresos extra, totales y subtotales
            {"egresos":egresos,
          "ingresos_extra":ingresos_extra,
          "egresos_dolares":egresos_dolares,
          "ingresos_extra_dolares":ingresos_extra_dolares,
            "subtotal_pesos": subtotal_pesos,
          "subtotal_dolares":subtotal_dolares,
          "total_general_pesos":total_pesos,
          "total_general_dolares":total_dolares},
            {"egresos_detalle": data_egresos},  # Detalles de los egresos
            {"ingresos_extra_detalle": data_ingresos_extra},  # Detalles de los ingresos extra
              # Fecha del reporte
        ]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Movimientos del Día"

        # Estilo para los encabezados
        header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")  # Amarillo
        header_font = Font(bold=True, color="000000")

        # Encabezados de la tabla de ventas
        headers = ["Tipo", "Cantidad Vendida", "Total Pesos", "Total Dólares"]
        ws.append(headers)

        # Aplicar estilo al encabezado
        for col_num, cell in enumerate(ws[1], 1):
            cell.fill = header_fill
            cell.font = header_font

        # Rellenar las filas con los datos de ventas
        for item in data:
            if "tipo" in item and "venta" in item and "total_pesos" in item and "total_dolares" in item:
                ws.append([item['tipo'], item['venta'], item['total_pesos'], item['total_dolares']])

        # Estilos para las celdas de la tabla de ventas
        for row in ws.iter_rows(min_row=2, max_row=len(data)+1, min_col=1, max_col=4):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if cell.row % 2 == 0:
                    cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Gris claro

        # Agregar los subtotales y totales generales debajo de la tabla de ventas

        ws.append([])  # Fila vacía para separar las secciones
        ws.append(["Datos en Pesos"])
        # ws.append([])
        ws.append(["Ingresos extra","Egresos","Subtotal", "Total"])
        # for item in data:
        # if "ingresos_extra" in item and "egresos" in item and "subtotal_pesos" in item and "total_pesos" in item:
        ws.append([ingresos_extra, egresos,subtotal_pesos,total_pesos])
                # ws.append([item['ingresos_extra'], item['egresos'],item['subtotal_pesos'],item['total_pesos']])

        ws.append([])  # Fila vacía para separar las secciones
        ws.append(["Datos en Dolares"])
        # ws.append([])
        ws.append(["Ingresos extra","Egresos","Subtotal", "Total"])
        # for item in data:
        # if "ingresos_extra_dolares" in item and "egresos_dolares" in item and "subtotal_dolares" in item and "total_dolares" in item:
        ws.append([ingresos_extra_dolares, egresos_dolares,subtotal_dolares,total_dolares])
                # ws.append(item['ingresos_extra_dolares'], item['egresos_dolares'],item['subtotal_dolares'],item['total_dolares']

        # Detalles de los egresos
        ws.append([])  # Fila vacía para separar las secciones
        ws.append(["Egresos Detalle"])
        # ws.append([])
        ws.append(["Monto","Motivo","Rubro"])
        for egreso in data_egresos:
            ws.append([egreso['valor'], egreso['motivo'],egreso['rubro']])

        # Detalles de los ingresos extra
        ws.append([])  # Fila vacía para separar las secciones
        ws.append(["Ingresos Extra Detalles"])
        # ws.append([])
        ws.append(["Monto","Motivo"])
        for ingreso in data_ingresos_extra:
            ws.append([ingreso['valor'], ingreso['motivo']])

        # Configuración de la respuesta para la descarga del archivo
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=Movimientos.xlsx'
        wb.save(response)
        return response


def baja_movimiento(req,id_caja,id_movimiento):
    try:
        if req.method == "POST":
            caja = Caja.objects.get(id=id_caja)
            movimiento = Movimientos.objects.get(id=id_movimiento)
            if movimiento.moneda == "Pesos":
                monto_a_quitar = float(movimiento.monto)
            else:
                dolar = PrecioDolar.objects.get(id=1)
                precio_dolar = float(dolar.precio_dolar_tienda)
                monto_a_quitar = float(movimiento.monto) * precio_dolar
            
            depositos_caja = float(caja.depositos)
            egresos_caja = float(caja.egresos)
            
            if movimiento.tipo == "Ingreso" or movimiento.tipo == "Ingreso extra":
                sumar_quitar_depositos = depositos_caja - monto_a_quitar
                caja.depositos = abs(sumar_quitar_depositos)
                # quitar_egresos = egresos_caja
            else:
                # sumar_quitar_depositos = depositos_caja + monto_a_quitar
                quitar_egresos = egresos_caja - monto_a_quitar
                caja.egresos = quitar_egresos
            
            if int(caja.diferencia) != 0:
                if movimiento.tipo == "Ingreso" or movimiento.tipo == "Ingreso extra":
                    nueva_diferencia = float(caja.diferencia) - monto_a_quitar
                else:
                    nueva_diferencia = float(caja.diferencia) + monto_a_quitar
                
                caja.diferencia = nueva_diferencia
            
            
            caja.save() 
            movimiento.delete()
            # messages.success(req, "Movimiento borrado con éxito")
            # return redirect(f"{reverse('MovimientosCaja',kwargs={'id_caja':id_caja})}")
            return render(req, "perfil_administrativo/arqueos/baja_movimiento.html", {"id_caja":id_caja,"message":"Movimiento borrado con éxito."})
        else:
            return render(req, "perfil_administrativo/arqueos/baja_movimiento.html", {"id_caja":id_caja})
    except Exception as e:
        pass

def nuevo_rubro_egreso(req,id_caja):
    try:
        rubro = req.POST['nuevo_rubro']
        nuevo_rubro = Rubros(
            rubro = rubro,
            habilitado = 1
        )
        nuevo_rubro.save()
        messages.success(req, "Rubro ingresado con éxito")
        return redirect(f"{reverse('MovimientosCaja',kwargs={'id_caja':id_caja})}")
    except Exception as e:
        pass

def servicios_en_gestion(req):
    # try:
        servicios_motos = (
        Servicios.objects
        .all()
        .select_related('moto', 'cliente')
        .values(
            'id',
            'fecha_ingreso', 
            'estado',
            'prioridad',
            'dias',
            'titulo',
            'moto__marca', 
            'moto__modelo',
            'cliente__nombre',
            'cliente__apellido'      
           ).filter(estado__in=["Pendiente","En espera"])
           .order_by('-id')
        )
        usuario = req.user
        usuario_actual = Personal.objects.filter(usuario=usuario.username).first()
        mecanico_usuario = Mecanico.objects.get(id=usuario_actual.id)
        if mecanico_usuario.jefe:
            jefe = True
        else:
            jefe = False
        datos = []
        for resultado in servicios_motos:
            mecanicos = MecanicosServicios.objects.filter(servicio_id=resultado['id'])
            datos_mecanicos = [
                f"{persona.nombre} {persona.apellido}"
                for mecanico in mecanicos
                for persona in Personal.objects.filter(id=mecanico.mecanico_id)
            ]

            cambiar_nombre_variable = MecanicosServicios.objects.filter(servicio_id=resultado['id'],mecanico_id = mecanico_usuario.id).first()
            if cambiar_nombre_variable or jefe:
                mostrar_boton = True
            else:
                mostrar_boton = False
            fecha_ingreso = resultado['fecha_ingreso']  # Suponiendo que 'fecha_ingreso' es un campo de la consulta
            dias_transcurridos = (date.today() - fecha_ingreso).days
            datos.append({
                'servicio': resultado if resultado else None,
                'mecanicos': datos_mecanicos if datos_mecanicos else None,  # Lista de nombres completos de mecánicos
                'dias':dias_transcurridos,
                "mostrar_boton":mostrar_boton
            })
        page_obj = funcion_paginas_varias(req,datos)
        return render(req,"perfil_taller/servicios/servicios_en_gestion.html",{"servicios":datos,"page_obj":page_obj})
    # except Exception as e:
    #     pass

def form_alta_servicio(req):
    try:
        return render(req,"perfil_taller/servicios/alta_servicio.html",{"buscar_moto_cliente":True,"div_mecanicos":False})
    except Exception as e:
        return render(req,"perfil_taller/servicios/alta_servicio.html",{"error_message":e,"buscar_moto_cliente":True})

def cliente_moto_servicio(req):
    # try:
        documento = req.POST['tipo_documento'] + str(req.POST['documento'])
        cliente = Cliente.objects.filter(documento=documento).first()
        tipo_busqueda = req.POST['buscar_moto_por']
        matricula = req.POST['letras_matricula'].upper() + str(req.POST['numeros_matricula'])
        if tipo_busqueda == "matricula":
            existe_matricula = Matriculas.objects.filter(matricula=matricula).first()
            if existe_matricula:
                moto = Moto.objects.get(id=existe_matricula.moto_id)
                matricula = existe_matricula.matricula
            else:
                moto = None
                matricula = None
        else:
            moto = Moto.objects.filter(num_motor=req.POST['numero_de_motor']).first()
            if moto:
                existe_matricula = Matriculas.objects.filter(matricula=matricula).first()
                if existe_matricula:
                    matricula = existe_matricula.matricula
                else:
                    matricula = None
            else:
                moto = None
                matricula = None
            
        
        if not cliente:
            return render(req,"perfil_taller/servicios/alta_servicio.html",{"error_message_cliente":"El cliente no existe, para ingresarlo haga clic ","buscar_moto_cliente":True})
        elif not moto:
            return render(req,"perfil_taller/servicios/alta_servicio.html",{"error_message":"La moto no se encuentra registrada en el sistema, debe ingresar la misma completando todos los datos requeridos","buscar_moto_cliente":True})
        else:
            usuario = req.user
            usuario_actual = Personal.objects.filter(usuario=usuario.username).first()
            mecanico_actual = Mecanico.objects.get(id=usuario_actual.id)
            if mecanico_actual.jefe:
                div_mecanicos = True
                mecanicos = (Mecanico.objects
                        .filter(activo=True)
                        .values('id', 'nombre', 'apellido')
                        .order_by('nombre'))
            else:
                div_mecanicos = False
                mecanicos = False
            
            correo = ClienteCorreo.objects.filter(cliente=cliente,principal=1).first()
            telefono = ClienteTelefono.objects.filter(cliente=cliente,principal=1).first()
            return render(req,"perfil_taller/servicios/alta_servicio.html",{"datos_moto":True,
                                                                            "moto":moto,
                                                                            "cliente":cliente,
                                                                            "matricula":matricula if matricula else "Sin matrícula",
                                                                            "buscar_moto_cliente":False,
                                                                            "mecanicos":mecanicos if mecanicos else None,
                                                                            "div_mecanicos":div_mecanicos,
                                                                            "correo":correo.correo if correo else "El cliente no tiene correo",
                                                                            "telefono":telefono.telefono})
    # except Exception as e:
    #     return render(req,"perfil_taller/servicios/alta_servicio.html",{"error_message":"La moto no se encuentra registrada en el sistema, debe ingresar la misma completando todos los datos requeridos"})

def alta_servicio(req,id_moto,id_cliente):
    # try:
        # print(id_moto)
        # print(id_cliente)piezas[]
        servicios_seleccionados = req.POST.getlist('servicios[]')
        mecanicos_seleccionados = req.POST.getlist('mecanicos[]')
        piezas_seleccionadas = req.POST.getlist('piezas_seleccionadas[]')
        cantidades = req.POST.getlist('cantidad_piezas[]')
        mecanicos = (Mecanico.objects
                       .filter(activo=True)
                       .values('id', 'nombre', 'apellido')
                       .order_by('nombre'))
        usuario = req.user
        usuario_actual = Personal.objects.filter(usuario=usuario.username).first()
        mecanico_actual = Mecanico.objects.get(id=usuario_actual.id)
        if piezas_seleccionadas:
            i = 0
            for pieza in piezas_seleccionadas:
                rp = RepuestosPiezas.objects.get(id=int(pieza))
                if int(cantidades[i]) > int(rp.stock):
                    error_pieza = True
                    break
                else:
                    error_pieza = False
                    break
        else:
            error_pieza = False

        moto = Moto.objects.get(id=id_moto)
        cliente = Cliente.objects.get(id=id_cliente)
        usuario = req.user
        usuario_actual = Personal.objects.filter(usuario=usuario.username).first()
        mecanico_actual = Mecanico.objects.get(id=usuario_actual.id)
        if mecanico_actual.jefe:
            div_mecanicos = True
            mecanicos = (Mecanico.objects
                    .filter(activo=True)
                    .values('id', 'nombre', 'apellido')
                    .order_by('nombre'))
        else:
            div_mecanicos = False
            mecanicos = False
        # if not servicios_seleccionados:
        #     return render(req,"perfil_taller/servicios/alta_servicio.html",{"error_message_alta":"No has seleccionado ningún servicio","datos_moto":True,"mecanicos":mecanicos if mecanicos else None,"moto":moto,"cliente":cliente,"div_mecanicos":div_mecanicos})
        # el
        if not mecanicos_seleccionados and mecanico_actual.jefe:   
            return render(req,"perfil_taller/servicios/alta_servicio.html",{"error_message_alta":"No has seleccionado ningún mecánico","datos_moto":True,"mecanicos":mecanicos if mecanicos else None,"moto":moto,"cliente":cliente,"div_mecanicos":div_mecanicos})
        elif error_pieza:
            return render(req,"perfil_taller/servicios/alta_servicio.html",{"error_message_alta":"Algunos de los repuestos o piezas exceden el stock","datos_moto":True,"mecanicos":mecanicos if mecanicos else None,"moto":moto,"cliente":cliente,"div_mecanicos":div_mecanicos})
        else:
            nuevo_servicio = Servicios(
                fecha_ingreso = datetime.now(),
                descripcion_ingreso = req.POST['anotaciones'],
                estado = "Pendiente",
                prioridad = req.POST['prioridad'],
                fecha_estimada_cierre = req.POST['fecha_estimada'],
                cliente_id = id_cliente,
                moto_id = id_moto,
                titulo = req.POST['titulo_servicio']
            )
            nuevo_servicio.save()
           #inputPrecio.name = `precio_${textoServicio}`;

            if piezas_seleccionadas:
                i = 0
                for pieza in piezas_seleccionadas:
                    rp = RepuestosPiezas.objects.get(id=int(pieza))
                    nueva_pieza = RepuestosPiezasServicios(
                        repuestopieza_id = pieza,
                        servicio_id = nuevo_servicio.id,
                        cantidad = cantidades[i]
                    )
                    nueva_pieza.save()
                    nuevo_stock = int(rp.stock) - int(cantidades[i])
                    rp.stock = nuevo_stock
                    rp.save()
                    if rp.stock_critico >= rp.stock:
                        insert_notificaciones(f"Hay poco stock de {rp.descripcion}","Bajo stock de pieza")
                    i = i + 1



            if servicios_seleccionados:
                for servicio in servicios_seleccionados:
                    # prueba = servicio.replace(" ","_")
                    # print(prueba)
                    # precio = req.POST[f'precio_{prueba}']
                    # print("PRUEBA: " + str(precio))
                    tareas_servicios = TareasServicios(
                        tarea = servicio,
                        servicio_id = nuevo_servicio.id,
                        # precio = precio
                    )
                    tareas_servicios.save()
            
            if not mecanico_actual.jefe:
                mecanicos_servicio = MecanicosServicios(
                        mecanico_id = mecanico_actual.id,
                        servicio_id = nuevo_servicio.id
                    )
                mecanicos_servicio.save()
            else:
                for mecanico in mecanicos_seleccionados:
                    mecanicos_servicio = MecanicosServicios(
                        mecanico_id = mecanico,
                        servicio_id = nuevo_servicio.id
                    )
                    mecanicos_servicio.save()
        
            messages.success(req, "Servicio ingresado correctamente.")
            return redirect('ServiciosEnGestion')
    # except Exception as e:
    #     pass

def cerrar_servicio(req,id_s):
    # try:
        servicio = Servicios.objects.get(id=id_s)
        if req.method == "POST":
            servicio.estado = "Completado"
            if req.POST['f_prox_servicio']:
                fecha = req.POST.get('f_prox_servicio')  # Cambiado a paréntesis
                f_prox_service = datetime.strptime(fecha, '%Y-%m-%d').date() if fecha else None
                servicio.fecha_prox_servicio = f_prox_service
            if req.POST['km_prox_servicio']:
                servicio.km_prox_servicio = req.POST['km_prox_servicio']
            
            servicio.precio = req.POST['precio_servicio']
            servicio.save()
            anotacion = req.POST['anotacion_cierre']
            usuario = req.user
            usuario_actual = Personal.objects.filter(usuario=usuario.username).first()
            if anotacion:
                anotacion_editada = "Servicio cerrado con la siguiente descripción: " + anotacion
            else:
                anotacion_editada = "Servicio cerrado sin descripción"
            ultima_anotacion = AnotacionesServicio(
                anotaciones = anotacion_editada,
                fecha = datetime.now(),
                mecanico_id = usuario_actual.id,
                servicio_id = id_s
            )
            ultima_anotacion.save()
            messages.success(req, "Servicio cerrado correctamente.")
            return redirect('ServiciosEnGestion')
        else:
            datos = json_para_servicio(id_s)
            return render(req,"perfil_taller/servicios/cerrar_servicio.html",{"datos_fijos":datos[0],"tareas":datos[1]})
    # except Exception as e:
    #     pass

def contexto_modificar_servicio(id_s,mensaje):
    servicio = Servicios.objects.get(id=id_s)
    cliente = Cliente.objects.get(id=servicio.cliente_id)
    moto = Moto.objects.get(id=servicio.moto_id)
    matricula = Matriculas.objects.filter(moto_id=servicio.moto_id).first()
    telefono = ClienteTelefono.objects.filter(cliente_id=servicio.cliente_id,principal=1).first()
    correo = ClienteCorreo.objects.filter(cliente_id=servicio.cliente_id,principal=1).first()
    tareas_realizadas = TareasServicios.objects.filter(servicio_id=id_s,realizado=1)
    tareas_pendientes = TareasServicios.objects.filter(servicio_id=id_s,realizado=0)

    data_mecanicos_servicio = []
    mecanicos_en_el_servicio = MecanicosServicios.objects.filter(servicio_id=id_s)

    for mecanico in mecanicos_en_el_servicio:
            mec = Mecanico.objects.filter(id=mecanico.mecanico_id,activo=1).first()
            # resto_de_mecanicos = Mecanico.objects.exclude(id=mecanico.mecanico_id)
            # for demas_mecanicos in resto_de_mecanicos:
            #     mecanico_otro = Mecanico.objects.get(id=demas_mecanicos.id)
            data_mecanicos_servicio.append({
                    "mecanico":mec,
                    # "resto_mecanicos":mecanico_otro
                })
    
    # resto_mecanicos = MecanicosServicios.objects.exclude(servicio_id=id_s)
    resto_mecanicos = Mecanico.objects.filter(activo=1)
    data_mecanicos_no_servicio = []
    for mecanico in resto_mecanicos:
        # mec = Mecanico.objects.get(id=mecanico.mecanico_id)
        mecanico_en_este_servicio = MecanicosServicios.objects.filter(servicio_id=id_s,mecanico=mecanico).first()
        if not mecanico_en_este_servicio:
            data_mecanicos_no_servicio.append({
                        # "mecanico":mec,
                        "resto_mecanicos":mecanico
                    })
    
    anotaciones = AnotacionesServicio.objects.filter(servicio_id=id_s)
    data_anotaciones = []
    for anotacion in anotaciones:
        mecanico = Mecanico.objects.get(id=anotacion.mecanico_id)
        data_anotaciones.append({
            "anotacion":anotacion,
            "mecanico":mecanico
        })

    f_cierre_formateada = servicio.fecha_estimada_cierre.strftime('%Y-%m-%d') if servicio.fecha_estimada_cierre else None

    piezas = RepuestosPiezasServicios.objects.filter(servicio_id=id_s)
    data_piezas = []
    for pieza in piezas:
        rp = RepuestosPiezas.objects.filter(id=pieza.repuestopieza_id).first()
        data_piezas.append({
            "piezas":rp,
            "cantidad":pieza.cantidad,
            "ide":pieza.id
        })
    data = [
        moto,
        cliente,
        matricula.matricula if matricula else "Sin matricula",
        telefono.telefono if telefono else "El cliente no cuenta con teléfono",
        correo.correo if correo else "El cliente no cuenta con correo",
        tareas_realizadas,
        tareas_pendientes,
        data_mecanicos_servicio,
        mensaje if mensaje else None,
        id_s,
        data_anotaciones,
        data_mecanicos_no_servicio,
        servicio,
        f_cierre_formateada,
        data_piezas
    ]

    return data

def form_modificar_servicio(req,id_s):
    contexto = contexto_modificar_servicio(id_s,None)
    usuario = req.user
    usuario_actual = Personal.objects.filter(usuario=usuario.username).first()
    mecanico_usuario = Mecanico.objects.get(id=usuario_actual.id)
    return render(req,"perfil_taller/servicios/modificar_servicio.html",{ "moto":contexto[0],
                                                                            "cliente":contexto[1],
                                                                            "matricula":contexto[2],
                                                                            "telefono":contexto[3],
                                                                            "correo":contexto[4],
                                                                            "tareas_realizadas":contexto[5],
                                                                            "tareas_pendientes":contexto[6],
                                                                            "mecanicos":contexto[7],
                                                                            "id_servicio":contexto[9],
                                                                            "anotaciones":contexto[10],
                                                                            "resto_mecanicos":contexto[11],
                                                                            "info_servicio":contexto[12],
                                                                            "fecha_cierre":contexto[13],
                                                                            "mostrar_boton":True if mecanico_usuario.jefe else False,
                                                                            "piezas":contexto[14]
                                                                            })


def agregar_quitar_servicios(req,id_s):
    try:
        servicios_seleccionados = req.POST.getlist('servicios_pendientes[]')
        nuevos_servicios = req.POST.getlist('nuevos_servicios[]')

        if nuevos_servicios:
            serv = Servicios.objects.get(id=id_s)
            for servicio in nuevos_servicios:
                # servicio = servicio.replace("_", " ")
                # prueba = servicio.replace(" ","_")
                # precio = req.POST[f'precio_{prueba}']
                nuevo_servicio = TareasServicios(
                    servicio = serv,
                    tarea = servicio,
                    realizado = 0,
                    # precio = precio
                )
                nuevo_servicio.save()
        
        for servicio in servicios_seleccionados:
            servicio_realizado = TareasServicios.objects.get(id=servicio)
            servicio_realizado.realizado = 1
            servicio_realizado.save()
        messages.success(req, "Servicios actualizados correctamente.")
        return redirect(reverse('FormModificarServicio', kwargs={'id_s': id_s}))
    except Exception as e:
        pass

def borrar_servicio(req,id_tarea):
    try:
        servicio = TareasServicios.objects.get(id=id_tarea)
        id_s = servicio.servicio_id
        servicio.delete()
        messages.success(req, "Servicios actualizados correctamente.")
        return redirect(reverse('FormModificarServicio', kwargs={'id_s': id_s}))
    except Exception as e:
        pass

def agregar_mecanico_servicio(req,id_s):
    try:
        mecanicos_seleccionados = req.POST.getlist('mecanicos_seleccionados[]')
        if not mecanicos_seleccionados:
            pass
        else:
            for mecanico in mecanicos_seleccionados:
                # servicio = Servicios.objects.get(id=id_s)
                # print("PRUEBA: " + str(mecanico))
                nuevo_mecanico = MecanicosServicios(
                    servicio_id = id_s,
                    mecanico_id = mecanico
                )
                nuevo_mecanico.save()
            messages.success(req, "Mecanicos actualizados correctamente.")
            return redirect(reverse('FormModificarServicio', kwargs={'id_s': id_s}))
    except Exception as e:
        pass

def borrar_mecanico_servicio(req,id_s,id_mecanico):
    try:
        mecanico = MecanicosServicios.objects.filter(servicio_id=id_s,mecanico_id=id_mecanico).first()
        mecanico.delete()
        messages.success(req, "Mecanicos actualizados correctamente.")
        return redirect(reverse('FormModificarServicio', kwargs={'id_s': id_s}))
    except Exception as e:
        pass


def agregar_anotacion_servicio(req,id_s):
    # try:
        usuario = req.user
        persona = Personal.objects.filter(usuario=usuario.username).first()
        nueva_anotacion = AnotacionesServicio(
            anotaciones = req.POST['anotaciones'],
            fecha = datetime.now(),
            mecanico_id = persona.id,
            servicio_id = id_s
        )
        nueva_anotacion.save()
        messages.success(req, "Anotación agregada correctamente.")
        return redirect(reverse('FormModificarServicio', kwargs={'id_s': id_s}))
    # except Exception as e:
    #     pass

def modificar_datos_servicio(req,id_s):
    try:
        servicio = Servicios.objects.get(id=id_s)
        servicio.titulo = req.POST['titulo_servicio']
        servicio.prioridad = req.POST['prioridad']
        servicio.fecha_estimada_cierre = req.POST['fecha_estimada']
        servicio.save()
        
        messages.success(req, "Datos del servicio modificados correctamente.")
        
        # messages.success(req, "PRUEBA")
        return redirect(reverse('FormModificarServicio', kwargs={'id_s': id_s}))

    except Exception as e:
        pass

def agregar_repuesto_pieza_servicio(req,id_s):
    # try:
        piezas_seleccionadas = req.POST.getlist('piezas_seleccionadas[]')
        cantidades = req.POST.getlist('cantidad_piezas[]')
        if piezas_seleccionadas:
            i = 0
            for pieza in piezas_seleccionadas:
                rp = RepuestosPiezas.objects.get(id=int(pieza))
                if int(cantidades[i]) > int(rp.stock):
                    messages.error(req, "Algunos elementos no pudieron asignarse al servicio dado que exceden su stock.")
                    return redirect(reverse('FormModificarServicio', kwargs={'id_s': id_s}))
                else:
                    existe_rp = RepuestosPiezasServicios.objects.filter(repuestopieza_id=pieza,servicio_id=id_s).first()
                    if existe_rp:
                        nueva_cantidad = int(existe_rp.cantidad) + int(cantidades[i])
                        existe_rp.cantidad = nueva_cantidad
                        existe_rp.save()
                    else:
                        nueva_pieza = RepuestosPiezasServicios(
                            repuestopieza_id = pieza,
                            servicio_id = id_s,
                            cantidad = cantidades[i]
                        )
                        nueva_pieza.save()
                    nuevo_stock = int(rp.stock) - int(cantidades[i])
                    rp.stock = nuevo_stock
                    rp.save()
                    if int(rp.stock_critico) >= int(rp.stock):
                        insert_notificaciones(f"Hay poco stock de {rp.descripcion}","Bajo stock de pieza")
                    
                # print("CANTIDAD DE " + str(pieza) + " ES " + str(cantidades[i]))
                i = i + 1
            messages.success(req, "Datos del servicio modificados correctamente.")
        else:
            messages.error(req, "Debe seleccionar alguna pieza o repuesto.")
        return redirect(reverse('FormModificarServicio', kwargs={'id_s': id_s}))
        
    # except Exception as e:
    #     pass

def borrar_repuesto_pieza_servicio(req,id_rp):
        rp = RepuestosPiezasServicios.objects.get(id=id_rp)
        id_s = rp.servicio_id
    # try:
        repuesto_pieza = RepuestosPiezas.objects.get(id=rp.repuestopieza_id)
        repuesto_pieza.stock = int(repuesto_pieza.stock) + int(rp.cantidad)
        repuesto_pieza.save()
        rp.delete()
        messages.success(req, "Datos del servicio modificados correctamente.")
        return redirect(reverse('FormModificarServicio', kwargs={'id_s': id_s}))
    # except Exception as e:
    #     contexto = contexto_modificar_servicio(id_s,None)
    #     usuario = req.user
    #     usuario_actual = Personal.objects.filter(usuario=usuario.username).first()
    #     mecanico_usuario = Mecanico.objects.get(id=usuario_actual.id)
    #     return render(req,"perfil_taller/servicios/modificar_servicio.html",{ "moto":contexto[0],
    #                                                                             "cliente":contexto[1],
    #                                                                             "matricula":contexto[2],
    #                                                                             "telefono":contexto[3],
    #                                                                             "correo":contexto[4],
    #                                                                             "tareas_realizadas":contexto[5],
    #                                                                             "tareas_pendientes":contexto[6],
    #                                                                             "mecanicos":contexto[7],
    #                                                                             "id_servicio":contexto[9],
    #                                                                             "anotaciones":contexto[10],
    #                                                                             "resto_mecanicos":contexto[11],
    #                                                                             "info_servicio":contexto[12],
    #                                                                             "fecha_cierre":contexto[13],
    #                                                                             "mostrar_boton":True if mecanico_usuario.jefe else False,
    #                                                                             "piezas":contexto[14],
    #                                                                             "error_message":str(e)
    #                                                                             })

def detalles_servicios(req,id_s):
    contexto = contexto_modificar_servicio(id_s,None)
    return render(req,"perfil_taller/servicios/detalles_servicio.html",{ "moto":contexto[0],
                                                                            "cliente":contexto[1],
                                                                            "matricula":contexto[2],
                                                                            "telefono":contexto[3],
                                                                            "correo":contexto[4],
                                                                            "tareas_realizadas":contexto[5],
                                                                            "tareas_pendientes":contexto[6],
                                                                            "mecanicos":contexto[7],
                                                                            "id_servicio":contexto[9],
                                                                            "anotaciones":contexto[10],
                                                                            "info_servicio":contexto[12],
                                                                            "fecha_cierre":contexto[13],
                                                                            })


def historial_de_servicios(req):
    try:
        servicios_motos = (
        Servicios.objects
        .all()
        .select_related('moto', 'cliente')
        .filter(estado="Completado")
        .values(
            'id',
            'fecha_ingreso', 
            'estado',
            'prioridad',
            'dias',
            'moto__marca', 
            'moto__modelo',
            'cliente__nombre',
            'cliente__apellido'      
           ).order_by('-fecha_ingreso')
        )

        datos = []
        for resultado in servicios_motos:
            mecanicos = MecanicosServicios.objects.filter(servicio_id=resultado['id'])
            datos.append({
            'servicio': resultado,
            })
        page_obj = funcion_paginas_varias(req,datos)
        return render(req,"perfil_taller/historial_de_servicios/historial_de_servicios.html",{"servicios":datos,"page_obj":page_obj})
    except Exception as e:
        pass

def detalles_servicios_cerrados(req,id_s):
    contexto = contexto_modificar_servicio(id_s,None)
    datos = json_para_servicio(id_s)
    return render(req,"perfil_taller/historial_de_servicios/detalles_servicio.html",{ "moto":contexto[0],
                                                                            "cliente":contexto[1],
                                                                            "matricula":contexto[2],
                                                                            "telefono":contexto[3],
                                                                            "correo":contexto[4],
                                                                            "tareas_realizadas":contexto[5],
                                                                            "tareas_pendientes":contexto[6],
                                                                            "mecanicos":contexto[7],
                                                                            "id_servicio":contexto[9],
                                                                            "anotaciones":contexto[10],
                                                                            "info_servicio":contexto[12],
                                                                            "fecha_cierre":contexto[13],
                                                                            "datos_fijos":datos[0],
                                                                            "tareas":datos[1]
                                                                            })

def repuestos(req):
    repuestos = RepuestosPiezas.objects.filter(activo=1).order_by('-id')
    #repuestos = RepuestosPiezas.objects.all()
    page_obj = funcion_paginas_varias(req,repuestos)
    return render(req,"perfil_taller/repuestos/repuestos.html",{'page_obj': page_obj})

def buscar_pieza(request):
    query = request.GET.get('query', '')
    if query:
        piezas = RepuestosPiezas.objects.filter(descripcion__icontains=query)[:10]  # Ajusta el campo a filtrar
        data = [{'id': pieza.id, 'nombre': pieza.descripcion, 'stock': pieza.stock} for pieza in piezas]
    else:
        data = []
    return JsonResponse(data, safe=False)

def alta_repuesto(req):
    try:
        if req.method == "POST":
            stock = int(req.POST['stock_repuesto'])
            stock_critico = int(req.POST['stock_critico'])
            precio_repuesto = int(req.POST['precio_repuesto'])
            if stock <= 0:
                return render(req,"perfil_taller/repuestos/alta_repuesto.html",{"error_message":"El stock ingresado es incorrecto"})
            elif precio_repuesto < 0:
                return render(req,"perfil_taller/repuestos/alta_repuesto.html",{"error_message":"El precio ingresado es incorrecto"})
            elif stock_critico < 0:
                return render(req,"perfil_taller/repuestos/alta_repuesto.html",{"error_message":"El stock crítico ingresado es incorrecto"})
            else:
                foto = req.FILES.get('foto_repuesto')
                nuevo_repuesto = RepuestosPiezas(
                    stock = stock,
                    descripcion = req.POST['descripcion_repuesto'],
                    activo = 1,
                    precio = precio_repuesto,
                    foto = foto,
                    stock_critico = stock_critico
                )
                nuevo_repuesto.save()
                messages.success(req, "Repuesto ingresado con éxito.")
                return redirect('Repuestos')
        else:
            return render(req,"perfil_taller/repuestos/alta_repuesto.html",{})
    except Exception as e:
        pass

def baja_repuesto(req,id_rp):
    try:
        if req.method == "POST":
            repuesto = RepuestosPiezas.objects.get(id=id_rp)
            repuesto.activo = 0
            repuesto.save()
            # return render(req,"perfil_taller/repuestos/baja_repuesto.html",{})
            return render(req, "perfil_taller/repuestos/baja_repuesto.html", {"message":"Repuesto borrado con éxito"})
        else:
            return render(req,"perfil_taller/repuestos/baja_repuesto.html",{})
    except Exception as e:
        pass

def modificacion_repuesto(req,id_rp):
    # try:
        repuesto = RepuestosPiezas.objects.get(id=id_rp)
        stock = int(repuesto.stock)
        stock_critico = int(repuesto.stock_critico)
        precio = int(repuesto.precio)
        if req.method == "POST":
            stock_txt = int(req.POST['stock_repuesto'])
            precio_txt = int(req.POST['precio_repuesto'])
            stock_critico_txt = int(req.POST['stock_critico_repuesto'])
            if stock_txt < 0:
                return render(req,"perfil_taller/repuestos/modificacion_repuesto.html",{"data":repuesto,
                                                                                        "stock":stock,
                                                                                        "stock_critico":stock_critico,
                                                                                        "precio":precio,
                                                                                        "error_message":"El stock ingresado es incorrecto"})
            elif precio_txt < 0:
                return render(req,"perfil_taller/repuestos/modificacion_repuesto.html",{"data":repuesto,
                                                                                        "stock":stock,
                                                                                        "stock_critico":stock_critico,
                                                                                        "precio":precio,
                                                                                        "error_message":"El precio ingresado es incorrecto"})
            elif stock_critico_txt < 0:
                return render(req,"perfil_taller/repuestos/modificacion_repuesto.html",{"data":repuesto,
                                                                                        "stock":stock,
                                                                                        "stock_critico":stock_critico,
                                                                                        "precio":precio,
                                                                                        "error_message":"El stock crítico ingresado es incorrecto"})
            else:
                foto = req.FILES.get('foto_repuesto')
                if foto:
                    repuesto.foto = foto
                repuesto.stock = stock_txt
                repuesto.stock_critico = stock_critico_txt
                repuesto.descripcion = req.POST['descripcion_repuesto']
                repuesto.precio = precio_txt
                repuesto.save()
                messages.success(req, "Repuesto modificado con éxito.")
                return redirect('Repuestos')
        else:
            
            return render(req,"perfil_taller/repuestos/modificacion_repuesto.html",{"data":repuesto,"stock":stock,"precio":precio,"stock_critico":stock_critico,})
    # except Exception as e:
    #     pass

def detalles_repuesto(req,id_rp):
    # try:
        repuesto = get_object_or_404(RepuestosPiezas, id=id_rp)

        # Obtener el año del parámetro GET, si no, usar el año actual
        anio = req.GET.get('year', datetime.now().year)
        anio = int(anio)  # Convertir a entero

        # Inicializar datos
        datos_repuestos = []
        total = 0

        # Obtener las cantidades por mes
        for mes in range(1, 13):
            cantidad_piezas = RepuestosPiezasServicios.objects.filter(
                servicio__fecha_ingreso__year=anio,
                servicio__fecha_ingreso__month=mes,
                repuestopieza_id=id_rp
            ).aggregate(total_cantidad=Sum('cantidad'))

            cantidad_ventas = ClienteRepuestosPiezas.objects.filter(
                 repuestospiezas_id=id_rp,
                 fecha_compra__year=anio,
                 fecha_compra__month=mes,
            ).aggregate(total_cantidad=Sum('cantidad'))

            cantidad = cantidad_piezas['total_cantidad'] or 0  # Si es None, usar 0
            cantidad_ventas = cantidad_ventas['total_cantidad'] or 0
            total += cantidad + cantidad_ventas
            datos_repuestos.append(cantidad)

        # Renderizar la plantilla
        return render(req, "perfil_taller/repuestos/detalles_repuesto.html", {
            "repuesto": repuesto,
            "cantidad": datos_repuestos,
            "total": total,
            "current_year": anio,
            "years": range(datetime.now().year - 10, datetime.now().year + 1),  # Últimos 10 años
        })
    # except Exception as e:
    #     pass

def stock_critico(req):
    try:
        repuestos = RepuestosPiezas.objects.filter(activo=1).order_by('-id')
        data = []
        for repuesto in repuestos:
            if int(repuesto.stock) <= int(repuesto.stock_critico):
                data.append({
                    "repuesto":repuesto
                })
        page_obj = funcion_paginas_varias(req,data)
        return render(req,"perfil_taller/stock_critico/stock_critico.html",{'page_obj': page_obj})
    except Exception as e:
        pass

def estadisticas_taller(req):
    try:
        anio = datetime.now().year
        mes_actual = datetime.now().month
        datos_servicios = []
        for n in range(1,mes_actual + 1):
            cant_servicios = Servicios.objects.filter(fecha_ingreso__year=anio,fecha_ingreso__month=n,estado="Completado").count()
            datos_servicios.append(int(cant_servicios))

        return render(req,"perfil_taller/estadisticas/estadisticas.html",{"datos":datos_servicios})
                                                                                  
    except Exception as e:
        pass

def clientes_taller(req):
    clientes = Cliente.objects.filter(
        cliente_telefono__principal=True
    ).values('id','nombre', 'apellido', 'cliente_telefono__telefono').order_by('nombre')
    page_obj = funcion_paginas_varias(req,clientes)
    return render(req,"perfil_taller/clientes/clientes.html",{'page_obj': page_obj,"clientes":clientes})

def buscar_por_doc_taller(req):
    tipo_doc = req.GET.get('tipo_doc_busq')
    doc_num = req.GET.get('documento')
    documento = tipo_doc + str(doc_num)

    cliente = Cliente.objects.filter(
         documento = documento,
         cliente_telefono__principal=True
     ).values('id','nombre', 'apellido', 'cliente_telefono__telefono')
    page_obj = funcion_paginas_varias(req,cliente)  # Obtiene la página solicitada
    return render(req,"perfil_taller/clientes/clientes.html",{'page_obj': page_obj,"clientes":cliente})

def buscar_nom_ape_taller(req):
    nombre = req.GET.get('nombre').capitalize()
    apellido = req.GET.get('apellido').capitalize()
    cliente = Cliente.objects.filter(
         nombre__icontains = nombre,
         apellido__icontains = apellido,
         cliente_telefono__principal=True
     ).values('id','nombre', 'apellido', 'cliente_telefono__telefono')
    page_obj = funcion_paginas_varias(req,cliente)  # Obtiene la página solicitada
    return render(req,"perfil_taller/clientes/clientes.html",{'page_obj': page_obj,"clientes":cliente})

def alta_cliente_taller(req):
    # try:
        if req.method == "POST":
            tipo_doc = req.POST['tipo_doc']
            doc = req.POST['doc']
            nombre = req.POST['nombre'].title()
            apellido = req.POST['apellido'].title()
            f_nac_str = req.POST.get('f_nac')  # Cambiado a paréntesis
            f_nac = datetime.strptime(f_nac_str, '%Y-%m-%d').date() if f_nac_str else None
            telefono_principal = req.POST['telefono_principal']
            telefono_secundario = req.POST['telefono_secundario']
            correo = req.POST['correo_1']
            dominio_correo = req.POST['dominio_correo']
            correo_2 = req.POST['correo_2']
            dominio_correo_2 = req.POST['dominio_correo_2']
            # localidad = req.POST['localidad'].title()
            # calle = req.POST['calle'].title()
            # numero = req.POST['numero']
            # num_apto = req.POST['num_apto']
            
            if correo:
                correo_principal = correo + dominio_correo
            else:
                correo_principal = None
            
            if correo_2:
                correo_secundario = correo_2 + dominio_correo_2
            else:
                correo_secundario = None
            
            doc_compuesto = tipo_doc + str(doc)
            existe_cliente = valid_cliente(doc_compuesto)
            if existe_cliente == "existe_cliente":
                return render(req,"perfil_taller/clientes/alta_cliente.html",{"error_message":"El cliente ya existe"})
            # elif existe_cliente == "existe_telefono_1":
            #     return render(req,"perfil_taller/clientes/alta_cliente.html",{"error_message":"El telefono 1 ya existe"})
            # elif existe_cliente == "existe_telefono_2":
            #     return render(req,"perfil_taller/clientes/alta_cliente.html",{"error_message":"El telefono 2 ya existe"})
            # elif existe_cliente == "existe_correo_1":
            #     return render(req,"perfil_taller/clientes/alta_cliente.html",{"error_message":"El correo 1 ya existe"})
            # elif existe_cliente == "existe_correo_2":
            #     return render(req,"perfil_taller/clientes/alta_cliente.html",{"error_message":"El correo 2 ya existe"})
            elif telefono_principal == telefono_secundario:
                return render(req,"perfil_taller/clientes/alta_cliente.html",{"error_message":"Los numeros de teléfono no pueden ser iguales"})
            elif (correo_principal and correo_secundario) and (correo_principal == correo_secundario):
                return render(req,"perfil_taller/clientes/alta_cliente.html",{"error_message":"Los correos no pueden ser iguales"})
            else:
                # if localidad == "Otro":
                #     ciudad = req.POST['localidad_otro'].title()
                # else:
                #     ciudad = localidad

                # if num_apto:
                #     n_a = num_apto
                # else:
                #     n_a = 0

                nuevo_cliente = Cliente(
                    documento = doc_compuesto,
                    nombre = nombre,
                    apellido = apellido,
                    fecha_nacimiento = f_nac,
                    domicilio = req.POST['domicilio'].title()
                    # ciudad = ciudad,
                    # calle = calle,
                    # numero = numero,
                    # num_apartamento = n_a
                )

                nuevo_cliente.save()
                insert_cliente_telefono(telefono_principal,1,nuevo_cliente.id)

                if telefono_secundario:
                    insert_cliente_telefono(telefono_secundario,0,nuevo_cliente.id)

                if correo_principal:
                    insert_cliente_correo(correo_principal,1,nuevo_cliente.id)
                
                if correo_secundario:
                    insert_cliente_correo(correo_secundario,0,nuevo_cliente.id)
                messages.success(req, "El cliente fue ingresado con éxito.")
                return redirect('ClientesTaller')
        else:
            return render(req,"perfil_taller/clientes/alta_cliente.html",{})
    # except Exception as e:
    #     pass  

def modificacion_cliente_taller(req,id_cliente):
    try:
        if req.method == "POST":
            tipo_doc = req.POST['tipo_doc']
            doc = req.POST['doc']
            documento = tipo_doc + str(doc)
            tel1 = req.POST['telefono_principal']
            tel2 = req.POST['telefono_secundario']
            correo1 = req.POST['correo_1'] + req.POST['dominio_correo']
            correo2 = req.POST['correo_2'] + req.POST['dominio_correo_2']
    
            valid_cliente = valid_cliente_mod(id_cliente,documento,tel1,tel2,correo1,correo2)
            if valid_cliente == "existe_cliente":
                contexto = contexto_para_cliente(id_cliente,"El documento ingresado ya existe")
                return render(req,"perfil_taller/clientes/modificacion_cliente.html",contexto)
            
            # elif valid_cliente == "existe_tel_principal":
            #     contexto = contexto_para_cliente(id_cliente,"El telefono 1 ingresado ya existe")
            #     return render(req,"perfil_taller/clientes/modificacion_cliente.html",contexto)
            # elif valid_cliente == "existe_tel_secundario":
            #     contexto = contexto_para_cliente(id_cliente,"El telefono 2 ingresado ya existe")
            #     return render(req,"perfil_taller/clientes/modificacion_cliente.html",contexto)
            # elif valid_cliente == "existe_correo_principal":
            #     contexto = contexto_para_cliente(id_cliente,"El correo 1 ingresado ya existe")
            #     return render(req,"perfil_taller/clientes/modificacion_cliente.html",contexto)
            # elif valid_cliente == "existe_correo_secundario":
            #     contexto = contexto_para_cliente(id_cliente,"El correo 2 ingresado ya existe")
            #     return render(req,"perfil_taller/clientes/modificacion_cliente.html",contexto)
            elif tel1 == tel2:
                contexto = contexto_para_cliente(id_cliente,"Los telefonos no pueden ser iguales")
                return render(req,"perfil_taller/clientes/modificacion_cliente.html",contexto)
            elif (correo1 == correo2) and (req.POST['correo_1'] and req.POST['correo_2']):
                contexto = contexto_para_cliente(id_cliente,"Los correos no pueden ser iguales")
                return render(req,"perfil_taller/clientes/modificacion_cliente.html",contexto)
            else:
            
                tel1_actual = ClienteTelefono.objects.filter(cliente_id=id_cliente,principal=1).first()
                if tel1_actual.telefono != tel1:
                    #SI EL TEL1 INGRESADO ES DISTINTO DEL ACTUAL --->>> BORRAR ACTUAL E INGRESAR NUEVO TEL1
                    tel1_actual.delete()
                    insert_cliente_telefono(tel1,1,id_cliente)

                if tel2:
                    tel2_actual = ClienteTelefono.objects.filter(cliente_id=id_cliente,principal=0).first()
                    checkbox = 'convert_to_tel1' in req.POST    
                    if tel2_actual.telefono != tel2:
                        #SI EL TEL2 INGRESADO ES DISTINTO DEL ACTUAL --->>> BORRAR ACTUAL E INGRESAR NUEVO TEL2
                        tel2_actual.delete()
                        insert_cliente_telefono(tel2,0,id_cliente)
                    if checkbox:
                        tel2_actual = ClienteTelefono.objects.filter(cliente_id=id_cliente,principal=0).first()
                        tel2_actual.principal = 1
                        tel1_actual.principal = 0
                        tel2_actual.save()
                        tel1_actual.save()

                if req.POST['correo_1']:
                    correo1_actual = ClienteCorreo.objects.filter(cliente_id=id_cliente,principal=1).first()
                    if correo1_actual.correo != correo1:
                        #SI EL CORREO1 INGRESADO ES DISTINTO DEL ACTUAL --->>> BORRAR ACTUAL E INGRESAR NUEVO CORREO1
                        correo1_actual.delete()
                        insert_cliente_correo(correo1,1,id_cliente)
                
                if req.POST['correo_2']:
                    correo2_actual = ClienteCorreo.objects.filter(cliente_id=id_cliente,principal=0).first()
                    checkbox_correo = 'convert_to_correo1' in req.POST
                    if correo2_actual.correo != correo2:
                        #SI EL CORREO2 INGRESADO ES DISTINTO DEL ACTUAL --->>> BORRAR ACTUAL E INGRESAR NUEVO CORREO2
                        correo2_actual.delete()
                        insert_cliente_correo(correo2,0,id_cliente)
                    
                    if checkbox_correo:
                        correo2_actual = ClienteCorreo.objects.filter(cliente_id=id_cliente,principal=0).first()
                        correo2_actual.principal = 1
                        correo1_actual.principal = 0 
                        correo1_actual.save()
                        correo2_actual.save()
                
                f_nac_str = req.POST.get('f_nac')  # Cambiado a paréntesis
                f_nac = datetime.strptime(f_nac_str, '%Y-%m-%d').date() if f_nac_str else None

                mod_cliente = Cliente.objects.get(id=id_cliente)
                mod_cliente.documento = documento
                mod_cliente.nombre = req.POST['nombre'].title()
                mod_cliente.apellido = req.POST['apellido'].title()
                mod_cliente.fecha_nacimiento = f_nac
                # mod_cliente.ciudad = req.POST['localidad'].title()
                # mod_cliente.calle = req.POST['calle'].title()
                # mod_cliente.numero = req.POST['numero']
                # mod_cliente.num_apartamento = req.POST['num_apto']
                mod_cliente.domicilio = req.POST['domicilio'].title()
                
                mod_cliente.save()
                messages.success(req, "Cliente modificado con éxito")
                return redirect('ClientesTaller')
        else:
            contexto = contexto_para_cliente(id_cliente,None)
            return render(req,"perfil_taller/clientes/modificacion_cliente.html",contexto)
    except Exception as e:
        return render(req,"perfil_taller/clientes/modificacion_cliente.html",{"error_message":e})


def detalles_cliente_taller(req,id_cliente):
    cliente = Cliente.objects.get(id=id_cliente)
    tel1 = ClienteTelefono.objects.filter(principal=1,cliente_id=id_cliente).first()
    tel_1 = tel1.telefono
    tel2 = ClienteTelefono.objects.filter(principal=0,cliente_id=id_cliente).first()

    correo1 = ClienteCorreo.objects.filter(principal=1,cliente_id=id_cliente).first()
    correo2 = ClienteCorreo.objects.filter(principal=0,cliente_id=id_cliente).first()
    if tel2:
        tel_2 = tel2.telefono
    else:
        tel_2 = None

    if correo1:
        c_1 = correo1.correo
    else:
        c_1 = None
    
    if correo2:
        c_2 = correo2.correo
    else:
        c_2 = None
    
    # resultados_motos = (
    #     Servicios.objects
    #     .filter(cliente__id=id_cliente)
    #     .select_related('moto', 'cliente')
    #     .values(
    #         'id',
    #         'moto__marca', 
    #         'moto__modelo',
    #     ).order_by('-id')
    # )
    resultados_motos = (
    Servicios.objects
    .filter(cliente__id=id_cliente)
    .select_related('moto', 'cliente')
    .values('moto__marca', 'moto__modelo','moto__id')  # Agrupar por estos campos
    # .annotate(total_servicios=Count('id'))  # Contar servicios por cada grupo
    # .order_by('-total_servicios')  # Ordenar por la cantidad de servicios
    .annotate(ultima_fecha_ingreso=Max('fecha_ingreso'))
    .order_by('-ultima_fecha_ingreso')
    )

    res_data = []
    for resultado in resultados_motos:
            matricula = Matriculas.objects.filter(moto_id=resultado['moto__id']).first()
            res_data.append({
            'moto': resultado,
            "matricula":matricula.matricula if matricula else "Sin matrícula"
        })
    
    compras_rp = (
        ClienteRepuestosPiezas.objects
        .filter(cliente_id=id_cliente)
        .select_related('repuestospiezas', 'cliente')
        .values('id','cantidad','fecha_compra','repuestospiezas__descripcion')
        .order_by('-id')
    )





    page_obj = funcion_paginas_varias(req,res_data)
    page_obj_rp = funcion_paginas_varias(req,compras_rp)

    return render(req,"perfil_taller/clientes/detalles_cliente.html",{
                                                                    "page_obj":page_obj,
                                                                    "cliente":cliente,
                                                                    "tel1":tel_1,
                                                                    "tel2":tel_2,
                                                                    "correo1":c_1,
                                                                    "correo2":c_2, 
                                                                    "page_obj_rp":page_obj_rp,
                                                                    })

def servicios_por_moto_de_cliente(req,id_moto,id_cliente):
    try:
        moto = Moto.objects.get(id=id_moto)
        matricula = Matriculas.objects.filter(moto_id=id_moto).first()
        matr_actual = matricula.matricula if matricula else None
        servicios_moto = (
        Servicios.objects
        .all()
        .select_related('moto')
        .filter(estado="Completado",moto_id=id_moto)
        .values(
            'id',
            'fecha_ingreso',  
            'titulo'     
           ).order_by('-fecha_ingreso')
        )
        page_obj = funcion_paginas_varias(req,servicios_moto)
        return render(req,"perfil_taller/clientes/servicios_por_moto.html",{"page_obj":page_obj,"id_cliente":id_cliente,"moto":moto,"matr_actual":matr_actual})                                                          
                                                                   
    except Exception as e:
        pass

def detalle_de_cada_servicio_de_moto(req,id_s,id_cliente):
    contexto = contexto_modificar_servicio(id_s,None)
    datos = json_para_servicio(id_s)
    return render(req,"perfil_taller/clientes/detalles_servicio.html",{ "moto":contexto[0],
                                                                            "cliente":contexto[1],
                                                                            "matricula":contexto[2],
                                                                            "telefono":contexto[3],
                                                                            "correo":contexto[4],
                                                                            "tareas_realizadas":contexto[5],
                                                                            "tareas_pendientes":contexto[6],
                                                                            "mecanicos":contexto[7],
                                                                            "id_servicio":contexto[9],
                                                                            "anotaciones":contexto[10],
                                                                            "info_servicio":contexto[12],
                                                                            "fecha_cierre":contexto[13],
                                                                            "id_cliente":id_cliente,
                                                                            "datos_fijos":datos[0],
                                                                            "tareas":datos[1],
                                                                            "observaciones":datos[2]
                                                                            })

def pedidos(req):
    # try:
        pedidos = Pedidos.objects.all().order_by('fecha')
        data = []
        for pedido in pedidos:
            cliente = Cliente.objects.get(id=pedido.cliente_id)
            cliente_telefono = ClienteTelefono.objects.filter(cliente_id=pedido.cliente_id,principal=1).first()

            data.append({
                "id_pedido":pedido.id,
                "pedido":pedido.detalle,
                "fecha":pedido.fecha,
                "cliente":cliente.nombre + " " + cliente.apellido,
                "telefono":cliente_telefono.telefono
            })
        page_obj = funcion_paginas_varias(req,data)
        return render(req,"perfil_administrativo/pedidos/pedidos.html",{'page_obj': page_obj})
    # except Exception as e:
    #     pass

def cliente_pedido(req):
    try:
        if req.method == "POST":
            documento = req.POST['tipo_doc'] + str(req.POST['doc'])
            cliente = Cliente.objects.filter(documento=documento).first()
            if not cliente:
                return render(req,"perfil_administrativo/pedidos/alta_pedido.html",{"ingresar_pedido":False,"error_message_cliente":"El cliente no se encuentra registrado en el sistema, para ingresarlo haga clic "})
            else:
                correo = ClienteCorreo.objects.filter(cliente_id=cliente.id,principal=1).first()
                if correo:
                    correo = correo.correo
                else:
                    correo = "El cliente no cuenta con correo"
                telefono = ClienteTelefono.objects.filter(cliente_id=cliente.id,principal=1).first()
                if telefono:
                    telefono = telefono.telefono
                else:
                    telefono = "El cliente no cuenta con teléfono"
                return render(req,"perfil_administrativo/pedidos/alta_pedido.html",{"ingresar_pedido":True,"cliente":cliente,"correo":correo,"telefono":telefono})
        else:
            return render(req,"perfil_administrativo/pedidos/alta_pedido.html",{"ingresar_pedido":False})
    except Exception as e:
        pass

def alta_pedido(req,id_cliente):
    try:
        nuevo_pedido = Pedidos(
            detalle = req.POST['pedido'],
            fecha = datetime.now(),
            cliente_id = id_cliente
        )
        nuevo_pedido.save()
        messages.success(req, "Pedido ingresado con éxito")
        return redirect('Pedidos')
    
    except Exception as e:
        pass

def baja_pedido(req,id_pedido):
    try:
        if req.method == "POST":
            pedido = Pedidos.objects.get(id=id_pedido)
            pedido.delete()
            messages.success(req, "Pedido borrado con éxito")
            return redirect('Pedidos')
        else:
            return render(req,"perfil_administrativo/pedidos/baja_pedido.html",{})
    except Exception as e:
        pass

def cerrar_pedido(req,id_pedido):
    try:
        pedido = Pedidos.objects.get(id=id_pedido)
        correo = ClienteCorreo.objects.filter(cliente_id=pedido.cliente_id,principal=1).first()
        if req.method == "POST":
            checkbox = 'notificar_cliente' in req.POST
            if correo and checkbox:
                cliente = Cliente.objects.get(id=pedido.cliente_id)
                titulo = "Encargue Umpierrez Motos"
                mensaje = f"Estimado/a {cliente.nombre} {cliente.apellido}, le informamos que su {pedido.detalle} ha llegado a nuestro local. Puede pasar por el mismo cuando desee."
                enviar_correo(titulo,mensaje,correo.correo)
            pedido.delete()
            messages.success(req, "Pedido cerrado con éxito")
            return redirect('Pedidos')
        else:
            if correo:
                mostrar_cbox = True
            else:
                mostrar_cbox = False
            return render(req,"perfil_administrativo/pedidos/cierre_pedido.html",{"mostrar_cbox":mostrar_cbox})
    except Exception as e:
        pass

def busqueda_pedido_por_doc_cliente(req):
    # try:
        documento = req.GET.get('tipo_documento') + str(req.GET.get('documento'))
        cliente_busq = Cliente.objects.filter(documento=documento).first()
        pedidos = Pedidos.objects.filter(cliente_id=cliente_busq.id).order_by('fecha')
        data = []
        for pedido in pedidos:
            cliente = Cliente.objects.get(id=pedido.cliente_id)
            cliente_telefono = ClienteTelefono.objects.filter(cliente_id=pedido.cliente_id,principal=1).first()

            data.append({
                "id_pedido":pedido.id,
                "pedido":pedido.detalle,
                "fecha":pedido.fecha,
                "cliente":cliente.nombre + " " + cliente.apellido,
                "telefono":cliente_telefono.telefono
            })
        page_obj = funcion_paginas_varias(req,data)
        return render(req,"perfil_administrativo/pedidos/pedidos.html",{'page_obj': page_obj})
    # except Exception as e:
    #     pass

def busqueda_pedido_por_nombre_apellido_cliente(req):
    # try:
        nombre = req.GET.get('nombre')
        apellido = req.GET.get('apellido')
        cliente_busq = Cliente.objects.filter(nombre__contains=nombre,apellido__contains=apellido).first()
        pedidos = Pedidos.objects.filter(cliente_id=cliente_busq.id).order_by('fecha')
        data = []
        for pedido in pedidos:
            cliente = Cliente.objects.get(id=pedido.cliente_id)
            cliente_telefono = ClienteTelefono.objects.filter(cliente_id=pedido.cliente_id,principal=1).first()

            data.append({
                "id_pedido":pedido.id,
                "pedido":pedido.detalle,
                "fecha":pedido.fecha,
                "cliente":cliente.nombre + " " + cliente.apellido,
                "telefono":cliente_telefono.telefono
            })
        page_obj = funcion_paginas_varias(req,data)
        return render(req,"perfil_administrativo/pedidos/pedidos.html",{'page_obj': page_obj})
    # except Exception as e:
    #     pass

def personal_taller(req):

    usuario = req.user
    usuario_actual = Personal.objects.filter(usuario=usuario.username).first()
    mecanico_actual = Mecanico.objects.filter(id=usuario_actual.id).first()
    mostrar_botones = True if mecanico_actual.jefe else False
    administrativos = (Administrativo.objects
                       .filter(activo=True).exclude(usuario="adminapp")
                       .values('id', 'nombre', 'apellido', 'telefono', 'correo', 'activo','usuario')
                       .order_by('nombre'))
    page_obj = funcion_paginas_varias(req,administrativos)

    mecanicos = (Mecanico.objects
                       .filter(activo=True).exclude(usuario="adminapp")
                       .values('id', 'nombre', 'apellido', 'telefono', 'correo', 'activo')
                       .order_by('nombre'))
    page_objMec = funcion_paginas_varias(req,mecanicos)

    return render(req,"perfil_taller/personal/personal.html",{"page_obj":page_obj,"page_objMec":page_objMec, "mostrar_botones":mostrar_botones})

def alta_personal_taller(req):
    try:
        if req.method == "POST":
            documento = req.POST['tipo_doc'] + str(req.POST['doc'])
            telefono = req.POST['telefono'] 
            correo_nombre = req.POST['correo'] 
            correo_dominio = req.POST['dominio_correo'] 
            correo = correo_nombre + correo_dominio
            correo = correo.lower()

            valid_personal = validar_personal(documento,telefono,correo)

            if valid_personal == "existe_mecanico":
                return render(req,"perfil_taller/personal/alta_personal.html",{"error_message":"La persona que desea ingresar ya existe en el sistema"})
            elif valid_personal == "existe_telefono":
                return render(req,"perfil_taller/personal/alta_personal.html",{"error_message":"El teléfono ingresado ya existe en el sistema, el mismo no debe estar repetido"})
            elif valid_personal == "existe_correo":
                return render(req,"perfil_taller/personal/alta_personal.html",{"error_message":"El correo ingresado ya existe en el sistema, el mismo no debe estar repetido"})
            elif valid_personal == "mecanico_desactivado":    
                persona = Personal.objects.filter(documento=documento).first()
                id = persona.id
                return render(req,"perfil_taller/personal/alta_personal.html",{"reingresar":True,"id_personal":id})
            else:
                    f_nac_str = req.POST.get('f_nac')  # Cambiado a paréntesis
                    f_nac = datetime.strptime(f_nac_str, '%Y-%m-%d').date() if f_nac_str else None
                    usuario = nombre_usuario(req.POST['nombre'],req.POST['apellido'])
                    
                    personal = Personal.objects.create(
                        documento=documento,
                        nombre=req.POST['nombre'].capitalize(),
                        apellido=req.POST['apellido'].capitalize(),
                        fecha_nacimiento=f_nac,
                        usuario=usuario,
                        contrasena="Inicio1234",
                        correo=correo,
                        telefono=telefono,
                        primera_sesion = True
                    )
                    personal.save()
                    # Crear registro en Administrativo asociado al Personal
                    administrativo = Administrativo.objects.create(
                        personal_ptr=personal,  # Asociar al registro de Personal
                        activo=False,
                        documento=documento,
                        nombre=req.POST['nombre'].capitalize(),
                        apellido=req.POST['apellido'].capitalize(),
                        fecha_nacimiento=f_nac,
                        usuario=usuario,
                        contrasena="Inicio1234",
                        correo=correo,
                        telefono=telefono,
                        primera_sesion = True
                    )
                    administrativo.save()

                    # Crear registro en Mecanico asociado al Personal

                    jefe = True if req.POST['permiso_del_mecanico'] == "jefe" else False
                    mecanico = Mecanico.objects.create(
                        personal_ptr=personal,  # Asociar al registro de Personal
                        activo=True,
                        jefe=jefe,
                        documento=documento,
                        nombre=req.POST['nombre'].capitalize(),
                        apellido=req.POST['apellido'].capitalize(),
                        fecha_nacimiento=f_nac,
                        usuario=usuario,
                        contrasena="Inicio1234",
                        correo=correo,
                        telefono=telefono,
                        primera_sesion = True
                    )
                    mecanico.save()

                    user = User.objects.create_user(
                        username=usuario,
                        password="Inicio1234",  # La contraseña predeterminada
                        email=correo,
                        first_name=req.POST['nombre'].capitalize(),
                        last_name=req.POST['apellido'].capitalize()
                    )
                    
                    # Asignar permisos o grupos si es necesario
                    user.is_staff = False  # Si deseas que tenga acceso al admin de Django
                    user.save()

                    messages.success(req, "Personal ingresado con éxito")
                    return redirect('PersonalTaller')
        else:
            return render(req,"perfil_taller/personal/alta_personal.html",{})
    except Exception as e:
        pass

def ingresar_taller(req,id_personal):
    try:
        jefe = True if req.POST['permiso_del_mecanico_reingreso'] == "jefe" else False
        mecanico = Mecanico.objects.get(personal_ptr_id=id_personal)
        mecanico.activo = 1
        mecanico.jefe = jefe
        mecanico.save()
        messages.success(req, "Personal reingresado con éxito")
        return redirect('PersonalTaller')
    except Exception as e:
        return render(req,"perfil_taller/personal/alta_personal.html",{"error_message":e})

def detalles_personal_taller(req,id_personal):
    try:
        personal = Personal.objects.get(id=id_personal)
        administrativo = Administrativo.objects.filter(personal_ptr_id = id_personal,activo = 1).first()
        if administrativo:
            admin = True
        else:
            admin = False
        
        mecanico = Mecanico.objects.filter(personal_ptr_id = id_personal,activo = 1).first()
        if mecanico:
            mec = True
            if mecanico.jefe == 1:
                jefe = True
            else:
                jefe = False
        else:
            mec = False
            jefe = False
        contexto = {
            "personal":personal,
            "admin":admin,
            "mec":mec,
            "jefe":jefe
        }
        return render(req,"perfil_taller/personal/detalles_personal.html",contexto)
    except Exception as e:
         pass

def resetear_usuario_taller(req,id_u):
    # try:
        if req.method == "POST":
            usuario = Personal.objects.get(id=id_u)
            usuario.contrasena = make_password("Inicio1234")
            usuario.primera_sesion = 1
            usuario.save()
            user = User.objects.filter(username=usuario.usuario).first()
            user.password = make_password("Inicio1234")
            user.save()
            return render(req,"perfil_taller/personal/resetear_usuario.html",{"message":"Usuario reseteado con éxito"})
        else:
            return render(req,"perfil_taller/personal/resetear_usuario.html",{})

def notificaciones_pagos_atrasados():
    #EJECUTAR ESTA FUNCION TODOS LOS DIAS A LAS 6AM
    pass

def notificaciones_taller(req):
    # try:
        #notificaciones_cumples()
        # notificaciones = [
        # {
        # "titulo": "Pago pendiente",
        # "fecha": "2024-11-28",
        # "descripcion": "El cliente Juan Pérez tiene un pago atrasado desde hace 7 días.",
        # "acciones": [
        #     {"nombre": "Ver detalle", "url": "/detalle_pago/123/"},
        #     {"nombre": "Contactar cliente", "url": "/contacto_cliente/123/"}
        # ]
        # },
        # {
        # "titulo": "Cumpleaños del cliente",
        # "fecha": "2024-11-29",
        # "descripcion": "Hoy es el cumpleaños de María López. Envíale un saludo especial.",
        # "acciones": [{"nombre": "Enviar saludo", "url": "/enviar_saludo/456/"}]
        #     }
        #     ]
        usuario = req.user
        usuario_actual = Personal.objects.filter(usuario=usuario.username).first()
        filter_option = req.GET.get('filter', 'all')
        if filter_option == 'leidas':
            # notificaciones = Notificaciones.objects.filter(leido=True).order_by('-id')  # Filtrar solo las leídas
            notificaciones = (
                NotificacionPersonal.objects
                .filter(leido=True,personal=usuario_actual)
                .values(
                    'id',
                    'notificacion__tipo', 
                    'notificacion__fecha', 
                    'notificacion__descripcion', 
                     'notificacion__id',
                )
            ).order_by('-id')
        else:
            # notificaciones = Notificaciones.objects.all().order_by('-id')  # Mostrar todas las notificaciones
            notificaciones = (
                NotificacionPersonal.objects
                .filter(leido=False,personal=usuario_actual)
                .values(
                    'id',
                    'notificacion__tipo', 
                    'notificacion__fecha', 
                    'notificacion__descripcion', 
                    'notificacion__id',
                     
                )
            ).order_by('-id')
        
        #notificaciones = Notificaciones.objects.all().order_by('-id')
        data = []
        for notificacion in notificaciones:
            if notificacion['notificacion__tipo'] == "Atraso en pago":
                acciones = {"nombre": "Enviar correo", "url": ""}
            elif notificacion['notificacion__tipo'] == "Cumpleaños":
                acciones = [{"nombre": "Ver detalle", "url": ""},]
            else:
                acciones = [{"nombre": "Ver detalle", "url": ""},]
            data.append({
                "notificacion":notificacion,
                "acciones":acciones
            })
        
        
        notif_no_leidas = NotificacionPersonal.objects.filter(leido=0,personal=usuario_actual)
        for notificacion in notif_no_leidas:
            notificacion.leido = 1
            notificacion.save()
        
        return render(req,"perfil_taller/notificaciones/notificaciones.html",{"data":data}) 
    # except Exception as e:
    #     return render(req,"perfil_taller/notificaciones/notificaciones.html",{"error_message":e})

def borrar_notificacion_taller(req,id_notificacion):
    # try:
        usuario = req.user
        usuario_actual = Personal.objects.filter(usuario=usuario.username).first()
        
        notificacion_usuario = NotificacionPersonal.objects.filter(notificacion_id=id_notificacion,personal=usuario_actual).first()
        notificacion_usuario.delete()
        return redirect('NotificacionesTaller')

    # except Exception as e:
    #     pass


def venta_repuesto_form(req,id_rp):
    # try:
        if req.method == "POST":
            tipo_doc = req.POST['tipo_documento']
            doc = req.POST['documento']
            documento = tipo_doc + str(doc)
            contexto = contexto_venta_repuesto(req,id_rp,None,documento)
            return render(req,"perfil_taller/repuestos/venta_repuesto.html",contexto)
        else:
            return render(req,"perfil_taller/repuestos/venta_repuesto.html",{})
    # except Exception as e:
    #     pass

def venta_repuesto(req,id_rp,id_cliente):
    try:
        rp = RepuestosPiezas.objects.get(id=id_rp)
        cliente = Cliente.objects.get(id=id_cliente)
        stock = int(rp.stock)
        cantidad = int(req.POST['cantidad_rp'])
        if cantidad > stock:
            contexto = contexto_venta_repuesto(req,id_rp,"La cantidad ingresada no puede exceder el stock del producto",cliente.documento)
            return render(req,"perfil_taller/repuestos/venta_repuesto.html",contexto)
        elif cantidad <= 0:
            contexto = contexto_venta_repuesto(req,id_rp,"La cantidad ingresada es incorrecta",cliente.documento)
            return render(req,"perfil_taller/repuestos/venta_repuesto.html",contexto)
        else:
            nueva_venta_rp = ClienteRepuestosPiezas(
                cliente_id = id_cliente,
                repuestospiezas_id = id_rp,
                cantidad = req.POST['cantidad_rp'],
                fecha_compra = datetime.now()
            )
            nueva_venta_rp.save()
            rp.stock = stock - cantidad
            rp.save()
            if int(rp.stock) <= int(rp.stock_critico):
                insert_notificaciones(f"Hay poco stock de {rp.descripcion}","Bajo stock de pieza")
            messages.success(req, "Venta generada con éxito")
            return redirect(f"{reverse('DetallesClienteTaller',kwargs={'id_cliente':id_cliente})}")
    except Exception as e:
        pass
    
def calculadora_pagos(req):
    return render(req, "perfil_administrativo/calculadora/calculadora.html", {})

def vista_inventario_motos_taller(req):
    motos = Moto.objects.filter(pertenece_taller=1).order_by('-fecha_ingreso')
    data_motos = []
    for moto in motos:
        matricula = Matriculas.objects.filter(moto_id=moto.id).first()
        servicio = Servicios.objects.filter(moto_id=moto.id).first()
        if servicio:
            cliente = Cliente.objects.get(id=servicio.cliente_id)
            datos_cliente = cliente.nombre + " " + cliente.apellido
        else:
            cliente = None
            datos_cliente = None
        data_motos.append({
            "moto":moto,
            "matricula":matricula.matricula if matricula else "Sin matrícula",
            "cliente":datos_cliente if cliente else "Sin cliente asignado"
        })
    logo_um = Logos.objects.get(id=1)
    page_obj = funcion_paginas_varias(req,data_motos)
    return render(req,"perfil_taller/motos/motos.html",{'page_obj': page_obj,"motos":motos,"logo_um":logo_um.logo_UM.url if logo_um.logo_UM else None,"active_page": 'Motos'})

def alta_moto_taller(req):
    # try:
        if req.method == "POST":
            checkbox_num_motor = 'sin_num_motor' in req.POST
            if checkbox_num_motor:
                num_motor = crear_num_motor()
                contiene_num_motor = 0
            else:
                num_motor = req.POST['num_motor_moto'].upper()
                contiene_num_motor = 1

            checkbox_num_chasis = 'sin_num_chasis' in req.POST
            if checkbox_num_chasis:
                num_chasis = crear_num_chasis()
                contiene_num_chasis = 0
            else:
                num_chasis = req.POST['num_chasis_moto'].upper()
                contiene_num_chasis = 1

            if req.POST['matricula_letras'] and req.POST['matricula_numeros']:
                matricula = req.POST['matricula_letras'].upper() + str(req.POST['matricula_numeros'])
                matr_moto = Matriculas.objects.filter(matricula=matricula).first()
                if matr_moto:
                    existe_matr = True
                else:
                    existe_matr = False
            else:
                matricula = None
                existe_matr = False

            existe_num_motor = Moto.objects.filter(num_motor=num_motor).first()
            existe_num_chasis = Moto.objects.filter(num_chasis=num_chasis).first()
            modificar = False
            if existe_num_motor:
                if existe_num_motor.pertenece_taller == 1:
                    return render(req,"perfil_taller/motos/alta_moto.html",{
                                                                            "active_page": 'Motos',
                                                                            "error_message":"Ya existe el número de motor ingresado",
                                                                            })
                else:
                    modificar = True
            elif existe_num_chasis:
                if existe_num_chasis.pertenece_taller == 1:
                    return render(req,"perfil_taller/motos/alta_moto.html",{
                                                                            "active_page": 'Motos',
                                                                            "error_message":"Ya existe el número de chasis ingresado",
                                                                            })
                else:
                    modificar = True
            elif existe_matr:
                return render(req,"perfil_taller/motos/alta_moto.html",{
                                                                            "active_page": 'Motos',
                                                                            "error_message":"Ya existe la matrícula ingresada",
                                                                            })
            else:
                if modificar:
                    moto = Moto.objects.filter(num_motor=num_motor).first()
                    moto.pertenece_taller = 1
                    moto.save()
                else:
                    marca = req.POST['marca_moto'].upper()
                    modelo = req.POST['modelo_moto'].upper()
                    foto = req.FILES.get('foto_moto')
                    nueva_moto = insert_moto(marca,modelo,None,"Usada",req.POST['motor_moto'],None,None,None,None,num_motor,num_chasis,req.POST['num_cilindros'],None,0,1,req.POST['descripcion_moto'],foto,req.POST['tipo_moto'],contiene_num_motor,contiene_num_chasis)
                    if req.POST['matricula_letras'] and req.POST['matricula_numeros']:
                        matricula = req.POST['matricula_letras'].upper() + str(req.POST['matricula_numeros'])
                        nueva_matr = Matriculas(
                            matricula = matricula,
                            moto_id = nueva_moto.id
                        )
                        nueva_matr.save()
                messages.success(req, "Moto ingresada con éxito")
                return redirect('MotosTaller')
        else:
            return render(req,"perfil_taller/motos/alta_moto.html",{})
    # except Exception as e:
    #     pass


def balanceo(req,busqueda):
    # try:
        # print("ID CAJA ES: " + str(id_caja))
        # req.session["id_caja_movimiento"] = id_caja
        if busqueda == "metodo":
            movimientos = Movimientos.objects.filter(metodo=req.GET.get('metodo')).order_by('-fecha')
        elif busqueda == "fecha":
            fecha_str = req.GET.get('f_movimiento')  # Cambiado a paréntesis
            fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date() if fecha_str else None
            movimientos = Movimientos.objects.filter(fecha__date=fecha)
        elif busqueda == "tipo":
            movimientos = Movimientos.objects.filter(tipo=req.GET.get('tipo_movimiento')).order_by('-fecha')
        else:
            movimientos = Movimientos.objects.all().order_by('-fecha')
        # caja = Caja.objects.get(id=id_caja)
        # monto_inicial = caja.monto_inicial
        dolar = PrecioDolar.objects.get(id=1)
        precio_dolar = float(dolar.precio_dolar_tienda)
        # print(mes_actual)
        data = []
        for movimiento in movimientos:
            usuario = Personal.objects.get(id=movimiento.usuario_id)
            es_pago_moto = MovimientoPagoMoto.objects.filter(movimiento_id=movimiento.id).first()
            es_pago_accesorio = MovimientoPagoAccesorio.objects.filter(movimiento_id=movimiento.id).first()

            if es_pago_moto or es_pago_accesorio:
                mostrar_boton = False
            else: 
                mostrar_boton = True
            data.append({
                "movimiento":movimiento,
                "descripcion":movimiento.movimiento if movimiento.movimiento else "Sin descripción",
                "usuario":usuario.nombre + " " + usuario.apellido,
                "mostrar_boton":mostrar_boton
            })
        
        
        mes_actual = datetime.now().month
        movimientos_mes_actual = Movimientos.objects.filter(fecha__month=mes_actual)
        ganancias_mes = 0
        egresos_mes = 0 
        ganancias_netas_mes = 0
        ingresos_extras_mes = 0
        
        for movimiento in movimientos_mes_actual:
            if movimiento.tipo == "Ingreso":
                if movimiento.moneda == "Pesos":
                    ganancias_mes = ganancias_mes + float(movimiento.monto)
                else:
                    ganancias_mes = ganancias_mes + (float(movimiento.monto) * precio_dolar)
            elif movimiento.tipo == "Egreso":
                if movimiento.moneda == "Pesos":
                    egresos_mes = egresos_mes + float(movimiento.monto)
                else:
                    egresos_mes = egresos_mes + (float(movimiento.monto) * precio_dolar)
            elif movimiento.tipo == "Ingreso extra":
                if movimiento.moneda == "Pesos":
                    ingresos_extras_mes = ingresos_extras_mes + float(movimiento.monto)
                else:
                    ingresos_extras_mes = ingresos_extras_mes + (float(movimiento.monto) * precio_dolar)
            else:
                pass
            

        ganancias_netas_mes = ganancias_mes + ingresos_extras_mes - egresos_mes

        # print("GANANCIAS DEL MES: " + str(ganancias_mes))
        # print("EGRESOS DEL MES: -" + str(egresos_mes))
        # print("INGRESOS EXTRAS DEL MES: " + str(ingresos_extras_mes))
        # print("GANANCIAS NETAS DEL MES: " + str(ganancias_netas_mes))
        
        
        
        ganancias_dia = 0
        egresos_dia = 0
        ganancias_netas_dia = 0
        ingresos_extras_dia = 0

        fecha_actual = datetime.now().date()
        movimientos_fecha_actual = Movimientos.objects.filter(fecha__date=fecha_actual)

        for movimiento in movimientos_fecha_actual:
            if movimiento.tipo == "Ingreso":
                if movimiento.moneda == "Pesos":
                    ganancias_dia = ganancias_dia + float(movimiento.monto)
                else:
                    ganancias_dia = ganancias_dia + (float(movimiento.monto) * precio_dolar)
            elif movimiento.tipo == "Egreso":
                if movimiento.moneda == "Pesos":
                    egresos_dia = egresos_dia + float(movimiento.monto)
                else:
                    egresos_dia = egresos_dia + (float(movimiento.monto) * precio_dolar)
            elif movimiento.tipo == "Ingreso extra":
                if movimiento.moneda == "Pesos":
                    ingresos_extras_dia = ingresos_extras_dia + float(movimiento.monto)
                else:
                    ingresos_extras_dia = ingresos_extras_dia + (float(movimiento.monto) * precio_dolar)
            else:
                pass
            

        ganancias_netas_dia = ganancias_dia + ingresos_extras_dia - egresos_dia
        # print("GANANCIAS DEL DIA: " + str(ganancias_dia))
        # print("EGRESOS DEL DIA: -" + str(egresos_dia))
        # print("INGRESOS EXTRAS DEL DIA: " + str(ingresos_extras_dia))
        # print("GANANCIAS NETAS DEL DIA: " + str(ganancias_netas_dia))

        

        todos_movimientos_pesos = Movimientos.objects.all()
        ingresos_efectivo_pesos = 0
        ingresos_transferencia_pesos = 0
        ingresos_debito_pesos = 0
        egresos_efectivo_pesos = 0
        egresos_transferencia_pesos = 0
        egresos_debito_pesos = 0

        ingresos_efectivo_dolares = 0
        ingresos_transferencia_dolares = 0
        ingresos_debito_dolares = 0
        egresos_efectivo_dolares = 0
        egresos_transferencia_dolares = 0
        egresos_debito_dolares = 0

        for mov in todos_movimientos_pesos:
            if mov.moneda == "Pesos":
                if mov.tipo == "Egreso":
                    if mov.metodo == "Efectivo":
                        egresos_efectivo_pesos = egresos_efectivo_pesos + float(mov.monto)
                    elif mov.metodo == "Transferencia":
                        egresos_transferencia_pesos = egresos_transferencia_pesos + float(mov.monto) 
                    else:
                        egresos_debito_pesos = egresos_debito_pesos + float(mov.monto)
                else:
                    if mov.metodo == "Efectivo":
                        ingresos_efectivo_pesos = ingresos_efectivo_pesos + float(mov.monto)
                    elif mov.metodo == "Transferencia":
                        ingresos_transferencia_pesos = ingresos_transferencia_pesos + float(mov.monto)
                    else:
                        ingresos_debito_pesos = ingresos_debito_pesos + float(mov.monto)
            else:
                if mov.tipo == "Egreso":
                    if mov.metodo == "Efectivo":
                        egresos_efectivo_dolares = egresos_efectivo_dolares + float(mov.monto)
                    elif mov.metodo == "Transferencia":
                        egresos_transferencia_dolares = egresos_transferencia_dolares + float(mov.monto) 
                    else:
                        egresos_debito_dolares = egresos_debito_dolares + float(mov.monto)
                else:
                    if mov.metodo == "Efectivo":
                        ingresos_efectivo_dolares = ingresos_efectivo_dolares + float(mov.monto)
                    elif mov.metodo == "Transferencia":
                        ingresos_transferencia_dolares = ingresos_transferencia_dolares + float(mov.monto)
                    else:
                        ingresos_debito_dolares = ingresos_debito_dolares + float(mov.monto)


        total_ingresos_dolares = ingresos_efectivo_dolares + ingresos_transferencia_dolares + ingresos_debito_dolares
        total_egresos_dolares = egresos_efectivo_dolares + egresos_transferencia_dolares + egresos_debito_dolares
        
        total_ingresos_pesos = ingresos_efectivo_pesos + ingresos_transferencia_pesos + ingresos_debito_pesos
        total_egresos_pesos = egresos_efectivo_pesos + egresos_transferencia_pesos + egresos_debito_pesos

        total_ingresos_efectivo_pesos = round(ingresos_efectivo_pesos + float(ingresos_efectivo_dolares * precio_dolar),2)
        total_ingresos_efectivo_dolares = round(ingresos_efectivo_dolares + float(ingresos_efectivo_pesos / precio_dolar),2)

        total_ingresos_transferencia_pesos = round(ingresos_transferencia_pesos + float(ingresos_transferencia_dolares * precio_dolar),2)
        total_ingresos_transferencia_dolares = round(ingresos_transferencia_dolares + float(ingresos_transferencia_pesos / precio_dolar),2)

        total_ingresos_debito_pesos = round(ingresos_debito_pesos + float(ingresos_debito_dolares * precio_dolar),2)
        total_ingresos_debito_dolares = round(ingresos_debito_dolares + float(ingresos_debito_pesos / precio_dolar),2)


        total_egresos_efectivo_pesos = round(egresos_efectivo_pesos + float(egresos_efectivo_dolares * precio_dolar),2)
        total_egresos_efectivo_dolares = round(egresos_efectivo_dolares + float(egresos_efectivo_pesos / precio_dolar),2)

        total_egresos_transferencia_pesos = round(egresos_transferencia_pesos + float(egresos_transferencia_dolares * precio_dolar),2)
        total_egresos_transferencia_dolares = round(egresos_transferencia_dolares + float(egresos_transferencia_pesos / precio_dolar),2)

        total_egresos_debito_pesos = round(egresos_debito_pesos + float(egresos_debito_dolares * precio_dolar),2)
        total_egresos_debito_dolares = round(egresos_debito_dolares + float(egresos_debito_pesos / precio_dolar),2)

        ingresos_totales = total_ingresos_efectivo_pesos + total_ingresos_transferencia_pesos + total_ingresos_debito_pesos
        egresos_totales = total_egresos_efectivo_pesos + total_egresos_transferencia_pesos + total_egresos_debito_pesos
        saldo_total = ingresos_totales - egresos_totales

        saldo_total_dolares = round(saldo_total / precio_dolar,2)        
        rubros = Rubros.objects.filter(habilitado=1).order_by('-cantidad_usos')
        page_obj = funcion_paginas_varias(req,data)

        last_records = CuotasMoto.objects.values('venta_id').annotate(last_id=Max('id'))
        pagos_motos = CuotasMoto.objects.filter(id__in=[record['last_id'] for record in last_records]).order_by('venta_id')
        creditos_a_cobrar_x_motos_pesos = 0
        creditos_a_cobrar_x_motos_dolares = 0
        creditos_a_cobrar_x_motos_pesos_dia = 0
        creditos_a_cobrar_x_motos_dolares_dia = 0
        creditos_a_cobrar_x_motos_pesos_mes = 0
        creditos_a_cobrar_x_motos_dolares_mes = 0
        for pago in pagos_motos:
            if pago.cant_restante_pesos != 0 or pago.cant_restante_dolares != 0:
                creditos_a_cobrar_x_motos_pesos = creditos_a_cobrar_x_motos_pesos + float(pago.cant_restante_pesos)
                creditos_a_cobrar_x_motos_dolares = creditos_a_cobrar_x_motos_dolares + float(pago.cant_restante_dolares)
                if pago.fecha_prox_pago ==  fecha_actual:
                    creditos_a_cobrar_x_motos_pesos_dia = creditos_a_cobrar_x_motos_pesos_dia + float(pago.cant_restante_pesos)
                    creditos_a_cobrar_x_motos_dolares_dia = creditos_a_cobrar_x_motos_dolares_dia + float(pago.cant_restante_dolares)

                if pago.fecha_prox_pago:
                    if pago.fecha_prox_pago.month == mes_actual:
                        creditos_a_cobrar_x_motos_pesos_mes = creditos_a_cobrar_x_motos_pesos_mes + float(pago.cant_restante_pesos)
                        creditos_a_cobrar_x_motos_dolares_mes = creditos_a_cobrar_x_motos_dolares_mes + float(pago.cant_restante_dolares)




        last_records_accesorios = CuotasAccesorios.objects.values('venta_id').annotate(last_id=Max('id'))
        pagos_accesorios = CuotasAccesorios.objects.filter(id__in=[record['last_id'] for record in last_records_accesorios]).order_by('venta_id')
        creditos_a_cobrar_x_accesorios_pesos = 0
        creditos_a_cobrar_x_accesorios_dolares = 0
        creditos_a_cobrar_x_accesorios_pesos_dia = 0
        creditos_a_cobrar_x_accesorios_dolares_dia = 0
        creditos_a_cobrar_x_accesorios_pesos_mes = 0
        creditos_a_cobrar_x_accesorios_dolares_mes = 0
        for pago in pagos_accesorios:
            if pago.cant_restante_pesos != 0 or pago.cant_restante_dolares != 0:
                creditos_a_cobrar_x_accesorios_pesos = creditos_a_cobrar_x_accesorios_pesos + float(pago.cant_restante_pesos)
                creditos_a_cobrar_x_accesorios_dolares = creditos_a_cobrar_x_accesorios_dolares + float(pago.cant_restante_dolares)
                if pago.fecha_pago ==  fecha_actual:
                    creditos_a_cobrar_x_accesorios_pesos_dia = creditos_a_cobrar_x_accesorios_pesos_dia + float(pago.cant_restante_pesos)
                    creditos_a_cobrar_x_accesorios_dolares_dia = creditos_a_cobrar_x_accesorios_dolares_dia + float(pago.cant_restante_dolares)

                if pago.fecha_pago.month == mes_actual:
                    creditos_a_cobrar_x_accesorios_pesos_mes = creditos_a_cobrar_x_accesorios_pesos_mes + float(pago.cant_restante_pesos)
                    creditos_a_cobrar_x_accesorios_dolares_mes = creditos_a_cobrar_x_accesorios_dolares_mes + float(pago.cant_restante_dolares)
        
        total_creditos_pesos = round(creditos_a_cobrar_x_motos_pesos + creditos_a_cobrar_x_accesorios_pesos,2)
        total_creditos_dolares = round(creditos_a_cobrar_x_motos_dolares + creditos_a_cobrar_x_accesorios_dolares,2) 
        
        total_creditos_pesos_dia = round(creditos_a_cobrar_x_motos_pesos_dia + creditos_a_cobrar_x_accesorios_pesos_dia,2)
        total_creditos_dolares_dia = round(creditos_a_cobrar_x_motos_dolares_dia + creditos_a_cobrar_x_accesorios_dolares_dia,2)
        total_creditos_pesos_mes = round(creditos_a_cobrar_x_motos_pesos_mes + creditos_a_cobrar_x_accesorios_pesos_mes,2)
        total_creditos_dolares_mes = round(creditos_a_cobrar_x_motos_dolares_mes + creditos_a_cobrar_x_accesorios_dolares_mes,2)
            


        return render(req, "perfil_administrativo/balance/balance.html", {"page_obj": page_obj,
                                                                              "creditos_a_cobrar_x_motos_pesos":round(creditos_a_cobrar_x_motos_pesos,2),
                                                                              "creditos_a_cobrar_x_motos_dolares":round(creditos_a_cobrar_x_motos_dolares,2),
                                                                              "creditos_a_cobrar_x_accesorios_pesos":round(creditos_a_cobrar_x_accesorios_pesos,2),
                                                                              "creditos_a_cobrar_x_accesorios_dolares":round(creditos_a_cobrar_x_accesorios_dolares,2),
                                                                              "total_creditos_pesos":total_creditos_pesos,
                                                                              "total_creditos_dolares":total_creditos_dolares,
                                                                              "total_creditos_pesos_dia":total_creditos_pesos_dia,
                                                                              "total_creditos_dolares_dia":total_creditos_dolares_dia,
                                                                              "total_creditos_pesos_mes":total_creditos_pesos_mes,
                                                                              "total_creditos_dolares_mes":total_creditos_dolares_mes,
                                                                              "saldo_total":saldo_total,
                                                                              "saldo_total_dolares":saldo_total_dolares,
                                                                              "ganancias_mes":ganancias_mes,
                                                                              "ganancias_dia":ganancias_dia,
                                                                              "egresos_dia":egresos_dia,
                                                                              "egresos_mes":egresos_mes,
                                                                              "ganancias_netas_mes":ganancias_netas_mes,
                                                                              "ganancias_netas_dia":ganancias_netas_dia,
                                                                              "ingresos_extra_dia":ingresos_extras_dia,
                                                                              "ingresos_extra_mes":ingresos_extras_mes,

                                                                              "ganancias_mes_dolares":round(ganancias_mes / precio_dolar,2),
                                                                              "ganancias_dia_dolares":round(ganancias_dia / precio_dolar,2),
                                                                              "egresos_dia_dolares":round(egresos_dia / precio_dolar,2),
                                                                              "egresos_mes_dolares":round(egresos_mes / precio_dolar,2),
                                                                              "ganancias_netas_mes_dolares":round(ganancias_netas_mes / precio_dolar,2),
                                                                              "ganancias_netas_dia_dolares":round(ganancias_netas_dia / precio_dolar,2),
                                                                              "ingresos_extra_dia_dolares":round(ingresos_extras_dia / precio_dolar,2),
                                                                              "ingresos_extra_mes_dolares":round(ingresos_extras_mes / precio_dolar,2),
                                                                              
                                                                              
                                                                              
                                                                            #   "id_caja":id_caja,
                                                                              "rubros":rubros,
                                                                              "total_ingresos_efectivo_pesos":total_ingresos_efectivo_pesos,
                                                                              "total_ingresos_efectivo_dolares":total_ingresos_efectivo_dolares,
                                                                              "total_ingresos_transferencia_pesos":total_ingresos_transferencia_pesos,
                                                                              "total_ingresos_transferencia_dolares":total_ingresos_transferencia_dolares,
                                                                              "total_ingresos_debito_pesos":total_ingresos_debito_pesos,
                                                                              "total_ingresos_debito_dolares":total_ingresos_debito_dolares,

                                                                              "total_egresos_efectivo_pesos":total_egresos_efectivo_pesos,
                                                                              "total_egresos_efectivo_dolares":total_egresos_efectivo_dolares,
                                                                              "total_egresos_transferencia_pesos":total_egresos_transferencia_pesos,
                                                                              "total_egresos_transferencia_dolares":total_egresos_transferencia_dolares,
                                                                              "total_egresos_debito_pesos":total_egresos_debito_pesos,
                                                                              "total_egresos_debito_dolares":total_egresos_debito_dolares,
                                                                              "saldo_caja":"" if None else "",
                                                                              "saldo_sistema":"" if None else "",
                                                                            #   "monto_inicial":monto_inicial
                                                                              
                                                                              
                                                                              })
    # except Exception as e:
    #     return render(req,"perfil_administrativo/arqueos/movimientos.html",{"error_message":e})

def baja_movimiento_movimiento(req,id_movimiento):
    # try:
        if req.method == "POST":

            movimiento = Movimientos.objects.get(id=id_movimiento)
            id_caja = movimiento.caja_id
            caja = Caja.objects.get(id=id_caja)
            if movimiento.moneda == "Pesos":
                monto_a_quitar = float(movimiento.monto)
            else:
                dolar = PrecioDolar.objects.get(id=1)
                precio_dolar = float(dolar.precio_dolar_tienda)
                monto_a_quitar = float(movimiento.monto) * precio_dolar
            
            depositos_caja = float(caja.depositos)
            egresos_caja = float(caja.egresos)
            
            if movimiento.tipo == "Ingreso" or movimiento.tipo == "Ingreso extra":
                sumar_quitar_depositos = depositos_caja - monto_a_quitar
                caja.depositos = abs(sumar_quitar_depositos)
                # quitar_egresos = egresos_caja
            else:
                # sumar_quitar_depositos = depositos_caja + monto_a_quitar
                quitar_egresos = egresos_caja - monto_a_quitar
                caja.egresos = quitar_egresos
            
            if int(caja.diferencia) != 0:
                if movimiento.tipo == "Ingreso" or movimiento.tipo == "Ingreso extra":
                    nueva_diferencia = float(caja.diferencia) - monto_a_quitar
                else:
                    nueva_diferencia = float(caja.diferencia) + monto_a_quitar
                
                caja.diferencia = nueva_diferencia
            
            
            caja.save() 
            movimiento.delete()
            # messages.success(req, "Movimiento borrado con éxito")
            # return redirect(f"{reverse('MovimientosCaja',kwargs={'id_caja':id_caja})}")
            return render(req, "perfil_administrativo/balance/baja_movimiento.html", {"id_caja":id_caja,"message":"Movimiento borrado con éxito."})
        else:
            return render(req, "perfil_administrativo/balance/baja_movimiento.html", {})
    # except Exception as e:
    #     pass

def buscar_detalles_balance_movimientos_x_fecha(req):
    # try:
        # fecha = req.POST['']
       
        f_detalle_str = req.POST['f_detalles']  # Cambiado a paréntesis
        fecha = datetime.strptime(f_detalle_str, '%Y-%m-%d').date() if f_detalle_str else None
        print(fecha)
        print(f_detalle_str)
        dolar = PrecioDolar.objects.get(id=1)
        precio_dolar = float(dolar.precio_dolar_tienda)

        #CANTIDAD MOTOS VENDIDAS EN LA FECHA INGRESADA, TOTAL DEL INGRESO EN PESOS Y DOLARES 
        movs_motos_vendidas_dia = Movimientos.objects.filter(fecha__date=fecha, es_moto=1)
        cantidad_vendidas_dia = ComprasVentas.objects.filter(fecha_compra=fecha,tipo="V").count()
        total_moto_dia_pesos = 0
        total_moto_dia_dolares = 0
        for movimiento in movs_motos_vendidas_dia:
            # moto = Moto.objects.get(id=movimiento.moto_id)
            if movimiento.tipo == "Ingreso":
                if movimiento.moneda == "Pesos":
                    total_moto_dia_pesos = total_moto_dia_pesos + float(movimiento.monto)
                    total_moto_dia_dolares = total_moto_dia_pesos / precio_dolar 
                elif movimiento.moneda == "Dolares":
                    total_moto_dia_dolares = total_moto_dia_dolares + float(movimiento.monto)
                    total_moto_dia_pesos = total_moto_dia_dolares * precio_dolar


        #CANTIDAD ACCESORIOS VENDIDOS EN LA FECHA INGRESADA, TOTAL DEL INGRESO EN PESOS Y DOLARES 
        movs_accesorios_vendidos_dia = Movimientos.objects.filter(fecha__date=fecha, es_accesorio=1)
        cantidad_vendidos_dia = ClienteAccesorio.objects.filter(fecha_compra=fecha).count()
        total_accesorio_dia_pesos = 0
        total_accesorio_dia_dolares = 0
        for movimiento in movs_accesorios_vendidos_dia:
            if movimiento.tipo == "Ingreso":
                if movimiento.moneda == "Pesos":
                    total_accesorio_dia_pesos = total_accesorio_dia_pesos + float(movimiento.monto)
                    total_accesorio_dia_dolares = total_accesorio_dia_pesos / precio_dolar 
                elif movimiento.moneda == "Dolares":
                    total_accesorio_dia_dolares = total_accesorio_dia_dolares + float(movimiento.monto)
                    total_accesorio_dia_pesos = total_accesorio_dia_dolares * precio_dolar


        movimientos_dia = Movimientos.objects.filter(fecha__date=fecha)
        ingresos_extra = 0
        egresos = 0 
        ingresos_extra_dolares = 0
        egresos_dolares = 0 
        data_egresos = []
        data_ingresos_extra = []
        for mov in movimientos_dia:
            if mov.tipo == "Ingreso extra":
                if mov.moneda == "Pesos":
                    ingresos_extra = ingresos_extra + float(mov.monto)
                    ingresos_extra_dolares = ingresos_extra / precio_dolar
                else:
                    ingresos_extra_dolares = ingresos_extra_dolares + float(mov.monto)
                    ingresos_extra = ingresos_extra_dolares * precio_dolar
                data_ingresos_extra.append({
                    "valor":mov.monto,
                    "motivo":mov.movimiento
                })
            elif mov.tipo == "Egreso":
                if mov.moneda == "Pesos":
                    egresos = egresos + float(mov.monto)
                    egresos_dolares = egresos / precio_dolar
                else:
                    egresos_dolares = egresos_dolares + float(mov.monto)
                    egresos = egresos_dolares * precio_dolar
                data_egresos.append({
                    "rubro":mov.rubro,
                    "valor":mov.monto,
                    "motivo":mov.movimiento
                })
            
        
        total_moto_dia_pesos = round(total_moto_dia_pesos, 2)
        total_moto_dia_dolares = round(total_moto_dia_dolares, 2)
        total_accesorio_dia_pesos = round(total_accesorio_dia_pesos, 2)
        total_accesorio_dia_dolares = round(total_accesorio_dia_dolares, 2)
        subtotal_pesos = total_moto_dia_pesos + total_accesorio_dia_pesos
        subtotal_pesos = round(subtotal_pesos, 2)
        subtotal_dolares = total_moto_dia_dolares + total_accesorio_dia_dolares
        subtotal_dolares = round(subtotal_dolares, 2)
        ingresos_extra = round(ingresos_extra, 2)
        egresos = round(egresos, 2)
        ingresos_extra_dolares = round(ingresos_extra_dolares, 2)
        egresos_dolares = round(egresos_dolares, 2)
        total_pesos = subtotal_pesos + ingresos_extra - egresos
        total_pesos = int(total_pesos)
        total_dolares = subtotal_dolares + ingresos_extra_dolares - egresos_dolares
        total_dolares = int(total_dolares)
        # fecha_json = datetime.strptime(f_detalle_str, '%d-%m-%Y').date() if f_detalle_str else None
        data = [
        {"tipo": "Motos", 
         "venta": cantidad_vendidas_dia, 
         "total_pesos": total_moto_dia_pesos, 
         "total_dolares": total_moto_dia_dolares},
        {"tipo": "Accesorios", 
         "venta": cantidad_vendidos_dia, 
         "total_pesos": total_accesorio_dia_pesos, 
         "total_dolares": total_accesorio_dia_dolares},
         {"egresos":egresos,
          "ingresos_extra":ingresos_extra,
          "egresos_dolares":egresos_dolares,
          "ingresos_extra_dolares":ingresos_extra_dolares,
            "subtotal_pesos": subtotal_pesos,
          "subtotal_dolares":subtotal_dolares,
          "total_general_pesos":total_pesos,
          "total_general_dolares":total_dolares},
          {"egresos_detalle": data_egresos},  # Pasamos la lista completa
            {"ingresos_extra_detalle": data_ingresos_extra},
            {"fecha":f_detalle_str}

        ]
        print("TOTAL ACCESORIOS DEL DIA EN PESOS: $" + str(total_accesorio_dia_pesos))
        print("TOTAL ACCESORIOS DEL DIA EN DOLARES: U$s" + str(total_accesorio_dia_dolares))
        print("CANTIDAD ACCESORIOS DE MOTOS VENDIDAS: " + str(cantidad_vendidos_dia))
        print("TOTAL EN PESOS: " + str(total_pesos))
        print("TOTAL EN DOLARES: " + str(total_dolares))
        print("EGRESOS: " + str(egresos))
        print("INGRESOS EXTRA: " + str(ingresos_extra))
        
        
        return JsonResponse(data, safe=False)
        
        # motos_vendidas_mes = ComprasVentas.objects.filter(fecha_compra__month=mes_actual,tipo="V")
        # cantidad_vendidas_mes = 0
        # for venta in motos_vendidas_mes:
        #     pass
    # except Exception as e:
    #     pass

def buscar_detalles_balance_movimientos_x_fecha_excel(req):
       
        f_detalle_str = req.POST['f_detalles']  # Cambiado a paréntesis
        fecha = datetime.strptime(f_detalle_str, '%Y-%m-%d').date() if f_detalle_str else None
        dolar = PrecioDolar.objects.get(id=1)
        precio_dolar = float(dolar.precio_dolar_tienda)
        #CANTIDAD MOTOS VENDIDAS EN LA FECHA INGRESADA, TOTAL DEL INGRESO EN PESOS Y DOLARES 
        movs_motos_vendidas_dia = Movimientos.objects.filter(fecha__date=fecha, es_moto=1)
        cantidad_vendidas_dia = ComprasVentas.objects.filter(fecha_compra=fecha,tipo="V").count()
        total_moto_dia_pesos = 0
        total_moto_dia_dolares = 0
        for movimiento in movs_motos_vendidas_dia:
            # moto = Moto.objects.get(id=movimiento.moto_id)
            if movimiento.tipo == "Ingreso":
                if movimiento.moneda == "Pesos":
                    total_moto_dia_pesos = total_moto_dia_pesos + float(movimiento.monto)
                    total_moto_dia_dolares = total_moto_dia_pesos / precio_dolar 
                elif movimiento.moneda == "Dolares":
                    total_moto_dia_dolares = total_moto_dia_dolares + float(movimiento.monto)
                    total_moto_dia_pesos = total_moto_dia_dolares * precio_dolar


        #CANTIDAD ACCESORIOS VENDIDOS EN LA FECHA INGRESADA, TOTAL DEL INGRESO EN PESOS Y DOLARES 
        movs_accesorios_vendidos_dia = Movimientos.objects.filter(fecha__date=fecha, es_accesorio=1)
        cantidad_vendidos_dia = ClienteAccesorio.objects.filter(fecha_compra=fecha).count()
        total_accesorio_dia_pesos = 0
        total_accesorio_dia_dolares = 0
        for movimiento in movs_accesorios_vendidos_dia:
            if movimiento.tipo == "Ingreso":
                if movimiento.moneda == "Pesos":
                    total_accesorio_dia_pesos = total_accesorio_dia_pesos + float(movimiento.monto)
                    total_accesorio_dia_dolares = total_accesorio_dia_pesos / precio_dolar 
                elif movimiento.moneda == "Dolares":
                    total_accesorio_dia_dolares = total_accesorio_dia_dolares + float(movimiento.monto)
                    total_accesorio_dia_pesos = total_accesorio_dia_dolares * precio_dolar


        movimientos_dia = Movimientos.objects.filter(fecha__date=fecha)
        ingresos_extra = 0
        egresos = 0 
        ingresos_extra_dolares = 0
        egresos_dolares = 0 
        data_egresos = []
        data_ingresos_extra = []
        for mov in movimientos_dia:
            if mov.tipo == "Ingreso extra":
                if mov.moneda == "Pesos":
                    ingresos_extra = ingresos_extra + float(mov.monto)
                    ingresos_extra_dolares = ingresos_extra / precio_dolar
                else:
                    ingresos_extra_dolares = ingresos_extra_dolares + float(mov.monto)
                    ingresos_extra = ingresos_extra_dolares * precio_dolar
                data_ingresos_extra.append({
                    "valor":mov.monto,
                    "motivo":mov.movimiento
                })
            elif mov.tipo == "Egreso":
                if mov.moneda == "Pesos":
                    egresos = egresos + float(mov.monto)
                    egresos_dolares = egresos / precio_dolar
                else:
                    egresos_dolares = egresos_dolares + float(mov.monto)
                    egresos = egresos_dolares * precio_dolar
                data_egresos.append({
                    "rubro":mov.rubro,
                    "valor":mov.monto,
                    "motivo":mov.movimiento
                })
            
        
        total_moto_dia_pesos = round(total_moto_dia_pesos, 2)
        total_moto_dia_dolares = round(total_moto_dia_dolares, 2)
        total_accesorio_dia_pesos = round(total_accesorio_dia_pesos, 2)
        total_accesorio_dia_dolares = round(total_accesorio_dia_dolares, 2)
        subtotal_pesos = total_moto_dia_pesos + total_accesorio_dia_pesos
        subtotal_pesos = round(subtotal_pesos, 2)
        subtotal_dolares = total_moto_dia_dolares + total_accesorio_dia_dolares
        subtotal_dolares = round(subtotal_dolares, 2)
        ingresos_extra = round(ingresos_extra, 2)
        egresos = round(egresos, 2)
        ingresos_extra_dolares = round(ingresos_extra_dolares, 2)
        egresos_dolares = round(egresos_dolares, 2)
        total_pesos = subtotal_pesos + ingresos_extra - egresos
        total_pesos = int(total_pesos)
        total_dolares = subtotal_dolares + ingresos_extra_dolares - egresos_dolares
        total_dolares = int(total_dolares)


        data = [
            {"tipo": "Motos", "venta": cantidad_vendidas_dia, "total_pesos": total_moto_dia_pesos, "total_dolares": total_moto_dia_dolares},
            {"tipo": "Accesorios", "venta": cantidad_vendidos_dia, "total_pesos": total_accesorio_dia_pesos, "total_dolares": total_accesorio_dia_dolares},
            # Datos adicionales como egresos, ingresos extra, totales y subtotales
            {"egresos":egresos,
          "ingresos_extra":ingresos_extra,
          "egresos_dolares":egresos_dolares,
          "ingresos_extra_dolares":ingresos_extra_dolares,
            "subtotal_pesos": subtotal_pesos,
          "subtotal_dolares":subtotal_dolares,
          "total_general_pesos":total_pesos,
          "total_general_dolares":total_dolares},
            {"egresos_detalle": data_egresos},  # Detalles de los egresos
            {"ingresos_extra_detalle": data_ingresos_extra},  # Detalles de los ingresos extra
            {"fecha": f_detalle_str}  # Fecha del reporte
        ]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Movimientos del Día"

        # Estilo para los encabezados
        header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")  # Amarillo
        header_font = Font(bold=True, color="000000")

        # Encabezados de la tabla de ventas
        headers = ["Tipo", "Cantidad Vendida", "Total Pesos", "Total Dólares"]
        ws.append(headers)

        # Aplicar estilo al encabezado
        for col_num, cell in enumerate(ws[1], 1):
            cell.fill = header_fill
            cell.font = header_font

        # Rellenar las filas con los datos de ventas
        for item in data:
            if "tipo" in item and "venta" in item and "total_pesos" in item and "total_dolares" in item:
                ws.append([item['tipo'], item['venta'], item['total_pesos'], item['total_dolares']])

        # Estilos para las celdas de la tabla de ventas
        for row in ws.iter_rows(min_row=2, max_row=len(data)+1, min_col=1, max_col=4):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if cell.row % 2 == 0:
                    cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Gris claro

        # Agregar los subtotales y totales generales debajo de la tabla de ventas

        ws.append([])  # Fila vacía para separar las secciones
        ws.append(["Datos en Pesos"])
        # ws.append([])
        ws.append(["Ingresos extra","Egresos","Subtotal", "Total"])
        # for item in data:
        # if "ingresos_extra" in item and "egresos" in item and "subtotal_pesos" in item and "total_pesos" in item:
        ws.append([ingresos_extra, egresos,subtotal_pesos,total_pesos])
                # ws.append([item['ingresos_extra'], item['egresos'],item['subtotal_pesos'],item['total_pesos']])

        ws.append([])  # Fila vacía para separar las secciones
        ws.append(["Datos en Dolares"])
        # ws.append([])
        ws.append(["Ingresos extra","Egresos","Subtotal", "Total"])
        # for item in data:
        # if "ingresos_extra_dolares" in item and "egresos_dolares" in item and "subtotal_dolares" in item and "total_dolares" in item:
        ws.append([ingresos_extra_dolares, egresos_dolares,subtotal_dolares,total_dolares])
                # ws.append(item['ingresos_extra_dolares'], item['egresos_dolares'],item['subtotal_dolares'],item['total_dolares']

        # Detalles de los egresos
        ws.append([])  # Fila vacía para separar las secciones
        ws.append(["Egresos Detalle"])
        # ws.append([])
        ws.append(["Monto","Motivo","Rubro"])
        for egreso in data_egresos:
            ws.append([egreso['valor'], egreso['motivo'],egreso['rubro']])

        # Detalles de los ingresos extra
        ws.append([])  # Fila vacía para separar las secciones
        ws.append(["Ingresos Extra Detalles"])
        # ws.append([])
        ws.append(["Monto","Motivo"])
        for ingreso in data_ingresos_extra:
            ws.append([ingreso['valor'], ingreso['motivo']])

        # Configuración de la respuesta para la descarga del archivo
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=Movimientos.xlsx'
        wb.save(response)
        return response

def buscar_detalles_balance_movimientos_x_mes_anio(req):
    # try:
        mes = int(req.POST['mes_reporte'])
        anio = int(req.POST['anio_reporte'])
        dolar = PrecioDolar.objects.get(id=1)
        precio_dolar = float(dolar.precio_dolar_tienda)
        #CANTIDAD MOTOS VENDIDAS EN LA FECHA INGRESADA, TOTAL DEL INGRESO EN PESOS Y DOLARES 
        movs_motos_vendidas_dia = Movimientos.objects.filter(fecha__month=mes,fecha__year=anio, es_moto=1)
        cantidad_vendidas_dia = ComprasVentas.objects.filter(fecha_compra__month=mes,fecha_compra__year=anio,tipo="V").count()
        total_moto_dia_pesos = 0
        total_moto_dia_dolares = 0
        for movimiento in movs_motos_vendidas_dia:
            # moto = Moto.objects.get(id=movimiento.moto_id)
            if movimiento.tipo == "Ingreso":
                if movimiento.moneda == "Pesos":
                    total_moto_dia_pesos = total_moto_dia_pesos + float(movimiento.monto)
                    total_moto_dia_dolares = total_moto_dia_pesos / precio_dolar 
                elif movimiento.moneda == "Dolares":
                    total_moto_dia_dolares = total_moto_dia_dolares + float(movimiento.monto)
                    total_moto_dia_pesos = total_moto_dia_dolares * precio_dolar


        #CANTIDAD ACCESORIOS VENDIDOS EN LA FECHA INGRESADA, TOTAL DEL INGRESO EN PESOS Y DOLARES 
        movs_accesorios_vendidos_dia = Movimientos.objects.filter(fecha__month=mes,fecha__year=anio, es_accesorio=1)
        cantidad_vendidos_dia = ClienteAccesorio.objects.filter(fecha_compra__month=mes,fecha_compra__year=anio).count()
        total_accesorio_dia_pesos = 0
        total_accesorio_dia_dolares = 0
        for movimiento in movs_accesorios_vendidos_dia:
            if movimiento.tipo == "Ingreso":
                if movimiento.moneda == "Pesos":
                    total_accesorio_dia_pesos = total_accesorio_dia_pesos + float(movimiento.monto)
                    total_accesorio_dia_dolares = total_accesorio_dia_pesos / precio_dolar 
                elif movimiento.moneda == "Dolares":
                    total_accesorio_dia_dolares = total_accesorio_dia_dolares + float(movimiento.monto)
                    total_accesorio_dia_pesos = total_accesorio_dia_dolares * precio_dolar


        movimientos_dia = Movimientos.objects.filter(fecha__month=mes,fecha__year=anio)
        ingresos_extra = 0
        egresos = 0 
        ingresos_extra_dolares = 0
        egresos_dolares = 0 
        data_egresos = []
        data_ingresos_extra = []
        for mov in movimientos_dia:
            if mov.tipo == "Ingreso extra":
                if mov.moneda == "Pesos":
                    ingresos_extra = ingresos_extra + float(mov.monto)
                    ingresos_extra_dolares = ingresos_extra / precio_dolar
                else:
                    ingresos_extra_dolares = ingresos_extra_dolares + float(mov.monto)
                    ingresos_extra = ingresos_extra_dolares * precio_dolar
                data_ingresos_extra.append({
                    "valor":mov.monto,
                    "motivo":mov.movimiento
                })
            elif mov.tipo == "Egreso":
                if mov.moneda == "Pesos":
                    egresos = egresos + float(mov.monto)
                    egresos_dolares = egresos / precio_dolar
                else:
                    egresos_dolares = egresos_dolares + float(mov.monto)
                    egresos = egresos_dolares * precio_dolar
                data_egresos.append({
                    "rubro":mov.rubro,
                    "valor":mov.monto,
                    "motivo":mov.movimiento
                })
            
        meses = ["","Enero","Febrero","Marzo","Abril","Mayo","Junio","Julio","Agosto","Setiembre","Octubre","Noviembre","Diciembre"]
        total_moto_dia_pesos = round(total_moto_dia_pesos, 2)
        total_moto_dia_dolares = round(total_moto_dia_dolares, 2)
        total_accesorio_dia_pesos = round(total_accesorio_dia_pesos, 2)
        total_accesorio_dia_dolares = round(total_accesorio_dia_dolares, 2)
        subtotal_pesos = total_moto_dia_pesos + total_accesorio_dia_pesos
        subtotal_pesos = round(subtotal_pesos, 2)
        subtotal_dolares = total_moto_dia_dolares + total_accesorio_dia_dolares
        subtotal_dolares = round(subtotal_dolares, 2)
        ingresos_extra = round(ingresos_extra, 2)
        egresos = round(egresos, 2)
        ingresos_extra_dolares = round(ingresos_extra_dolares, 2)
        egresos_dolares = round(egresos_dolares, 2)
        total_pesos = subtotal_pesos + ingresos_extra - egresos
        total_pesos = int(total_pesos)
        total_dolares = subtotal_dolares + ingresos_extra_dolares - egresos_dolares
        total_dolares = int(total_dolares)

        print("TOTAL MOTOS DEL DIA EN PESOS: $" + str(total_moto_dia_pesos))
        print("TOTAL MOTOS DEL DIA EN DOLARES: U$s" + str(total_moto_dia_dolares))
        print("CANTIDAD MOTOS VENDIDOS: " + str(cantidad_vendidas_dia))
        print("TOTAL ACCESORIOS DEL DIA EN PESOS: $" + str(total_accesorio_dia_pesos))
        print("TOTAL ACCESORIOS DEL DIA EN DOLARES: U$s" + str(total_accesorio_dia_dolares))
        print("CANTIDAD ACCESORIOS VENDIDOS: " + str(cantidad_vendidos_dia))
        print("TOTAL EN PESOS: " + str(total_pesos))
        print("TOTAL EN DOLARES: " + str(total_dolares))
        print("EGRESOS: " + str(egresos))
        print("INGRESOS EXTRA: " + str(ingresos_extra))
        # fecha_json = datetime.strptime(f_detalle_str, '%d-%m-%Y').date() if f_detalle_str else None
        data = [
        {"tipo": "Motos", 
         "venta": cantidad_vendidas_dia, 
         "total_pesos": total_moto_dia_pesos, 
         "total_dolares": total_moto_dia_dolares},
        {"tipo": "Accesorios", 
         "venta": cantidad_vendidos_dia, 
         "total_pesos": total_accesorio_dia_pesos, 
         "total_dolares": total_accesorio_dia_dolares},
         {"egresos":egresos,
          "ingresos_extra":ingresos_extra,
          "egresos_dolares":egresos_dolares,
          "ingresos_extra_dolares":ingresos_extra_dolares,
            "subtotal_pesos": subtotal_pesos,
          "subtotal_dolares":subtotal_dolares,
          "total_general_pesos":total_pesos,
          "total_general_dolares":total_dolares},
          {"egresos_detalle": data_egresos},  # Pasamos la lista completa
            {"ingresos_extra_detalle": data_ingresos_extra},
            {"mes":meses[mes],
             "anio":anio}

        ]
        
        return JsonResponse(data, safe=False)

def buscar_detalles_balance_movimientos_x_mes_anio_excel(req):
        mes = int(req.POST['mes_reporte'])
        anio = int(req.POST['anio_reporte'])
        dolar = PrecioDolar.objects.get(id=1)
        precio_dolar = float(dolar.precio_dolar_tienda)
        
        #CANTIDAD MOTOS VENDIDAS EN LA FECHA INGRESADA, TOTAL DEL INGRESO EN PESOS Y DOLARES 
        movs_motos_vendidas_dia = Movimientos.objects.filter(fecha__month=mes,fecha__year=anio, es_moto=1)
        cantidad_vendidas_dia = ComprasVentas.objects.filter(fecha_compra__month=mes,fecha_compra__year=anio,tipo="V").count()
        total_moto_dia_pesos = 0
        total_moto_dia_dolares = 0
        for movimiento in movs_motos_vendidas_dia:
            # moto = Moto.objects.get(id=movimiento.moto_id)
            if movimiento.tipo == "Ingreso":
                if movimiento.moneda == "Pesos":
                    total_moto_dia_pesos = total_moto_dia_pesos + float(movimiento.monto)
                    total_moto_dia_dolares = total_moto_dia_pesos / precio_dolar 
                elif movimiento.moneda == "Dolares":
                    total_moto_dia_dolares = total_moto_dia_dolares + float(movimiento.monto)
                    total_moto_dia_pesos = total_moto_dia_dolares * precio_dolar


        #CANTIDAD ACCESORIOS VENDIDOS EN LA FECHA INGRESADA, TOTAL DEL INGRESO EN PESOS Y DOLARES 
        movs_accesorios_vendidos_dia = Movimientos.objects.filter(fecha__month=mes,fecha__year=anio, es_accesorio=1)
        cantidad_vendidos_dia = ClienteAccesorio.objects.filter(fecha_compra__month=mes,fecha_compra__year=anio).count()
        total_accesorio_dia_pesos = 0
        total_accesorio_dia_dolares = 0
        for movimiento in movs_accesorios_vendidos_dia:
            if movimiento.tipo == "Ingreso":
                if movimiento.moneda == "Pesos":
                    total_accesorio_dia_pesos = total_accesorio_dia_pesos + float(movimiento.monto)
                    total_accesorio_dia_dolares = total_accesorio_dia_pesos / precio_dolar 
                elif movimiento.moneda == "Dolares":
                    total_accesorio_dia_dolares = total_accesorio_dia_dolares + float(movimiento.monto)
                    total_accesorio_dia_pesos = total_accesorio_dia_dolares * precio_dolar


        movimientos_dia = Movimientos.objects.filter(fecha__month=mes,fecha__year=anio)
        ingresos_extra = 0
        egresos = 0 
        ingresos_extra_dolares = 0
        egresos_dolares = 0 
        data_egresos = []
        data_ingresos_extra = []
        for mov in movimientos_dia:
            if mov.tipo == "Ingreso extra":
                if mov.moneda == "Pesos":
                    ingresos_extra = ingresos_extra + float(mov.monto)
                    ingresos_extra_dolares = ingresos_extra / precio_dolar
                else:
                    ingresos_extra_dolares = ingresos_extra_dolares + float(mov.monto)
                    ingresos_extra = ingresos_extra_dolares * precio_dolar
                data_ingresos_extra.append({
                    "valor":mov.monto,
                    "motivo":mov.movimiento
                })
            elif mov.tipo == "Egreso":
                if mov.moneda == "Pesos":
                    egresos = egresos + float(mov.monto)
                    egresos_dolares = egresos / precio_dolar
                else:
                    egresos_dolares = egresos_dolares + float(mov.monto)
                    egresos = egresos_dolares * precio_dolar
                data_egresos.append({
                    "rubro":mov.rubro,
                    "valor":mov.monto,
                    "motivo":mov.movimiento
                })
            
        
        total_moto_dia_pesos = round(total_moto_dia_pesos, 2)
        total_moto_dia_dolares = round(total_moto_dia_dolares, 2)
        total_accesorio_dia_pesos = round(total_accesorio_dia_pesos, 2)
        total_accesorio_dia_dolares = round(total_accesorio_dia_dolares, 2)
        subtotal_pesos = total_moto_dia_pesos + total_accesorio_dia_pesos
        subtotal_pesos = round(subtotal_pesos, 2)
        subtotal_dolares = total_moto_dia_dolares + total_accesorio_dia_dolares
        subtotal_dolares = round(subtotal_dolares, 2)
        ingresos_extra = round(ingresos_extra, 2)
        egresos = round(egresos, 2)
        ingresos_extra_dolares = round(ingresos_extra_dolares, 2)
        egresos_dolares = round(egresos_dolares, 2)
        total_pesos = subtotal_pesos + ingresos_extra - egresos
        total_pesos = int(total_pesos)
        total_dolares = subtotal_dolares + ingresos_extra_dolares - egresos_dolares
        total_dolares = int(total_dolares)

        print("TOTAL MOTOS DEL DIA EN PESOS: $" + str(total_moto_dia_pesos))
        print("TOTAL MOTOS DEL DIA EN DOLARES: U$s" + str(total_moto_dia_dolares))
        print("CANTIDAD MOTOS VENDIDOS: " + str(cantidad_vendidas_dia))
        print("TOTAL ACCESORIOS DEL DIA EN PESOS: $" + str(total_accesorio_dia_pesos))
        print("TOTAL ACCESORIOS DEL DIA EN DOLARES: U$s" + str(total_accesorio_dia_dolares))
        print("CANTIDAD ACCESORIOS VENDIDOS: " + str(cantidad_vendidos_dia))
        print("TOTAL EN PESOS: " + str(total_pesos))
        print("TOTAL EN DOLARES: " + str(total_dolares))
        print("EGRESOS: " + str(egresos))
        print("INGRESOS EXTRA: " + str(ingresos_extra))


        data = [
            {"tipo": "Motos", "venta": cantidad_vendidas_dia, "total_pesos": total_moto_dia_pesos, "total_dolares": total_moto_dia_dolares},
            {"tipo": "Accesorios", "venta": cantidad_vendidos_dia, "total_pesos": total_accesorio_dia_pesos, "total_dolares": total_accesorio_dia_dolares},
            # Datos adicionales como egresos, ingresos extra, totales y subtotales
            {"egresos":egresos,
          "ingresos_extra":ingresos_extra,
          "egresos_dolares":egresos_dolares,
          "ingresos_extra_dolares":ingresos_extra_dolares,
            "subtotal_pesos": subtotal_pesos,
          "subtotal_dolares":subtotal_dolares,
          "total_general_pesos":total_pesos,
          "total_general_dolares":total_dolares},
            {"egresos_detalle": data_egresos},  # Detalles de los egresos
            {"ingresos_extra_detalle": data_ingresos_extra},  # Detalles de los ingresos extra
              # Fecha del reporte
        ]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Movimientos del Día"

        # Estilo para los encabezados
        header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")  # Amarillo
        header_font = Font(bold=True, color="000000")

        # Encabezados de la tabla de ventas
        headers = ["Tipo", "Cantidad Vendida", "Total Pesos", "Total Dólares"]
        ws.append(headers)

        # Aplicar estilo al encabezado
        for col_num, cell in enumerate(ws[1], 1):
            cell.fill = header_fill
            cell.font = header_font

        # Rellenar las filas con los datos de ventas
        for item in data:
            if "tipo" in item and "venta" in item and "total_pesos" in item and "total_dolares" in item:
                ws.append([item['tipo'], item['venta'], item['total_pesos'], item['total_dolares']])

        # Estilos para las celdas de la tabla de ventas
        for row in ws.iter_rows(min_row=2, max_row=len(data)+1, min_col=1, max_col=4):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if cell.row % 2 == 0:
                    cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Gris claro

        # Agregar los subtotales y totales generales debajo de la tabla de ventas

        ws.append([])  # Fila vacía para separar las secciones
        ws.append(["Datos en Pesos"])
        # ws.append([])
        ws.append(["Ingresos extra","Egresos","Subtotal", "Total"])
        # for item in data:
        # if "ingresos_extra" in item and "egresos" in item and "subtotal_pesos" in item and "total_pesos" in item:
        ws.append([ingresos_extra, egresos,subtotal_pesos,total_pesos])
                # ws.append([item['ingresos_extra'], item['egresos'],item['subtotal_pesos'],item['total_pesos']])

        ws.append([])  # Fila vacía para separar las secciones
        ws.append(["Datos en Dolares"])
        # ws.append([])
        ws.append(["Ingresos extra","Egresos","Subtotal", "Total"])
        # for item in data:
        # if "ingresos_extra_dolares" in item and "egresos_dolares" in item and "subtotal_dolares" in item and "total_dolares" in item:
        ws.append([ingresos_extra_dolares, egresos_dolares,subtotal_dolares,total_dolares])
                # ws.append(item['ingresos_extra_dolares'], item['egresos_dolares'],item['subtotal_dolares'],item['total_dolares']

        # Detalles de los egresos
        ws.append([])  # Fila vacía para separar las secciones
        ws.append(["Egresos Detalle"])
        # ws.append([])
        ws.append(["Monto","Motivo","Rubro"])
        for egreso in data_egresos:
            ws.append([egreso['valor'], egreso['motivo'],egreso['rubro']])

        # Detalles de los ingresos extra
        ws.append([])  # Fila vacía para separar las secciones
        ws.append(["Ingresos Extra Detalles"])
        # ws.append([])
        ws.append(["Monto","Motivo"])
        for ingreso in data_ingresos_extra:
            ws.append([ingreso['valor'], ingreso['motivo']])

        # Configuración de la respuesta para la descarga del archivo
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=Movimientos.xlsx'
        wb.save(response)
        return response

@admin_required
def ingresos_balance(req):
    # try:
        if int(req.POST['ingresos']) <= 0:
            # return render(req,"perfil_administrativo/arqueos/arqueos.html",{"error_message":"Ingrese un valor correcto"})
            messages.error(req,"El monto del ingreso es incorrecto.")
            return redirect('Arqueos')
        else:
            moneda = req.POST['moneda_monto_ingreso']
            if moneda == "Pesos":
                monto = int(req.POST['ingresos'])
            else:
                dolar = PrecioDolar.objects.get(id=1)
                precio_dolar = float(dolar.precio_dolar_tienda)
                monto = float(precio_dolar * float(req.POST['ingresos']))
            caja = Caja.objects.filter(estado="Abierto").first()
            if not caja:
                caja = Caja.objects.all().order_by('-id')[1]
            id_caja = caja.id
            if int(caja.diferencia) != 0:
                sumar_diferencia = float(caja.diferencia) + monto
                caja.diferencia = sumar_diferencia
            ingresos = float(caja.depositos) + monto
            caja.depositos = ingresos
            caja.save()
            usuario = req.user
            personal = Personal.objects.filter(usuario=usuario.username).first()
            moneda = req.POST['moneda_monto_ingreso']
            metodo = req.POST['metodo_monto_ingreso']
            
            insert_movimientos_caja(req.POST['descripcion_ingreso'],"Ingreso extra",req.POST['ingresos'],id_caja,personal.id,moneda,None,metodo,0,0,None,None)
            #insert_movimientos_caja(movimiento_descripcion,tipo,monto,id_caja,id_personal,moneda,rubro,metodo,es_moto,es_accesorio,id_venta,producto)
            messages.success(req, "Ingreso extra ingresado con éxito")
            return redirect(reverse('Balance', kwargs={'busqueda': "todos"}))
    # except Exception as e:
    #     return render(req,"perfil_administrativo/arqueos/arqueos.html",{"error_message":e})

@admin_required
def egresos_balance(req):
    # try:
       
        if int(req.POST['egresos']) < 0 or int(req.POST['egresos']) == 0:
            # return render(req,"perfil_administrativo/arqueos/arqueos.html",{"error_message":"Ingrese un valor correcto"})
            messages.error(req,"El monto del egreso es incorrecto.")
            return redirect(reverse('Balance', kwargs={'busqueda': "todos"}))
        elif not req.POST['rubro_egreso']:  
            messages.error(req,"Debe ingresar un nuevo rubro.")
            return redirect(reverse('Balance', kwargs={'busqueda': "todos"}))
        else:
            id_rubro = req.POST['rubro_egreso']
            moneda = req.POST['moneda_monto_egreso']
            if moneda == "Pesos":
                monto = int(req.POST['egresos'])
            else:
                dolar = PrecioDolar.objects.get(id=1)
                precio_dolar = float(dolar.precio_dolar_tienda)
                monto = float(precio_dolar * float(req.POST['egresos']))
            caja = Caja.objects.filter(estado="Abierto").first()
            if not caja:
                caja = Caja.objects.all().order_by('-id')[1]
            id_caja = caja.id
            if int(caja.diferencia) != 0:
                restar_diferencia = float(caja.diferencia) - monto
                caja.diferencia = restar_diferencia
            egresos = float(caja.egresos) + monto
            caja.egresos = egresos
            caja.save()
            usuario = req.user
            personal = Personal.objects.filter(usuario=usuario.username).first()
            rubro = Rubros.objects.get(id=id_rubro)
            nombre_rubro = rubro.rubro
            rubro.cantidad_usos = int(rubro.cantidad_usos) + 1
            rubro.save()
            moneda = req.POST['moneda_monto_egreso']
            metodo = req.POST['metodo_monto_egreso']
            insert_movimientos_caja(req.POST['descripcion_egreso'],"Egreso",req.POST['egresos'],id_caja,personal.id,moneda,nombre_rubro,metodo,0,0,None,None)
            #insert_movimientos_caja(movimiento_descripcion,tipo,monto,id_caja,id_personal,moneda,rubro,metodo,es_moto,es_accesorio,id_venta,producto)
            messages.success(req, "Egreso extra ingresado con éxito")
            return redirect(reverse('Balance', kwargs={'busqueda': "todos"}))
    # except Exception as e:
    #     return render(req,"perfil_administrativo/arqueos/arqueos.html",{"error_message":e})

def nuevo_rubro_egreso_balance(req):
    try:
        rubro = req.POST['nuevo_rubro']
        nuevo_rubro = Rubros(
            rubro = rubro,
            habilitado = 1
        )
        nuevo_rubro.save()
        messages.success(req, "Rubro ingresado con éxito")
        # return redirect(f"{reverse('Balance')}")
        return redirect(reverse('Balance', kwargs={'busqueda': "todos"}))
    except Exception as e:
        pass

def cambiar_a_perfil_taller(req):
    return render(req,"perfil_taller/padre_perfil_taller.html",{})

def cambiar_a_perfil_tienda(req):
    return render(req,"perfil_administrativo/bienvenida.html",{})

